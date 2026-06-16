#!/usr/bin/env python3
"""
velo_final.py — VeloCloud profile-aware config collector + structured security review
─────────────────────────────────────────────────────────────────────────────────────
Architecture:
  Portal API v1  (/portal/)   → edge list (with profile join), profile catalog
  APIv2          (/api/sdwan/v2/) → profile deviceSettings, edge deviceSettings

Detection modules:
  1. Firewall rule analysis       — firewall module; rule scope, default deny, logging
  2. Edge access / NAT / overrides — services block, overlay→mgmt, DNAT, override governance
  3. Business policy / WAN        — segment rules, QoS override, WAN link encryption
  4. Segmentation analysis        — profile vs edge drift + segment isolation (VPN)
  5. Structural diff              — LAN/WAN drift when deviceSettings override enabled
  6. Device hardening             — NTP, syslog, SNMP, BFD, DNS, HA, overlay encryption
  6b. System certificate auth    — edge authentication mode + PKI certificate health
  6c. HA pair parity             — CPE0/CPE1 config symmetry (NTP, syslog, SNMP, AAA, BFD)
  7. Edge inventory               — offline/stale edges, activation, version review
  8. Enterprise events (optional) — --events: filtered MGD_CONF_* and auth events
  9. Methodology coverage         — XLSX-linked findings, manual checklist, Dradis CSV
 10. Deep collectors (--deep)     — route table, gateway assignments
 11. VPN review                    — encryption strength, certificate validation, key rotation
 12. Enterprise management         — API token inventory, auth mode, dormant users
 13. Isolation review              — shared resources, profile inheritance dedupe, wireless security

Output layout:
  vco_output/<timestamp>/
    portal_edge_list.json
    portal_profile_catalog.json
    edge_profile_join.json
    profiles/<profileLogicalId>.json
    edges/<edgeName>.json
    combined/<edgeName>_combined.json
    enterprise_events.json          (when --events, raw events)
    enterprise_event_findings.json  (when --events, filtered findings)
    wan/<edge>.json
    routes/enterprise_route_table.json   (--deep)
    gateways/gateway_<id>_assignments.json (--deep)
    certs/<edge>.json
    mgmt/enterprise.json
    mgmt/enterprise_users.json
    mgmt/api_tokens.json
    findings/
      findings.json
      findings.csv
      hardening.json
      methodology_coverage.json / .csv
      manual_review_checklist.json
      review_summary_by_scope.json
      dradis_import.csv
      summary.json

Required environment variables:
  VCO_HOSTNAME
  VCO_ENTERPRISE_LOGICAL_ID  (GUID, from enterprise/getEnterprise or script output)
  VCO_ENTERPRISE_NUMERIC_ID  (integer, from Orchestrator URL /enterprises/<id>/)
  VCO_TOKEN                  (APIv2 + Portal JSON-RPC via Authorization: Token header)

Optional:
  VCO_SCOPE_EDGES        comma-separated edge names to limit scope
  VCO_VERIFY_TLS         true/false  (default: false)
  VCO_STALE_EDGE_DAYS    days without contact before stale finding (default: 30)
  VCO_MIN_EDGE_VERSION   minimum acceptable edge software version, e.g. 6.0.0
  VCO_EVENTS_HOURS       event lookback when using --events (default: 168)
  VCO_EVENTS             false/0 to disable --events (enabled by default)
  VCO_PHASE4             false/0 to disable --deep (enabled by default)
  VCO_MINIMAL            1/true to disable both --events and --deep
  VCO_METHODOLOGY_XLSX   path to methodology workbook (default: velocloud_config_review_FINAL.xlsx)
  VCO_DORMANT_USER_DAYS  days since last login before dormant user finding (default: 90)

Usage:
  export VCO_HOSTNAME=vco.example.velocloud.net
  export VCO_TOKEN='YOUR_API_TOKEN_HERE'
  export VCO_ENTERPRISE_LOGICAL_ID='xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx'
  export VCO_ENTERPRISE_NUMERIC_ID='1234'
  export VCO_SCOPE_EDGES='EDGE-01,EDGE-02'   # optional
  python3 velo_final.py                      # full review: config + events + deep
  python3 velo_final.py --no-events          # skip enterprise event fetch
  python3 velo_final.py --no-deep            # skip route table + gateway collectors
  python3 velo_final.py --scope-edges EDGE-01,EDGE-02
  python3 velo_final.py --xlsx velocloud_config_review_FINAL.xlsx
  python3 velo_final.py --output-dir vco_output/manual_run
  python3 velo_final.py --events-hours 72
"""

import argparse
import csv
import ipaddress
import json
import os
import re
import sys
import time
from copy import deepcopy
from datetime import datetime, timedelta, timezone

import requests
import urllib3

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — read from environment
# ─────────────────────────────────────────────────────────────────────────────
VCO_HOSTNAME               = os.getenv("VCO_HOSTNAME")
VCO_TOKEN                  = os.getenv("VCO_TOKEN")
VCO_ENTERPRISE_LOGICAL_ID  = os.getenv("VCO_ENTERPRISE_LOGICAL_ID")
VCO_ENTERPRISE_NUMERIC_ID  = os.getenv("VCO_ENTERPRISE_NUMERIC_ID")
VCO_SCOPE_EDGES            = [x.strip() for x in os.getenv("VCO_SCOPE_EDGES", "").split(",") if x.strip()]
VERIFY_TLS                 = os.getenv("VCO_VERIFY_TLS", "false").strip().lower() in {"1", "true", "yes", "on"}
TIMEOUT                    = 45
PAGE_SIZE                  = 100
RATE_LIMIT_MAX_RETRIES     = 5
RATE_LIMIT_BASE_DELAY      = 2.0
STALE_EDGE_DAYS            = int(os.getenv("VCO_STALE_EDGE_DAYS", "30"))
MIN_EDGE_VERSION           = os.getenv("VCO_MIN_EDGE_VERSION", "").strip()
EVENTS_HOURS_DEFAULT       = int(os.getenv("VCO_EVENTS_HOURS", "168"))
def _feature_enabled(env_name: str, default: bool = True) -> bool:
    """Parse env var as feature toggle; empty means *default*."""
    raw = os.getenv(env_name, "").strip().lower()
    if not raw:
        return default
    if raw in {"0", "false", "no", "off"}:
        return False
    if raw in {"1", "true", "yes", "on"}:
        return True
    return default


_MINIMAL_RUN = os.getenv("VCO_MINIMAL", "").strip().lower() in {"1", "true", "yes", "on"}
EVENTS_DEFAULT = _feature_enabled("VCO_EVENTS", default=not _MINIMAL_RUN)
PHASE4_DEFAULT = _feature_enabled("VCO_PHASE4", default=not _MINIMAL_RUN)
DORMANT_USER_DAYS          = int(os.getenv("VCO_DORMANT_USER_DAYS", "90"))
METHODOLOGY_XLSX           = os.getenv(
    "VCO_METHODOLOGY_XLSX",
    "velocloud_config_review_FINAL.xlsx",
)

_SENSITIVE_NAT_PORTS       = {22, 23, 3389, 445, 1433, 3306, 5432, 5900, 8080, 8443}
_EDGE_ACCESS_SERVICES      = ("ssh", "snmp", "localUi", "console", "icmp", "post")

_SECURITY_EVENT_NAMES = {
    "MGD_CONF_FAILED", "MGD_CONF_ROLLBACK", "MGD_CONF_PENDING",
    "EDGE_LOCALUI_LOGIN", "EDGE_SSH_LOGIN", "EDGE_MGMT_LOGIN",
    "USER_LOGIN_FAILED", "LOGIN_FAILED", "AUTHENTICATION_FAILED",
}

_IGNORED_EVENT_NAMES = {
    "BROWSER_ENTERPRISE_LOGIN",
    "USER_LOGIN",
}

_AGGREGATE_FAILURE_EVENTS = {
    "USER_LOGIN_FAILURE",
    "USER_LOGIN_FAILED",
    "LOGIN_FAILED",
    "AUTHENTICATION_FAILED",
}

_ADMIN_ACTIVITY_EVENTS = {
    "CREATE_API_TOKEN", "DOWNLOAD_API_TOKEN", "REVOKE_API_TOKEN",
    "GRANT_SSH_ACCESS", "REVOKE_SSH_ACCESS", "ROLE_RESET",
    "REMOTE_ACTION", "EDIT_PROFILE",
    "EDGE_LOCALUI_LOGIN", "EDGE_SSH_LOGIN", "EDGE_MGMT_LOGIN",
}

_CONFIG_CHANGE_EVENTS = {
    "MGD_CONF_APPLIED", "MGD_CONF_FAILED", "MGD_CONF_ROLLBACK", "MGD_CONF_PENDING",
    "EDIT_PROFILE", "FIREWALL_ENABLE", "FIREWALL_DISABLE",
}

_EVENT_REVIEW_ALERT_EVENTS = {
    "REMOTE_ACTION", "GRANT_SSH_ACCESS", "REVOKE_SSH_ACCESS",
    "CREATE_API_TOKEN", "DOWNLOAD_API_TOKEN", "REVOKE_API_TOKEN",
    "MGD_CONF_FAILED", "MGD_CONF_ROLLBACK", "MGD_CONF_PENDING",
    "USER_LOGIN_FAILURE", "USER_LOGIN_FAILED", "LOGIN_FAILED",
    "FIREWALL_ENABLE", "FIREWALL_DISABLE",
}

_AGGREGATE_EVENT_NAMES = {
    "USER_LOGIN_FAILURE", "USER_LOGIN_FAILED", "LOGIN_FAILED", "AUTHENTICATION_FAILED",
    "MGD_CONF_APPLIED", "FIREWALL_ENABLE", "EDIT_PROFILE", "REMOTE_ACTION",
    "CREATE_API_TOKEN", "DOWNLOAD_API_TOKEN", "ROLE_RESET",
}

_SEGMENT_MERGE_KEYS = ("syslog", "ntp", "bfd")
_TOP_LEVEL_MERGE_KEYS = ("ntp", "bfd")

_required = {
    "VCO_HOSTNAME": VCO_HOSTNAME,
    "VCO_TOKEN": VCO_TOKEN,
    "VCO_ENTERPRISE_LOGICAL_ID": VCO_ENTERPRISE_LOGICAL_ID,
    "VCO_ENTERPRISE_NUMERIC_ID": VCO_ENTERPRISE_NUMERIC_ID,
}
_missing = [k for k, v in _required.items() if not v]


def _require_env() -> None:
    if _missing:
        print("[ERROR] Missing required environment variables:", ", ".join(_missing))
        sys.exit(1)

API2_BASE   = f"https://{VCO_HOSTNAME}/api/sdwan/v2"
PORTAL_URL  = f"https://{VCO_HOSTNAME}/portal/"
OUTPUT_DIR  = ""  # resolved in main() from --output-dir or vco_output/<timestamp>


def _resolve_output_dir(cli_path: str = None) -> str:
    if cli_path:
        return cli_path
    return os.path.join(
        "vco_output",
        datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S"),
    )

if not VERIFY_TLS:
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────────────────────────────────────
# HTTP CLIENTS (APIv2 + Portal JSON-RPC, both use VCO_TOKEN)
# ─────────────────────────────────────────────────────────────────────────────
api2_client = requests.Session()
api2_client.verify = VERIFY_TLS
api2_client.headers.update({
    "Authorization": f"Token {VCO_TOKEN}",
    "Content-Type":  "application/json",
})

portal_client = requests.Session()
portal_client.verify = VERIFY_TLS
portal_client.headers.update({
    "Authorization": f"Token {VCO_TOKEN}",
    "Content-Type":  "application/json",
})


# ─────────────────────────────────────────────────────────────────────────────
# GENERIC HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def save_json(path: str, data) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)


def save_csv(path: str, rows: list, fieldnames: list) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def safe_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name or "unknown").strip("_") or "unknown"


def get_nested(obj, *keys, default=None):
    cur = obj
    for k in keys:
        if isinstance(cur, dict) and k in cur:
            cur = cur[k]
        else:
            return default
    return cur


def index_by(items: list, key: str) -> dict:
    out = {}
    if not isinstance(items, list):
        return out
    for item in items:
        if isinstance(item, dict) and key in item:
            out[item[key]] = item
    return out


# ─────────────────────────────────────────────────────────────────────────────
# API TRANSPORT
# ─────────────────────────────────────────────────────────────────────────────

def _retry_after_seconds(response, attempt: int) -> float:
    """Parse Retry-After header or fall back to exponential backoff."""
    retry_after = response.headers.get("Retry-After")
    if retry_after:
        try:
            return max(float(retry_after), 1.0)
        except ValueError:
            pass
    return min(RATE_LIMIT_BASE_DELAY * (2 ** attempt), 60.0)


def api2_get(path: str, params: dict = None):
    url = f"{API2_BASE}{path}"
    for attempt in range(RATE_LIMIT_MAX_RETRIES + 1):
        r = api2_client.get(url, params=params, timeout=TIMEOUT)
        if r.status_code == 429:
            if attempt < RATE_LIMIT_MAX_RETRIES:
                delay = _retry_after_seconds(r, attempt)
                print(
                    f"[WARN] Rate limited: {path} — retrying in {delay:.0f}s "
                    f"({attempt + 1}/{RATE_LIMIT_MAX_RETRIES})"
                )
                time.sleep(delay)
                continue
            print(f"[ERROR] Rate limited: {path} — exhausted {RATE_LIMIT_MAX_RETRIES} retries")
            return None
        if r.status_code != 200:
            print(f"[ERROR] APIv2 GET {path} → HTTP {r.status_code}: {r.text[:300]}")
            return None
        try:
            return r.json()
        except Exception:
            print(f"[ERROR] APIv2 GET {path} → non-JSON response")
            return None
    return None


def api2_get_paginated(path: str) -> list:
    results = []
    params = {"limit": PAGE_SIZE}
    while True:
        data = api2_get(path, params=params)
        if not data:
            break
        results.extend(data.get("data", []))
        npl = get_nested(data, "metaData", "nextPageLink")
        if not npl:
            break
        params = {"nextPageLink": npl}
    return results


def portal_rpc(method: str, params: dict, request_id: int = 1):
    body = {"id": request_id, "jsonrpc": "2.0", "method": method, "params": params}
    for attempt in range(RATE_LIMIT_MAX_RETRIES + 1):
        r = portal_client.post(PORTAL_URL, data=json.dumps(body), timeout=TIMEOUT)
        if r.status_code == 429:
            if attempt < RATE_LIMIT_MAX_RETRIES:
                delay = _retry_after_seconds(r, attempt)
                print(
                    f"[WARN] Rate limited: Portal RPC {method} — retrying in {delay:.0f}s "
                    f"({attempt + 1}/{RATE_LIMIT_MAX_RETRIES})"
                )
                time.sleep(delay)
                continue
            print(f"[ERROR] Rate limited: Portal RPC {method} — exhausted {RATE_LIMIT_MAX_RETRIES} retries")
            return None
        if r.status_code != 200:
            print(f"[ERROR] Portal RPC {method} → HTTP {r.status_code}: {r.text[:300]}")
            return None
        try:
            data = r.json()
        except Exception:
            print(f"[ERROR] Portal RPC {method} → non-JSON response")
            return None
        if "error" in data:
            print(f"[ERROR] Portal RPC {method} → JSON-RPC error: {json.dumps(data['error'])}")
            return None
        return data.get("result")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# PORTAL DATA COLLECTION
# ─────────────────────────────────────────────────────────────────────────────

def get_portal_edge_list() -> list:
    params = {
        "enterpriseId": int(VCO_ENTERPRISE_NUMERIC_ID),
        "with": ["ha", "configuration", "licenses", "analyticsMode", "selfHealing"],
        "sortBy": [{"attribute": "name", "type": "ASC"}],
        "limit": 2000,
        "_filterSpec": True,
    }
    result = portal_rpc("enterprise/getEnterpriseEdgeList", params, request_id=25)
    if not result:
        return []
    return result.get("data", [])


def _extract_profile_records(obj) -> list:
    """
    Recursively walk enterprise/getEnterpriseServices response and extract
    dicts that look like profile metadata (id + logicalId + enterpriseLogicalId).
    """
    found = []
    if isinstance(obj, dict):
        if (
            isinstance(obj.get("id"), int)
            and isinstance(obj.get("logicalId"), str)
            and isinstance(obj.get("name"), str)
            and (
                isinstance(obj.get("enterpriseLogicalId"), str)
                or obj.get("configurationType") in {"PROFILE", "profile"}
                or "profile" in str(obj.get("configurationType", "")).lower()
            )
        ):
            found.append({
                "id":                  obj["id"],
                "name":                obj["name"],
                "logicalId":           obj["logicalId"],
                "enterpriseLogicalId": obj.get("enterpriseLogicalId") or VCO_ENTERPRISE_LOGICAL_ID,
                "description":         obj.get("description"),
                "version":             obj.get("version"),
            })
        for v in obj.values():
            found.extend(_extract_profile_records(v))
    elif isinstance(obj, list):
        for v in obj:
            found.extend(_extract_profile_records(v))
    return found


def get_enterprise_events(hours: int = EVENTS_HOURS_DEFAULT) -> list:
    start_ms = int((datetime.now(timezone.utc) - timedelta(hours=hours)).timestamp() * 1000)
    params = {
        "enterpriseId": int(VCO_ENTERPRISE_NUMERIC_ID),
        "interval": {"start": start_ms},
    }
    result = portal_rpc("event/getEnterpriseEvents", params, request_id=90)
    if not result:
        return []
    data = result.get("data", result) if isinstance(result, dict) else result
    return data if isinstance(data, list) else []


def get_enterprise_services_payload():
    params = {
        "enterpriseId": int(VCO_ENTERPRISE_NUMERIC_ID),
        "with": ["profileCount", "edgeUsage", "configuration", "gateways"],
    }
    result = portal_rpc("enterprise/getEnterpriseServices", params, request_id=48)
    if isinstance(result, (dict, list)):
        return result
    return {}


def get_profile_catalog() -> list:
    result = get_enterprise_services_payload()
    if not result:
        return []
    records = _extract_profile_records(result)
    uniq = {r["id"]: r for r in records}
    return list(uniq.values())


# ─────────────────────────────────────────────────────────────────────────────
# APIv2 CONFIG FETCH
# ─────────────────────────────────────────────────────────────────────────────

def get_profile_device_settings(profile_lid: str):
    return api2_get(f"/enterprises/{VCO_ENTERPRISE_LOGICAL_ID}/profiles/{profile_lid}/deviceSettings")


def get_edge_device_settings(edge_lid: str):
    return api2_get(f"/enterprises/{VCO_ENTERPRISE_LOGICAL_ID}/edges/{edge_lid}/deviceSettings")


def get_edge_configuration_modules(edge_numeric_id: int) -> dict:
    """Portal RPC: effective configuration modules for an edge (includes firewall)."""
    result = portal_rpc(
        "edge/getEdgeConfigurationModules",
        {"edgeId": int(edge_numeric_id)},
        request_id=60,
    )
    return result if isinstance(result, dict) else {}


def get_edge_certificates(edge_numeric_id: int) -> list:
    """Portal RPC: edge certificate history (used for VPN cert validation / rotation)."""
    result = portal_rpc(
        "edge/getEdgeCertificates",
        {"edgeId": int(edge_numeric_id)},
        request_id=72 + (int(edge_numeric_id) % 1000),
    )
    if isinstance(result, list):
        return result
    return []


def get_enterprise_record() -> dict:
    """Portal RPC: enterprise metadata (domain, endpointPkiMode, …)."""
    result = portal_rpc(
        "enterprise/getEnterprise",
        {"enterpriseId": int(VCO_ENTERPRISE_NUMERIC_ID)},
        request_id=80,
    )
    return result if isinstance(result, dict) else {}


def get_enterprise_users() -> list:
    """Portal RPC: enterprise user accounts."""
    result = portal_rpc(
        "enterprise/getEnterpriseUsers",
        {"enterpriseId": int(VCO_ENTERPRISE_NUMERIC_ID)},
        request_id=81,
    )
    return result if isinstance(result, list) else []


def get_enterprise_api_tokens() -> list:
    """Portal RPC: enterprise API token metadata (no secret values returned)."""
    result = portal_rpc(
        "enterprise/getApiTokens",
        {"enterpriseId": int(VCO_ENTERPRISE_NUMERIC_ID)},
        request_id=82,
    )
    return result if isinstance(result, list) else []


def get_profile_configuration_modules(profile_numeric_id: int) -> list:
    """Portal RPC: profile configuration modules (deviceSettings, firewall, WAN, …)."""
    result = portal_rpc(
        "configuration/getConfigurationModules",
        {
            "enterpriseId": int(VCO_ENTERPRISE_NUMERIC_ID),
            "configurationId": int(profile_numeric_id),
        },
        request_id=int(profile_numeric_id) % 10000,
    )
    return result if isinstance(result, list) else []


def _portal_modules_dict(modules) -> dict:
    """Normalise Portal module list to {name: module_dict}."""
    if isinstance(modules, dict):
        return modules
    if not isinstance(modules, list):
        return {}
    return {
        m["name"]: m for m in modules
        if isinstance(m, dict) and m.get("name")
    }


def _module_data(modules: dict, name: str) -> dict:
    """Extract the data payload from a named configuration module."""
    mod = modules.get(name) if isinstance(modules, dict) else None
    if not isinstance(mod, dict):
        return {}
    data = mod.get("data")
    return data if isinstance(data, dict) else mod


# ─────────────────────────────────────────────────────────────────────────────
# JOIN LOGIC
# ─────────────────────────────────────────────────────────────────────────────

def build_edge_profile_join(edges: list, profile_catalog: list) -> list:
    profile_by_num_id = {p["id"]: p for p in profile_catalog}
    joined = []
    for e in edges:
        edge_name = e.get("name")
        if VCO_SCOPE_EDGES and edge_name not in VCO_SCOPE_EDGES:
            continue
        cfg       = e.get("configuration", {}) or {}
        prof      = cfg.get("enterprise", {}) or {}
        prof_nid  = prof.get("id")
        prof_meta = profile_by_num_id.get(prof_nid, {})
        joined.append({
            "edgeName":          edge_name,
            "edgeLogicalId":     e.get("logicalId"),
            "edgeNumericId":     e.get("id"),
            "profileNumericId":  prof_nid,
            "profileName":       prof.get("name"),
            "profileLogicalId":  prof_meta.get("logicalId"),
            "profileMeta":       prof_meta,
            "edgeOverrides":     cfg.get("edgeOverrides", {}) or {},
            "edgePortalRecord":  e,
        })
    return joined


# ─────────────────────────────────────────────────────────────────────────────
# FINDINGS BUILDER
# ─────────────────────────────────────────────────────────────────────────────

_FINDING_KEYS = [
    "category", "severity", "edgeName", "edgeLogicalId",
    "profileName", "profileNumericId", "profileLogicalId",
    "title", "fieldPath", "evidence", "note",
    "methodologyRef", "automation",
]

def _finding(
    ctx: dict,
    category: str,
    severity: str,
    title: str,
    field_path: str,
    evidence: str,
    note: str,
    methodology_ref: str = None,
    automation: str = "automated",
) -> dict:
    result = {
        "category":          category,
        "severity":          severity,
        "edgeName":          ctx["edgeName"],
        "edgeLogicalId":     ctx["edgeLogicalId"],
        "profileName":       ctx["profileName"],
        "profileNumericId":  ctx["profileNumericId"],
        "profileLogicalId":  ctx["profileLogicalId"],
        "title":             title,
        "fieldPath":         field_path,
        "evidence":          evidence,
        "note":              note,
        "methodologyRef":    methodology_ref or "",
        "automation":        automation,
    }
    return result


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 1 — FIREWALL RULE ANALYSIS
#
# Inspects every segment's business-policy rules and the ccFirewall block.
# Flags:
#   • any/any allow (src=any AND dst=any AND no port/proto constraint)
#   • overly broad CIDR  (prefix length < BROAD_CIDR_THRESHOLD for allow rules)
#   • missing protocol/port constraint on non-deny rules
#   • default action = allow/permit
# ─────────────────────────────────────────────────────────────────────────────

BROAD_CIDR_THRESHOLD = 16   # /15 or shorter prefix is flagged as overly broad


def _is_any(value) -> bool:
    """True if value represents 'any' / unrestricted."""
    if value is None:
        return True
    if isinstance(value, (int, float)) and int(value) == -1:
        return True
    s = str(value).strip().lower()
    return s in {"any", "all", "0.0.0.0/0", "0.0.0.0", "::/0", "", "*"}


def _is_broad_cidr(cidr_str: str) -> bool:
    """True if a CIDR is defined but has a prefix length below the threshold."""
    if not cidr_str or _is_any(cidr_str):
        return False
    try:
        net = ipaddress.ip_network(cidr_str, strict=False)
        return net.prefixlen < BROAD_CIDR_THRESHOLD and net.prefixlen > 0
    except ValueError:
        return False


def _is_allow(action) -> bool:
    if action is None:
        return False
    if isinstance(action, dict):
        action = action.get("allow_or_deny") or action.get("action")
    return str(action).strip().lower() in {"allow", "accept", "permit"}


def _normalise_rule(rule: dict) -> dict:
    """Return a normalised view of a firewall/policy rule dict regardless of key naming."""
    lk = {k.lower(): v for k, v in rule.items()}
    return {
        "action":   lk.get("action") or lk.get("policy") or lk.get("defaultaction"),
        "src":      lk.get("sip") or lk.get("source") or lk.get("srcip") or lk.get("sourceaddress"),
        "dst":      lk.get("dip") or lk.get("destination") or lk.get("dstip") or lk.get("destinationaddress"),
        "proto":    lk.get("proto") or lk.get("protocol"),
        "sport":    lk.get("sport") or lk.get("sport_low") or lk.get("sourceport"),
        "dport":    lk.get("dport") or lk.get("dport_low") or lk.get("destinationport"),
        "name":     lk.get("name") or lk.get("rulename") or "(unnamed)",
        "enabled":  lk.get("enabled"),
        "log":      lk.get("log") or lk.get("logging"),
        "_raw":     rule,
    }


def _analyse_rule(ctx, nr: dict, location: str, findings: list) -> None:
    """Inspect one normalised rule and append findings."""
    if not _is_allow(nr["action"]):
        return   # we only care about allow/permit rules for these checks

    src_any   = _is_any(nr["src"])
    dst_any   = _is_any(nr["dst"])
    no_proto  = _is_any(nr["proto"])
    no_port   = _is_any(nr["sport"]) and _is_any(nr["dport"])
    src_broad = _is_broad_cidr(str(nr["src"] or ""))
    dst_broad = _is_broad_cidr(str(nr["dst"] or ""))
    name      = nr["name"]

    # any/any allow (typically AllowAny catch-all — common VeloCloud design)
    if src_any and dst_any:
        findings.append(_finding(
            ctx, "firewall_rules", "low",
            "Allow rule with no source or destination constraint (any/any)",
            f"{location}.rule[{name}]",
            f"action={nr['action']}, src={nr['src']}, dst={nr['dst']}",
            "AllowAny-style catch-all is typical VeloCloud design after scoped rules. "
            "Tighten only if strict egress is required; [FW] Default Deny (stateful "
            "firewall) is the higher-priority control.",
            methodology_ref="[FW] Rule Scope",
        ))
        return  # no point stacking more findings on the same rule

    # overly broad source CIDR
    if src_broad:
        findings.append(_finding(
            ctx, "firewall_rules", "high",
            f"Allow rule with overly broad source CIDR (/{ipaddress.ip_network(str(nr['src']), strict=False).prefixlen})",
            f"{location}.rule[{name}].src",
            f"action={nr['action']}, src={nr['src']}",
            f"Source CIDR is broader than /{BROAD_CIDR_THRESHOLD}. "
            "Tighten to the minimum required source range.",
        ))

    # overly broad destination CIDR
    if dst_broad:
        findings.append(_finding(
            ctx, "firewall_rules", "high",
            f"Allow rule with overly broad destination CIDR (/{ipaddress.ip_network(str(nr['dst']), strict=False).prefixlen})",
            f"{location}.rule[{name}].dst",
            f"action={nr['action']}, dst={nr['dst']}",
            f"Destination CIDR is broader than /{BROAD_CIDR_THRESHOLD}. "
            "Restrict to the minimum required destination range.",
        ))

    # missing protocol and port constraint
    if no_proto and no_port:
        findings.append(_finding(
            ctx, "firewall_rules", "medium",
            "Allow rule missing protocol and port constraints",
            f"{location}.rule[{name}]",
            f"action={nr['action']}, proto={nr['proto']}, sport={nr['sport']}, dport={nr['dport']}",
            "Define explicit protocol (TCP/UDP/ICMP) and port constraints. "
            "Unrestricted allow rules expose all services.",
        ))
    elif no_proto:
        findings.append(_finding(
            ctx, "firewall_rules", "low",
            "Allow rule missing protocol constraint",
            f"{location}.rule[{name}].proto",
            f"action={nr['action']}, proto={nr['proto']}",
            "Specify a protocol to limit the allow rule scope.",
        ))


def _collect_rules_from_block(block, prefix: str) -> list:
    """Yield (location, normalised_rule) pairs from any structure that may contain rules."""
    pairs = []
    if isinstance(block, list):
        for i, item in enumerate(block):
            pairs.extend(_collect_rules_from_block(item, f"{prefix}[{i}]"))
    elif isinstance(block, dict):
        for k, v in block.items():
            lk = k.lower()
            if lk in {"rules", "rule", "firewallrules", "policyrules"}:
                if isinstance(v, list):
                    for j, r in enumerate(v):
                        if isinstance(r, dict):
                            pairs.append((f"{prefix}.{k}[{j}]", _normalise_rule(r)))
            elif isinstance(v, (dict, list)):
                pairs.extend(_collect_rules_from_block(v, f"{prefix}.{k}"))
    return pairs


def _flatten_firewall_rule(rule: dict, default_action: str = "deny") -> dict:
    """Normalise VeloCloud firewall-module or business-policy rule dict."""
    match = rule.get("match") or {}
    action = rule.get("action") or {}
    return {
        "name":   rule.get("name") or rule.get("ruleLogicalId") or "(unnamed)",
        "action": get_nested(action, "allow_or_deny")
                  or get_nested(action, "routePolicy")
                  or default_action,
        "sip":    get_nested(match, "sip"),
        "dip":    get_nested(match, "dip"),
        "proto":  get_nested(match, "proto"),
        "sport":  get_nested(match, "sport_low") or get_nested(match, "sport"),
        "dport":  get_nested(match, "dport_low") or get_nested(match, "dport"),
    }


def _firewall_has_rules(firewall_cfg: dict) -> bool:
    if not isinstance(firewall_cfg, dict):
        return False
    for direction in ("inbound", "outbound", "inboundV6", "outboundV6"):
        if firewall_cfg.get(direction):
            return True
    for seg in firewall_cfg.get("segments") or []:
        if not isinstance(seg, dict):
            continue
        for direction in ("inbound", "outbound", "inboundV6", "outboundV6"):
            if seg.get(direction):
                return True
    return False


def _analyse_firewall_module(
    ctx: dict,
    firewall_cfg: dict,
    findings: list,
    segment_context: dict = None,
) -> None:
    """Inspect effective firewall module (inbound/outbound rules per segment)."""
    if not isinstance(firewall_cfg, dict):
        return

    if firewall_cfg.get("stateful_firewall_enabled") is False:
        evidence = "stateful_firewall_enabled=False"
        if isinstance(segment_context, dict) and segment_context.get("segments"):
            evidence = json.dumps({
                "stateful_firewall_enabled": False,
                "interSegmentRouting": segment_context.get("interSegmentRouting"),
                "segmentCount": segment_context.get("segmentCount"),
                "segments": segment_context.get("segments"),
            }, ensure_ascii=False)
        findings.append(_finding(
            ctx, "firewall_rules", "high",
            "Stateful firewall disabled",
            "firewall.stateful_firewall_enabled",
            evidence,
            "Enable stateful firewall with deny-by-default policy.",
            methodology_ref="[FW] Default Deny",
        ))

    if (
        firewall_cfg.get("firewall_enabled") is False
        and _firewall_has_rules(firewall_cfg)
    ):
        findings.append(_finding(
            ctx, "firewall_rules", "medium",
            "Firewall disabled but rules are defined",
            "firewall.firewall_enabled",
            "firewall_enabled=False with active rule sets",
            "Enable firewall or remove orphaned rules.",
            methodology_ref="[FW] Default Deny",
        ))

    for direction in ("inbound", "outbound", "inboundV6", "outboundV6"):
        rules = firewall_cfg.get(direction) or []
        if isinstance(rules, list):
            for i, rule in enumerate(rules):
                if isinstance(rule, dict):
                    nr = _normalise_rule(_flatten_firewall_rule(rule))
                    _analyse_rule(ctx, nr, f"firewall.{direction}[{i}]", findings)

    segments = firewall_cfg.get("segments") or []
    for seg in (segments if isinstance(segments, list) else []):
        if not isinstance(seg, dict):
            continue
        seg_name = get_nested(seg, "segment", "name") or seg.get("segmentId", "?")
        for direction in ("inbound", "outbound", "inboundV6", "outboundV6"):
            rules = seg.get(direction) or []
            for i, rule in enumerate(rules if isinstance(rules, list) else []):
                if isinstance(rule, dict):
                    nr = _normalise_rule(_flatten_firewall_rule(rule))
                    _analyse_rule(
                        ctx, nr,
                        f"firewall.segments[{seg_name}].{direction}[{i}]",
                        findings,
                    )

    for settings_key in ("statefulFirewallSettings", "statefulFirewallSettingsV6"):
        settings = firewall_cfg.get(settings_key) or {}
        if isinstance(settings, dict):
            default_action = (
                settings.get("defaultAction")
                or settings.get("default_action")
                or settings.get("defaultFirewallAction")
            )
            if default_action and _is_allow(default_action):
                findings.append(_finding(
                    ctx, "firewall_rules", "high",
                    "Firewall default action is ALLOW",
                    f"firewall.{settings_key}.defaultAction",
                    f"defaultAction={default_action}",
                    "Change the default firewall policy to DENY. "
                    "Use explicit allow rules for required traffic only.",
                    methodology_ref="[FW] Default Deny",
                ))


def _segment_enforcement_lan_context(
    profile_cfg: dict,
    edge_cfg: dict,
    firewall_cfg: dict,
    edge_ds: dict,
    profile_modules: dict = None,
    profile_fw: dict = None,
) -> dict:
    """Build multi-segment LAN context for Default Deny evidence (no duplicate finding)."""
    effective = _effective_config(profile_cfg, edge_cfg)
    name_lookup = _build_segment_name_lookup(
        profile_modules, profile_cfg, profile_fw, firewall_cfg,
    )
    lan_map = _build_segment_lan_map(effective, name_lookup)
    inter_seg, inter_reason = _inter_segment_routing_enabled(lan_map, edge_ds)
    if not inter_seg:
        return {}
    lan_segments = []
    for seg_key, info in lan_map.items():
        if info.get("network_count", 0) <= 0:
            continue
        lan_segments.append({
            "segment": info.get("name") or seg_key,
            "lanNetworks": info.get("networks") or [],
            "networkCount": info.get("network_count", 0),
        })
    if not lan_segments:
        return {}
    return {
        "interSegmentRouting": inter_reason,
        "segmentCount": len(lan_segments),
        "segments": lan_segments,
    }


def analyse_firewall_rules(
    ctx: dict,
    profile_cfg: dict,
    edge_cfg: dict,
    findings: list,
    firewall_cfg: dict = None,
    profile_fw: dict = None,
    edge_ds: dict = None,
    profile_modules: dict = None,
) -> None:
    """
    Entry point for firewall rule analysis.
    Primary source: edge firewall module from Portal API when present.
    Falls back to profile firewall module, then deviceSettings ccFirewall / rules.
    """
    segment_context = _segment_enforcement_lan_context(
        profile_cfg, edge_cfg, firewall_cfg or {}, edge_ds,
        profile_modules=profile_modules, profile_fw=profile_fw,
    )
    if isinstance(firewall_cfg, dict) and firewall_cfg:
        _analyse_firewall_module(ctx, firewall_cfg, findings, segment_context)
        return

    if isinstance(profile_fw, dict) and profile_fw:
        _analyse_firewall_module(ctx, profile_fw, findings, segment_context)
        return

    for source_label, cfg in [("profile", profile_cfg), ("edge", edge_cfg)]:
        if not isinstance(cfg, dict):
            continue

        segments = cfg.get("segments") or []
        for seg in (segments if isinstance(segments, list) else []):
            seg_name = get_nested(seg, "segment", "name") or seg.get("segmentId", "?")
            rules_block = seg.get("rules") or seg.get("businessPolicyRules") or []
            for i, rule in enumerate(rules_block if isinstance(rules_block, list) else []):
                if isinstance(rule, dict):
                    flat = _flatten_firewall_rule(rule, default_action="allow")
                    _analyse_rule(ctx, _normalise_rule(flat),
                                  f"{source_label}.segments[{seg_name}].rules", findings)

        for location, nr in _collect_rules_from_block(cfg.get("ccFirewall") or {}, f"{source_label}.ccFirewall"):
            _analyse_rule(ctx, nr, location, findings)

        cc = cfg.get("ccFirewall") or {}
        for seg_fw in (cc.get("segments") or cc.get("rules") or []):
            default_action = None
            if isinstance(seg_fw, dict):
                default_action = (
                    seg_fw.get("defaultAction")
                    or seg_fw.get("default_action")
                    or get_nested(seg_fw, "defaults", "action")
                )
            if default_action and _is_allow(default_action):
                findings.append(_finding(
                    ctx, "firewall_rules", "high",
                    "Firewall default action is ALLOW",
                    f"{source_label}.ccFirewall.defaultAction",
                    f"defaultAction={default_action}",
                    "Change the default firewall policy to DENY. "
                    "Use explicit allow rules for required traffic only.",
                ))


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 1a — SEGMENT ENFORCEMENT (partial / assisted)
# ─────────────────────────────────────────────────────────────────────────────

_GLOBAL_SEGMENT_LABELS = {"global segment", "0"}


def _is_sensitive_vrf_segment(label: str, seg_id=None) -> bool:
    """Non-Global segments treated as sensitive VRFs for inbound posture review."""
    if str(label).strip().lower() in _GLOBAL_SEGMENT_LABELS:
        return False
    if seg_id is not None and str(seg_id) == "0":
        return False
    return True


def _match_field_to_network(match: dict, field: str):
    """Parse VeloCloud firewall/QoS match sip/dip to ip_network."""
    if not isinstance(match, dict):
        return None
    value = match.get(field)
    if value is None or _is_any(value):
        return None
    rule_type_key = "s_rule_type" if field == "sip" else "d_rule_type"
    rule_type = str(match.get(rule_type_key) or "prefix").lower()
    try:
        if rule_type == "netmask":
            mask_key = "ssm" if field == "sip" else "dsm"
            mask = match.get(mask_key)
            if mask:
                ip_int = int(ipaddress.ip_address(value))
                mask_int = int(ipaddress.ip_address(mask))
                prefix_len = bin(mask_int).count("1")
                return ipaddress.ip_network(f"{value}/{prefix_len}", strict=False)
        text = str(value)
        if "/" not in text:
            text = f"{text}/32"
        return ipaddress.ip_network(text, strict=False)
    except ValueError:
        return None


def _normalised_rule_prefixes(nr: dict) -> list:
    prefixes = []
    for field in ("src", "dst"):
        if nr.get(field) and not _is_any(nr.get(field)):
            try:
                text = str(nr[field])
                if "/" not in text:
                    text = f"{text}/32"
                prefixes.append(ipaddress.ip_network(text, strict=False))
            except ValueError:
                continue
    return prefixes


def _build_segment_lan_map(effective_cfg: dict, name_lookup: dict = None) -> dict:
    """
    Map segment key → {name, network_count, prefixes, networks}.
    Uses APIv2 effective deviceSettings lan.networks segment href binding.
    """
    seg_by_key = {
        _segment_key(seg): seg
        for seg in (effective_cfg.get("segments") or [])
        if isinstance(seg, dict)
    }
    lan_map = {}
    for net in get_nested(effective_cfg, "lan", "networks") or []:
        if not isinstance(net, dict) or net.get("disabled"):
            continue
        seg_ref = net.get("segment") or {}
        href_id = _segment_href_id({"segment": seg_ref})
        seg_key = None
        seg_label = None
        for key, seg in seg_by_key.items():
            if key == href_id:
                seg_key = key
                seg_label = _segment_display_name(seg, name_lookup)
                break
        if not seg_key:
            seg_key = href_id or "unknown"
            seg_label = href_id or "unknown"
        entry = lan_map.setdefault(seg_key, {
            "name": seg_label,
            "network_count": 0,
            "prefixes": [],
            "networks": [],
        })
        entry["network_count"] += 1
        entry["networks"].append(net.get("name") or "(unnamed)")
        cidr_ip = net.get("cidrIp")
        cidr_prefix = net.get("cidrPrefix")
        if cidr_ip is not None and cidr_prefix is not None:
            try:
                entry["prefixes"].append(
                    ipaddress.ip_network(f"{cidr_ip}/{cidr_prefix}", strict=False)
                )
            except ValueError:
                pass
    return lan_map


def _inter_segment_routing_enabled(lan_map: dict, edge_ds: dict) -> tuple:
    """True when multiple LAN-bearing segments exist or portal DS has segment routes."""
    lan_segments = [k for k, v in lan_map.items() if v.get("network_count", 0) > 0]
    if len(lan_segments) >= 2:
        return True, f"{len(lan_segments)} segments with LAN networks"
    for seg in (edge_ds or {}).get("segments") or []:
        if not isinstance(seg, dict):
            continue
        routes = seg.get("routes")
        if isinstance(routes, list) and routes:
            label = _segment_display_name(seg)
            return True, f"segment routes on {label}"
        route_maps = seg.get("routeMaps")
        if isinstance(route_maps, list) and route_maps:
            label = _segment_display_name(seg)
            return True, f"routeMaps on {label}"
    return False, ""


def _firewall_seg_signature(fseg: dict) -> tuple:
    parts = []
    for direction in ("inbound", "outbound", "inboundV6", "outboundV6"):
        for rule in fseg.get(direction) or []:
            if not isinstance(rule, dict):
                continue
            flat = _flatten_firewall_rule(rule)
            parts.append((
                rule.get("name") or rule.get("ruleLogicalId"),
                flat.get("action"),
                flat.get("sip"),
                flat.get("dip"),
                flat.get("proto"),
            ))
    return tuple(parts)


def _firewall_default_is_allow(fw_cfg: dict) -> bool:
    for settings_key in ("statefulFirewallSettings", "statefulFirewallSettingsV6"):
        settings = (fw_cfg or {}).get(settings_key) or {}
        if not isinstance(settings, dict):
            continue
        default_action = (
            settings.get("defaultAction")
            or settings.get("default_action")
            or settings.get("defaultFirewallAction")
        )
        if default_action and _is_allow(default_action):
            return True
    return False


def _weak_segment_deny_posture(fw_cfg: dict, fseg: dict) -> bool:
    """True when deny-by-default / inbound enforcement is weak for cross-segment allows."""
    if not isinstance(fw_cfg, dict):
        return True
    if fw_cfg.get("firewall_enabled") is False:
        return True
    if fw_cfg.get("stateful_firewall_enabled") is not True:
        return True
    if _firewall_default_is_allow(fw_cfg):
        return True
    inbound = (fseg.get("inbound") or []) + (fseg.get("inboundV6") or [])
    if not inbound:
        return True
    return not any(
        isinstance(r, dict) and not _is_allow(_flatten_firewall_rule(r).get("action"))
        for r in inbound
    )


def _qos_rule_is_permit(rule: dict) -> bool:
    if not isinstance(rule, dict):
        return False
    action = rule.get("action") or {}
    if action.get("allow_or_deny") and not _is_allow(action.get("allow_or_deny")):
        return False
    return bool(
        action.get("QoS")
        or action.get("edge2EdgeRouteAction")
        or action.get("edge2DataCenterRouteAction")
        or action.get("edge2CloudRouteAction")
    )


def _cross_segment_target(src_seg_key: str, lan_map: dict, prefixes: list):
    """Return target segment name if prefixes overlap another segment's LAN."""
    for prefix in prefixes or []:
        for key, info in lan_map.items():
            if key == src_seg_key:
                continue
            for lan_prefix in info.get("prefixes") or []:
                if prefix.overlaps(lan_prefix):
                    return info.get("name") or key
    return None


def analyse_segment_enforcement(
    ctx: dict,
    profile_cfg: dict,
    edge_cfg: dict,
    findings: list,
    firewall_cfg: dict = None,
    profile_fw: dict = None,
    edge_ds: dict = None,
    profile_ds: dict = None,
    edge_qos: dict = None,
    profile_modules: dict = None,
    profile_fw_module: dict = None,
) -> None:
    """
    [FW] Segment Enforcement — partial / assisted checks (5 heuristics).
    Requires portal firewall + deviceSettings + QOS modules where available.
    """
    if not isinstance(firewall_cfg, dict) or not firewall_cfg:
        return

    effective = _effective_config(profile_cfg, edge_cfg)
    name_lookup = _build_segment_name_lookup(
        profile_modules, profile_cfg, profile_fw_module or profile_fw, firewall_cfg,
    )
    lan_map = _build_segment_lan_map(effective, name_lookup)
    inter_seg, inter_reason = _inter_segment_routing_enabled(lan_map, edge_ds)
    edge_fw_by_key = {
        _segment_key(seg): seg
        for seg in (firewall_cfg.get("segments") or [])
        if isinstance(seg, dict)
    }
    profile_fw_by_key = {
        _segment_key(seg): seg
        for seg in ((profile_fw or {}).get("segments") or [])
        if isinstance(seg, dict)
    }
    firewall_override = bool((ctx.get("edgeOverrides") or {}).get("firewall"))

    # 1) Firewall globally disabled + LAN + inter-segment routing
    #    (stateless mode is reported under [FW] Default Deny with segment context)
    if inter_seg and firewall_cfg.get("firewall_enabled") is False:
        lan_segments = []
        for seg_key, info in lan_map.items():
            if info.get("network_count", 0) <= 0:
                continue
            lan_segments.append({
                "segment": info.get("name") or seg_key,
                "lanNetworks": info.get("networks") or [],
                "networkCount": info.get("network_count", 0),
            })
        if lan_segments:
            findings.append(_finding(
                ctx, "firewall_rules", "medium",
                "Segment firewall disabled on edge with multi-segment LAN and inter-segment routing",
                "firewall.firewall_enabled",
                json.dumps({
                    "interSegmentRouting": inter_reason,
                    "firewall_enabled": firewall_cfg.get("firewall_enabled"),
                    "segmentCount": len(lan_segments),
                    "segments": lan_segments,
                }, ensure_ascii=False),
                "Enable segment firewall and confirm boundaries between VRFs/segments.",
                methodology_ref="[FW] Segment Enforcement",
                automation="partial",
            ))

    # 2) No inbound rules on sensitive VRFs when stateful firewall on
    if firewall_cfg.get("stateful_firewall_enabled") is True:
        for seg_key, fseg in edge_fw_by_key.items():
            label = _segment_display_name(fseg, name_lookup)
            seg_id = get_nested(fseg, "segment", "segmentId")
            if not _is_sensitive_vrf_segment(label, seg_id):
                continue
            inbound = (fseg.get("inbound") or []) + (fseg.get("inboundV6") or [])
            if inbound:
                continue
            findings.append(_finding(
                ctx, "firewall_rules", "medium",
                f"Sensitive VRF segment {label} has no inbound firewall rules with stateful firewall enabled",
                f"firewall.segments[{label}].inbound",
                json.dumps({
                    "segment": label,
                    "stateful_firewall_enabled": True,
                    "inboundRuleCount": 0,
                }, ensure_ascii=False),
                "Add inbound rules or document accepted exposure for this VRF segment.",
                methodology_ref="[FW] Segment Enforcement",
                automation="partial",
            ))

    # 3) Edge segment firewall diverges from profile baseline
    for seg_key, efseg in edge_fw_by_key.items():
        pfseg = profile_fw_by_key.get(seg_key)
        if not pfseg:
            continue
        if _firewall_seg_signature(efseg) == _firewall_seg_signature(pfseg):
            continue
        label = _segment_display_name(efseg, name_lookup)
        findings.append(_finding(
            ctx, "firewall_rules", "medium",
            f"Edge segment firewall diverges from profile on segment {label}",
            f"firewall.segments[{label}]",
            json.dumps({
                "segment": label,
                "edgeFirewallOverride": firewall_override,
                "edgeOutboundRules": len(efseg.get("outbound") or []),
                "profileOutboundRules": len(pfseg.get("outbound") or []),
                "edgeInboundRules": len(efseg.get("inbound") or []),
                "profileInboundRules": len(pfseg.get("inbound") or []),
            }, ensure_ascii=False),
            "Confirm edge firewall divergence is intentional; align with profile if not.",
            methodology_ref="[FW] Segment Enforcement",
            automation="partial",
        ))

    # 4) Business policy (QOS) permits cross-segment flows
    qos_cfg = edge_qos if isinstance(edge_qos, dict) else {}
    for qseg in qos_cfg.get("segments") or []:
        if not isinstance(qseg, dict):
            continue
        src_key = _segment_key(qseg)
        src_label = _segment_display_name(qseg, name_lookup)
        for rule in qseg.get("rules") or []:
            if not isinstance(rule, dict) or not _qos_rule_is_permit(rule):
                continue
            match = rule.get("match") or {}
            prefixes = []
            for field in ("sip", "dip"):
                net = _match_field_to_network(match, field)
                if net:
                    prefixes.append(net)
            if _is_any(match.get("dip")):
                continue
            target = _cross_segment_target(src_key, lan_map, prefixes)
            if not target:
                continue
            findings.append(_finding(
                ctx, "firewall_rules", "medium",
                f"Business policy on segment {src_label} permits cross-segment flow toward {target}",
                f"QOS.segments[{src_label}].rules[{rule.get('name', '?')}]",
                json.dumps({
                    "sourceSegment": src_label,
                    "targetSegment": target,
                    "rule": rule.get("name"),
                    "match": {
                        "sip": match.get("sip"),
                        "dip": match.get("dip"),
                    },
                }, ensure_ascii=False),
                "Confirm cross-segment business policy is intended; restrict if segments must stay isolated.",
                methodology_ref="[FW] Segment Enforcement",
                automation="partial",
            ))

    # 5) Explicit inter-segment firewall allow without deny/default posture
    for seg_key, fseg in edge_fw_by_key.items():
        src_label = _segment_display_name(fseg, name_lookup)
        if _weak_segment_deny_posture(firewall_cfg, fseg):
            weak_posture = True
        else:
            weak_posture = False
        if not weak_posture:
            continue
        for direction in ("inbound", "outbound", "inboundV6", "outboundV6"):
            for rule in fseg.get(direction) or []:
                if not isinstance(rule, dict):
                    continue
                flat = _flatten_firewall_rule(rule)
                if not _is_allow(flat.get("action")):
                    continue
                nr = _normalise_rule(flat)
                if nr.get("name") == "AllowAny" and _is_any(nr.get("src")) and _is_any(nr.get("dst")):
                    continue
                prefixes = _normalised_rule_prefixes(nr)
                if _is_any(nr.get("dst")):
                    continue
                target = _cross_segment_target(seg_key, lan_map, prefixes)
                if not target:
                    continue
                findings.append(_finding(
                    ctx, "firewall_rules", "medium",
                    f"Inter-segment firewall allow on {src_label} toward {target} without deny/default posture",
                    f"firewall.segments[{src_label}].{direction}[{rule.get('name', '?')}]",
                    json.dumps({
                        "sourceSegment": src_label,
                        "targetSegment": target,
                        "rule": rule.get("name"),
                        "direction": direction,
                        "stateful_firewall_enabled": firewall_cfg.get("stateful_firewall_enabled"),
                        "defaultActionAllow": _firewall_default_is_allow(firewall_cfg),
                    }, ensure_ascii=False),
                    "Add deny-by-default posture or scoped allows before permitting cross-segment traffic.",
                    methodology_ref="[FW] Segment Enforcement",
                    automation="partial",
                ))


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 1b — FIREWALL ACCESS, LOGGING, OVERRIDES, PATCH, NAT
# ─────────────────────────────────────────────────────────────────────────────

def _service_allow_list_broad(allow_list) -> bool:
    if not isinstance(allow_list, list) or not allow_list:
        return True
    for entry in allow_list:
        s = str(entry).strip().lower()
        if _is_any(s) or s in {"0.0.0.0/0", "::/0"}:
            return True
        try:
            net = ipaddress.ip_network(s, strict=False)
            if net.prefixlen == 0:
                return True
        except ValueError:
            continue
    return False


def analyse_edge_access(ctx: dict, firewall_cfg: dict, findings: list) -> None:
    """[FW] Edge Local Access Restrictions — firewall services block."""
    if not isinstance(firewall_cfg, dict):
        return
    services = firewall_cfg.get("services") or {}
    if not isinstance(services, dict):
        return

    for svc_name in _EDGE_ACCESS_SERVICES:
        svc = services.get(svc_name)
        if not isinstance(svc, dict) or not svc.get("enabled"):
            continue
        allow_ips = svc.get("allowSelectedIp") or []
        if not allow_ips:
            findings.append(_finding(
                ctx, "edge_access", "high",
                f"Edge management service {svc_name} enabled without source IP restriction",
                f"firewall.services.{svc_name}",
                f"enabled=True, allowSelectedIp={allow_ips}",
                "Disable the service or restrict to approved source IPs.",
                methodology_ref="[FW] Edge Local Access Restrictions",
            ))
        elif _service_allow_list_broad(allow_ips):
            findings.append(_finding(
                ctx, "edge_access", "medium",
                f"Edge management service {svc_name} allows overly broad source range",
                f"firewall.services.{svc_name}.allowSelectedIp",
                f"allowSelectedIp={allow_ips}",
                "Restrict management access to minimum required source prefixes.",
                methodology_ref="[FW] Edge Local Access Restrictions",
            ))

    usb_disabled = services.get("usb.disabled")
    if usb_disabled is False:
        findings.append(_finding(
            ctx, "edge_access", "medium",
            "USB access not disabled on edge",
            "firewall.services.usb.disabled",
            "usb.disabled=False",
            "Disable USB access unless a documented exception exists.",
            methodology_ref="[FW] Edge Local Access Restrictions",
        ))


_ORCHESTRATOR_ALLOW_NETS = (
    ipaddress.ip_network("169.254.0.0/16"),
)


def _extract_lan_prefixes(cfg: dict) -> list:
    """LAN/VLAN networks from deviceSettings (overlay-reachable peer site prefixes)."""
    prefixes = []
    if not isinstance(cfg, dict):
        return prefixes
    for net in (cfg.get("lan") or {}).get("networks") or []:
        if not isinstance(net, dict) or net.get("disabled"):
            continue
        cidr_ip = net.get("cidrIp")
        prefix_len = net.get("cidrPrefix")
        if not cidr_ip or prefix_len is None:
            continue
        try:
            prefixes.append(ipaddress.ip_network(f"{cidr_ip}/{prefix_len}", strict=False))
        except ValueError:
            continue
    return prefixes


def _is_orchestrator_allow_ip(value: str) -> bool:
    """Exclude VeloCloud/Orchestrator link-local allow entries from overlay peer checks."""
    try:
        if "/" in str(value):
            net = ipaddress.ip_network(value, strict=False)
            return any(net.subnet_of(orch) for orch in _ORCHESTRATOR_ALLOW_NETS)
        ip = ipaddress.ip_address(value)
        return any(ip in orch for orch in _ORCHESTRATOR_ALLOW_NETS)
    except ValueError:
        return False


def _allow_overlaps_prefix(allow_entry: str, peer_prefix) -> bool:
    try:
        if "/" in str(allow_entry):
            allow_net = ipaddress.ip_network(allow_entry, strict=False)
        else:
            allow_net = ipaddress.ip_network(f"{allow_entry}/32", strict=False)
    except ValueError:
        return False
    if peer_prefix.version != allow_net.version:
        return False
    return peer_prefix.overlaps(allow_net)


def _profile_has_branch_to_branch(profile_cfg: dict, edge_cfg: dict) -> bool:
    """True when effective config has Cloud VPN branch-to-branch on any segment."""
    effective = _effective_config(profile_cfg, edge_cfg)
    for seg in effective.get("segments") or []:
        if not isinstance(seg, dict):
            continue
        vpn = seg.get("vpn") or {}
        if vpn.get("enabled") and vpn.get("edgeToEdge"):
            return True
    return False


def build_overlay_peer_prefixes(combined_records: list) -> dict:
    """
    Per edge: LAN prefixes of sibling edges on the same profile (potential overlay sources).
    Returns {edgeName: [{"prefix": ip_network, "peerEdge": str}, ...]}.
    """
    by_profile = {}
    edge_effective = {}
    for rec in combined_records:
        en = rec.get("edgeName")
        pid = rec.get("profileNumericId")
        if not en:
            continue
        by_profile.setdefault(pid, []).append(en)
        edge_effective[en] = _effective_config(
            rec.get("profileConfig") or {},
            rec.get("edgeConfig") or {},
        )

    result = {}
    for rec in combined_records:
        en = rec.get("edgeName")
        pid = rec.get("profileNumericId")
        peers = []
        for peer_name in by_profile.get(pid, []):
            if peer_name == en:
                continue
            for prefix in _extract_lan_prefixes(edge_effective.get(peer_name, {})):
                peers.append({"prefix": prefix, "peerEdge": peer_name})
        result[en] = peers
    return result


def analyse_overlay_management_access(
    ctx: dict,
    firewall_cfg: dict,
    profile_cfg: dict,
    edge_cfg: dict,
    peer_prefixes: list,
    findings: list,
) -> None:
    """
    [Net] Overlay to Management Access (partial) — config-only phase 1.
    Flags management allow lists that overlap same-profile peer LAN prefixes, or
    unrestricted management services while branch-to-branch VPN is enabled.
    """
    if not isinstance(firewall_cfg, dict):
        return
    services = firewall_cfg.get("services") or {}
    if not isinstance(services, dict):
        return

    branch_to_branch = _profile_has_branch_to_branch(profile_cfg, edge_cfg)

    for svc_name in _EDGE_ACCESS_SERVICES:
        svc = services.get(svc_name)
        if not isinstance(svc, dict) or not svc.get("enabled"):
            continue
        allow_ips = svc.get("allowSelectedIp") or []

        if not allow_ips and branch_to_branch:
            findings.append(_finding(
                ctx, "edge_access", "medium",
                f"Edge management service {svc_name} unrestricted with branch-to-branch VPN enabled",
                f"firewall.services.{svc_name}",
                json.dumps({
                    "service": svc_name,
                    "allowSelectedIp": allow_ips,
                    "branchToBranch": True,
                }, ensure_ascii=False),
                "Restrict management access to NOC/jump sources not reachable from overlay peers.",
                methodology_ref="[Net] Overlay to Management Access",
            ))
            continue

        for allow in allow_ips:
            if _is_orchestrator_allow_ip(allow):
                continue
            for peer in peer_prefixes or []:
                prefix = peer.get("prefix")
                if prefix is None:
                    continue
                if _allow_overlaps_prefix(allow, prefix):
                    findings.append(_finding(
                        ctx, "edge_access", "medium",
                        f"Edge management service {svc_name} allow list overlaps "
                        f"overlay-routable peer prefix",
                        f"firewall.services.{svc_name}.allowSelectedIp",
                        json.dumps({
                            "service": svc_name,
                            "allow": allow,
                            "peerEdge": peer.get("peerEdge"),
                            "peerPrefix": str(prefix),
                        }, ensure_ascii=False),
                        "Limit management allow lists to sources not routable from remote branches.",
                        methodology_ref="[Net] Overlay to Management Access",
                    ))
                    break


def analyse_firewall_logging(
    ctx: dict,
    firewall_cfg: dict,
    effective: dict,
    findings: list,
) -> None:
    """[FW] Firewall Event Logging + [Monitoring] Firewall Log Collection."""
    if not isinstance(firewall_cfg, dict):
        return

    fw_active = _firewall_module_active(firewall_cfg)
    global_log = firewall_cfg.get("firewall_logging_enabled")
    syslog_fwd = firewall_cfg.get("syslog_forwarding")
    segment_logging = any(
        isinstance(seg, dict) and seg.get("firewall_logging_enabled")
        for seg in (firewall_cfg.get("segments") or [])
    )
    syslog_ok = _segment_syslog_configured(effective)

    if fw_active and not global_log and not segment_logging:
        findings.append(_finding(
            ctx, "firewall_logging", "medium",
            "Firewall active but logging disabled",
            "firewall.firewall_logging_enabled",
            f"stateful={firewall_cfg.get('stateful_firewall_enabled')}, logging={global_log}",
            "Enable firewall event logging on active segments.",
            methodology_ref="[FW] Firewall Event Logging",
        ))
        # Do not also flag [Monitoring] Firewall Log Collection here — same root
        # cause; collection findings fire only when logging is enabled below.

    if (global_log or segment_logging) and not syslog_ok:
        findings.append(_finding(
            ctx, "firewall_logging", "medium",
            "Firewall logging enabled but syslog collectors not configured",
            "segments[].syslog",
            f"firewall_logging={global_log or segment_logging}, syslog_collectors=0",
            "Forward firewall logs via syslog collectors or Orchestrator log storage.",
            methodology_ref="[Monitoring] Firewall Log Collection",
        ))

    if (global_log or segment_logging) and syslog_fwd is False:
        findings.append(_finding(
            ctx, "firewall_logging", "low",
            "Firewall syslog forwarding disabled",
            "firewall.syslog_forwarding",
            "syslog_forwarding=False",
            "Enable syslog forwarding for firewall events.",
            methodology_ref="[Monitoring] Firewall Log Collection",
        ))


_OVERRIDE_LABELS = {
    "deviceSettings": "device settings",
    "firewall":       "firewall",
    "qos":            "QoS",
    "firewallDisabled": "firewall disabled flag",
}


def analyse_override_governance(ctx: dict, findings: list) -> None:
    """[System] Edge Override Governance — report active edge overrides."""
    overrides = ctx.get("edgeOverrides") or {}
    if not isinstance(overrides, dict):
        return
    active = {
        k: v for k, v in overrides.items()
        if v is True and k in _OVERRIDE_LABELS
    }
    if not active:
        return
    for key in sorted(active):
        label = _OVERRIDE_LABELS.get(key, key)
        findings.append(_finding(
            ctx, "override_governance", "low",
            f"Edge configuration override active: {label}",
            f"edgeOverrides.{key}",
            json.dumps(overrides, ensure_ascii=False),
            "Confirm override is documented, approved, and security-reviewed.",
            methodology_ref="[System] Edge Override Governance",
        ))


def analyse_software_update(ctx: dict, profile_cfg: dict, edge_cfg: dict, findings: list) -> None:
    """[System] Patch Levels — software update policy."""
    effective = _effective_config(profile_cfg, edge_cfg)
    sw = effective.get("softwareUpdate") or {}
    if not sw or sw.get("enabled") is False:
        findings.append(_finding(
            ctx, "device_hardening", "medium",
            "Software update policy not configured or disabled",
            "softwareUpdate",
            f"enabled={sw.get('enabled') if isinstance(sw, dict) else None}",
            "Configure software update policy for timely patch application.",
            methodology_ref="[System] Patch Levels",
        ))
        return
    rules = sw.get("rules") or sw.get("schedule") or sw.get("windows")
    if not rules:
        findings.append(_finding(
            ctx, "device_hardening", "low",
            "Software update enabled without defined schedule or rules",
            "softwareUpdate.rules",
            json.dumps(sw, ensure_ascii=False)[:200],
            "Define maintenance windows or update rules.",
            methodology_ref="[System] Patch Levels",
        ))


def _nat_rule_exposure(rule: dict) -> tuple:
    """Return (is_exposed, reason) for a NAT rule dict."""
    if not isinstance(rule, dict):
        return False, ""
    match = rule.get("match") or rule.get("source") or {}
    if isinstance(match, dict):
        src = match.get("sip") or match.get("source") or match.get("src")
    else:
        src = None
    if _is_any(src):
        dport = (
            rule.get("translatedPort")
            or rule.get("dport")
            or get_nested(rule, "match", "dport")
            or get_nested(rule, "match", "dport_low")
        )
        try:
            port_num = int(dport) if dport not in (None, -1, "") else None
        except (TypeError, ValueError):
            port_num = None
        if port_num in _SENSITIVE_NAT_PORTS:
            return True, f"sensitive_port={port_num}, src=any"
        return True, "source unrestricted (any)"
    return False, ""


def analyse_nat_exposure(ctx: dict, profile_cfg: dict, edge_cfg: dict, findings: list) -> None:
    """[FW] NAT Exposure — permissive DNAT in effective deviceSettings."""
    effective = _effective_config(profile_cfg, edge_cfg)
    for seg in effective.get("segments") or []:
        if not isinstance(seg, dict):
            continue
        seg_label = _segment_name(seg)
        for rules_key in ("natRules", "dualNatRules", "nat"):
            rules = seg.get(rules_key)
            if rules_key == "nat" and isinstance(rules, dict):
                rules = rules.get("rules")
            if not isinstance(rules, list):
                continue
            for i, rule in enumerate(rules):
                if not isinstance(rule, dict):
                    continue
                exposed, reason = _nat_rule_exposure(rule)
                if exposed:
                    findings.append(_finding(
                        ctx, "nat_exposure", "high",
                        f"Permissive NAT rule on segment {seg_label}",
                        f"segments[{seg_label}].{rules_key}[{i}]",
                        reason,
                        "Restrict NAT to required sources and avoid exposing sensitive services.",
                        methodology_ref="[FW] NAT Exposure",
                    ))


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 2 — SEGMENTATION ANALYSIS (baseline-driven)
#
# Compares profile segment definitions against the edge config.
# Flags:
#   • segments present in profile but absent on edge
#   • segments present on edge but not in profile (unexpected addition)
#   • VPN / inter-segment routing delta between profile and edge
# ─────────────────────────────────────────────────────────────────────────────

def _segment_name(seg: dict) -> str:
    return (
        get_nested(seg, "segment", "name")
        or str(seg.get("segmentId") or "unknown")
    )


def _segment_lookup_keys(seg: dict) -> list:
    """All keys used to correlate the same segment across APIv2 and Portal modules."""
    keys = []
    for candidate in (
        _segment_href_id(seg),
        get_nested(seg, "segment", "segmentLogicalId"),
        get_nested(seg, "segment", "segmentId"),
        seg.get("segmentId"),
        get_nested(seg, "segment", "name"),
    ):
        if candidate is None or candidate == "":
            continue
        if candidate not in keys:
            keys.append(candidate)
    return keys


def _build_segment_name_lookup(
    profile_modules: dict = None,
    profile_cfg: dict = None,
    profile_fw: dict = None,
    edge_fw: dict = None,
) -> dict:
    """Map segment identifiers (href id, logicalId, segmentId) → display name."""
    lookup = {}

    def absorb(segs) -> None:
        for seg in segs or []:
            if not isinstance(seg, dict):
                continue
            name = get_nested(seg, "segment", "name")
            if not name:
                continue
            for key in _segment_lookup_keys(seg):
                lookup[key] = str(name)

    prof_ds = _module_data(profile_modules or {}, "deviceSettings")
    prof_ds_segs = prof_ds.get("segments") or []
    absorb(prof_ds_segs)
    # APIv2 profile deviceSettings often has segment._href only; align by index with Portal names.
    prof_cfg_segs = (profile_cfg.get("segments") or []) if isinstance(profile_cfg, dict) else []
    for i, pmod_seg in enumerate(prof_ds_segs):
        if not isinstance(pmod_seg, dict):
            continue
        name = get_nested(pmod_seg, "segment", "name")
        if not name or i >= len(prof_cfg_segs):
            continue
        cfg_seg = prof_cfg_segs[i]
        if isinstance(cfg_seg, dict):
            for key in _segment_lookup_keys(cfg_seg):
                lookup[key] = str(name)
    if isinstance(profile_cfg, dict):
        absorb(profile_cfg.get("segments"))
    if isinstance(profile_fw, dict):
        absorb(profile_fw.get("segments"))
    if isinstance(edge_fw, dict):
        absorb(edge_fw.get("segments"))
    return lookup


def _segment_display_name(seg: dict, name_lookup: dict = None) -> str:
    """Prefer Portal segment.name; fall back to lookup, segmentId, unknown."""
    if name_lookup:
        for key in _segment_lookup_keys(seg):
            if key in name_lookup:
                return name_lookup[key]
    name = get_nested(seg, "segment", "name")
    if name:
        return str(name)
    seg_id = get_nested(seg, "segment", "segmentId")
    if seg_id is not None:
        return str(seg_id)
    return _segment_name(seg)


def _vpn_isolation_weak(vpn: dict) -> tuple:
    """
    True when branch-to-branch VPN is on without isolation groups / isolate profile.
    Maps to UI: Cloud VPN on + Branch to Branch enabled + Isolate Profile off.
    """
    if not isinstance(vpn, dict) or not vpn.get("enabled"):
        return False, ""
    if not vpn.get("edgeToEdge"):
        return False, ""
    detail = vpn.get("edgeToEdgeDetail") or {}
    if (detail.get("isolation") or {}).get("enabled"):
        return False, ""
    if detail.get("isolationGroups"):
        return False, ""
    dynamic = detail.get("dynamic") or {}
    if dynamic.get("enabled") and not (dynamic.get("isolation") or {}).get("enabled"):
        return True, "dynamic branch-to-branch without isolation"
    return True, "branch-to-branch VPN without isolation"


def analyse_segment_isolation(
    ctx: dict,
    profile_cfg: dict,
    edge_cfg: dict,
    findings: list,
    profile_modules: dict = None,
    profile_fw: dict = None,
    edge_fw: dict = None,
) -> None:
    """
    [Net] Segment Isolation — VPN branch-to-branch without isolation only (high/medium).
    Uses effective profile+edge deviceSettings. Firewall rules are handled separately.
    """
    effective = _effective_config(profile_cfg, edge_cfg)
    name_lookup = _build_segment_name_lookup(
        profile_modules, profile_cfg, profile_fw, edge_fw,
    )
    profile_seg_by_key = {
        _segment_key(seg): seg
        for seg in (profile_cfg.get("segments") or [])
        if isinstance(seg, dict)
    }

    for seg in effective.get("segments") or []:
        if not isinstance(seg, dict):
            continue
        label = _segment_display_name(seg, name_lookup)
        vpn = seg.get("vpn") or {}
        weak, reason = _vpn_isolation_weak(vpn)
        if weak:
            detail = vpn.get("edgeToEdgeDetail") or {}
            findings.append(_finding(
                ctx, "segmentation", "high",
                "Branch-to-branch VPN enabled without segment isolation",
                f"segments[{label}].vpn",
                json.dumps({
                    "segment": label,
                    "edgeToEdge": vpn.get("edgeToEdge"),
                    "vpnEnabled": vpn.get("enabled"),
                    "isolation": (detail.get("isolation") or {}).get("enabled"),
                    "isolationGroupCount": len(detail.get("isolationGroups") or []),
                    "dynamicEnabled": (detail.get("dynamic") or {}).get("enabled"),
                    "detail": reason,
                }, ensure_ascii=False),
                "Enable Isolate Profile / VPN isolation groups, or disable branch-to-branch "
                "on segments that must not communicate laterally.",
                methodology_ref="[Net] Segment Isolation",
            ))

        prof_seg = profile_seg_by_key.get(_segment_key(seg))
        if prof_seg:
            p_weak, _ = _vpn_isolation_weak(prof_seg.get("vpn") or {})
            if not p_weak and weak:
                findings.append(_finding(
                    ctx, "segmentation", "medium",
                    "Edge override enables branch-to-branch VPN without isolation",
                    f"segments[{label}].vpn",
                    json.dumps({"segment": label, "profileIsolated": True, "edgeWeak": True}, ensure_ascii=False),
                    "Edge deviceSettings override widens VPN isolation vs profile baseline.",
                    methodology_ref="[Net] Segment Isolation",
                ))


def _vpn_edge_to_edge_enabled(vpn: dict) -> bool:
    return isinstance(vpn, dict) and bool(vpn.get("edgeToEdge"))


def analyse_edge_to_edge_communication(
    ctx: dict,
    profile_cfg: dict,
    edge_cfg: dict,
    findings: list,
    profile_modules: dict = None,
    profile_fw: dict = None,
    edge_fw: dict = None,
) -> None:
    """
    [Net] Edge-to-Edge Communication — edge must not widen branch-to-branch VPN vs profile.
    Profile-level mesh VPN is covered by [Net] Segment Isolation.
    """
    if not isinstance(profile_cfg, dict) or not isinstance(edge_cfg, dict):
        return

    effective = _effective_config(profile_cfg, edge_cfg)
    name_lookup = _build_segment_name_lookup(
        profile_modules, profile_cfg, profile_fw, edge_fw,
    )
    profile_seg_by_key = {
        _segment_key(seg): seg
        for seg in (profile_cfg.get("segments") or [])
        if isinstance(seg, dict) and _segment_key(seg)
    }

    for seg in effective.get("segments") or []:
        if not isinstance(seg, dict):
            continue
        prof_seg = profile_seg_by_key.get(_segment_key(seg))
        if not prof_seg:
            continue
        label = _segment_display_name(seg, name_lookup)
        profile_e2e = _vpn_edge_to_edge_enabled(prof_seg.get("vpn") or {})
        effective_e2e = _vpn_edge_to_edge_enabled(seg.get("vpn") or {})
        if not profile_e2e and effective_e2e:
            findings.append(_finding(
                ctx, "segmentation", "high",
                "Edge override enables branch-to-branch VPN wider than profile baseline",
                f"segments[{label}].vpn.edgeToEdge",
                json.dumps({
                    "segment": label,
                    "profileEdgeToEdge": profile_e2e,
                    "effectiveEdgeToEdge": effective_e2e,
                }, ensure_ascii=False),
                "Remove edge-level branch-to-branch widening or update profile baseline with approval.",
                methodology_ref="[Net] Edge-to-Edge Communication",
            ))


def analyse_segmentation(ctx: dict, profile_cfg: dict, edge_cfg: dict, findings: list) -> None:
    if not isinstance(profile_cfg, dict) or not isinstance(edge_cfg, dict):
        return

    p_segs = profile_cfg.get("segments") or []
    e_segs = edge_cfg.get("segments") or []

    if not isinstance(p_segs, list) or not isinstance(e_segs, list):
        return

    p_names = {_segment_name(s) for s in p_segs}
    e_names = {_segment_name(s) for s in e_segs}

    # Segments defined in profile but missing from edge override
    for missing in sorted(p_names - e_names):
        findings.append(_finding(
            ctx, "segmentation", "medium",
            "Profile-defined segment missing from edge config",
            "segments",
            f"Profile segment '{missing}' not present in edge deviceSettings",
            "Verify the edge inherits this segment from the profile or confirm intentional removal.",
            methodology_ref="[Isolation] Profile Inheritance",
        ))

    # Segments present on edge but not in profile baseline
    for added in sorted(e_names - p_names):
        findings.append(_finding(
            ctx, "segmentation", "medium",
            "Edge contains segment not defined in profile baseline",
            "segments",
            f"Edge-only segment '{added}' has no counterpart in the profile",
            "Confirm this segment was intentionally added as an edge-specific override "
            "and that its security controls have been reviewed.",
            methodology_ref="[Isolation] Profile Inheritance",
        ))

    # Per-segment: compare VPN and inter-segment routing flags
    p_seg_map = index_by(p_segs, "segmentId") or {_segment_name(s): s for s in p_segs}
    e_seg_map = index_by(e_segs, "segmentId") or {_segment_name(s): s for s in e_segs}

    for seg_key in set(p_seg_map) & set(e_seg_map):
        p_seg = p_seg_map[seg_key]
        e_seg = e_seg_map[seg_key]
        seg_label = _segment_name(p_seg)

        # VPN edge-to-datacentre
        p_e2dc = get_nested(p_seg, "vpn", "edgeToDataCenter")
        e_e2dc = get_nested(e_seg, "vpn", "edgeToDataCenter")
        if p_e2dc is not None and e_e2dc is not None and p_e2dc != e_e2dc:
            findings.append(_finding(
                ctx, "segmentation", "medium",
                "VPN edge-to-datacentre setting differs from profile baseline",
                f"segments[{seg_label}].vpn.edgeToDataCenter",
                f"profile={p_e2dc}, edge={e_e2dc}",
                "Confirm the edge-level override on this VPN setting is approved and documented.",
                methodology_ref="[Isolation] Profile Inheritance",
            ))

        # NAT override check
        p_nat = get_nested(p_seg, "nat", "rules")
        e_nat = get_nested(e_seg, "nat", "rules")
        if isinstance(p_nat, list) and isinstance(e_nat, list):
            if len(e_nat) > len(p_nat):
                findings.append(_finding(
                    ctx, "segmentation", "low",
                    "Edge has more NAT rules than profile baseline",
                    f"segments[{seg_label}].nat.rules",
                    f"profile rules={len(p_nat)}, edge rules={len(e_nat)}",
                    "Review the additional edge-specific NAT rules to confirm they are intended.",
                    methodology_ref="[Isolation] Profile Inheritance",
                ))


# ─────────────────────────────────────────────────────────────────────────────
# SHARED CONFIG MERGE
# ─────────────────────────────────────────────────────────────────────────────

def _is_empty_override(value) -> bool:
    return value in (None, {}, [], "")


def _segment_href_id(seg: dict) -> str:
    href = get_nested(seg, "segment", "_href") or ""
    if isinstance(href, str) and href:
        return href.rstrip("/").split("/")[-1]
    return ""


def _segment_key(seg: dict):
    return (
        _segment_href_id(seg)
        or get_nested(seg, "segment", "segmentLogicalId")
        or get_nested(seg, "segment", "segmentId")
        or seg.get("segmentId")
        or get_nested(seg, "segment", "name")
    )


def _deep_merge_dict(base: dict, override: dict) -> dict:
    """Merge override into base; empty override values preserve base fields."""
    merged = deepcopy(base) if isinstance(base, dict) else {}
    if not isinstance(override, dict):
        return merged
    for key, value in override.items():
        if _is_empty_override(value):
            continue
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            nested = deepcopy(merged[key])
            for sub_key, sub_val in value.items():
                if not _is_empty_override(sub_val):
                    nested[sub_key] = sub_val
            merged[key] = nested
        else:
            merged[key] = value
    return merged


def _merge_segment(profile_seg: dict, edge_seg: dict) -> dict:
    """Merge one edge segment onto profile baseline, preserving inherited telemetry."""
    if not isinstance(profile_seg, dict):
        return deepcopy(edge_seg) if isinstance(edge_seg, dict) else {}
    if not isinstance(edge_seg, dict):
        return deepcopy(profile_seg)

    merged = deepcopy(profile_seg)
    for key, value in edge_seg.items():
        if key in _SEGMENT_MERGE_KEYS:
            merged[key] = _deep_merge_dict(profile_seg.get(key) or {}, value or {})
        elif not _is_empty_override(value):
            merged[key] = value
    return merged


def _effective_config(profile_cfg: dict, edge_cfg: dict) -> dict:
    """
    Profile baseline with edge overrides applied.
    Segment-level deep merge for syslog/ntp/bfd so partial edge segment arrays
    do not drop inherited profile telemetry settings.
    """
    effective = deepcopy(profile_cfg) if isinstance(profile_cfg, dict) else {}
    if not isinstance(edge_cfg, dict):
        return effective

    for key, value in edge_cfg.items():
        if key == "segments":
            continue
        if key in _TOP_LEVEL_MERGE_KEYS:
            effective[key] = _deep_merge_dict(effective.get(key) or {}, value or {})
        elif not _is_empty_override(value):
            effective[key] = value

    profile_segments = profile_cfg.get("segments") if isinstance(profile_cfg, dict) else []
    edge_segments = edge_cfg.get("segments") or []
    if isinstance(edge_segments, list) and edge_segments:
        profile_map = {
            _segment_key(seg): seg
            for seg in (profile_segments if isinstance(profile_segments, list) else [])
            if isinstance(seg, dict)
        }
        effective["segments"] = [
            _merge_segment(profile_map.get(_segment_key(seg), {}), seg)
            for seg in edge_segments
            if isinstance(seg, dict)
        ]
    return effective


def _parse_version_tuple(version: str) -> tuple:
    parts = []
    for piece in re.split(r"[.\-]", version or ""):
        if piece.isdigit():
            parts.append(int(piece))
        elif parts:
            break
    return tuple(parts)


def _version_below_minimum(version: str, minimum: str) -> bool:
    if not minimum:
        return False
    return _parse_version_tuple(version) < _parse_version_tuple(minimum)


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 3 — DEVICE HARDENING FINDINGS
# ─────────────────────────────────────────────────────────────────────────────

def _segment_syslog_configured(effective: dict) -> bool:
    for seg in effective.get("segments") or []:
        if not isinstance(seg, dict):
            continue
        syslog = seg.get("syslog") or {}
        if syslog.get("enabled") and syslog.get("collectors"):
            return True
    return False


def _firewall_module_active(firewall_cfg: dict) -> bool:
    if not isinstance(firewall_cfg, dict):
        return False
    return bool(
        firewall_cfg.get("stateful_firewall_enabled")
        or firewall_cfg.get("firewall_enabled")
    )


_PSK_AUTH_MODES = {
    "CERTIFICATE_DEACTIVATED",
    "DEACTIVATED", "PSK", "PRE_SHARED_KEY", "PRE_SHARED",
}
_CERT_AUTH_MODES = {
    "CERTIFICATE_ACQUIRE",
    "CERTIFICATE_REQUIRED",
}
_HA_CPE_SUFFIX_RE = re.compile(r"^(?P<base>.+)-(CPE0|CPE1)$", re.IGNORECASE)


def _normalize_auth_mode(mode) -> str:
    return re.sub(r"[\s_\-]+", "_", str(mode or "").strip().upper())


def _edge_portal_auth_mode(portal: dict) -> str:
    return _normalize_auth_mode((portal or {}).get("authenticationMode"))


def _is_psk_auth_mode(auth_mode: str) -> bool:
    if not auth_mode:
        return False
    if auth_mode in _PSK_AUTH_MODES:
        return True
    return "DEACTIVATED" in auth_mode or auth_mode.endswith("_PSK")


def _is_cert_auth_mode(auth_mode: str) -> bool:
    if not auth_mode:
        return False
    if auth_mode in _CERT_AUTH_MODES:
        return True
    return "ACQUIRE" in auth_mode or "REQUIRED" in auth_mode


def _aaa_service_configured(service) -> bool:
    if isinstance(service, list):
        return any(_aaa_service_configured(item) for item in service)
    if not isinstance(service, dict) or not service:
        return False
    if service.get("enabled") is False:
        return False
    for key in ("server", "servers", "host", "ipAddress", "primary", "secondary"):
        val = service.get(key)
        if isinstance(val, dict) and val:
            return True
        if isinstance(val, list) and val:
            return True
        if isinstance(val, str) and val.strip():
            return True
    return False


def _admin_aaa_configured(effective: dict) -> tuple:
    """Return (configured, mechanism) for edge administrative AAA."""
    tacacs = effective.get("tacacs") or {}
    if _aaa_service_configured(tacacs):
        return True, "TACACS"

    radius = effective.get("radius") or {}
    if _aaa_service_configured(radius):
        return True, "RADIUS"

    auth = effective.get("authentication") or {}
    if isinstance(auth, dict):
        if auth.get("radiusEnabled") or auth.get("useRadius"):
            if _aaa_service_configured(auth.get("radius") or auth):
                return True, "RADIUS"
        if auth.get("tacacsEnabled") or auth.get("useTacacs"):
            if _aaa_service_configured(auth.get("tacacs") or auth):
                return True, "TACACS"

    for svc in effective.get("authServices") or []:
        if not isinstance(svc, dict) or svc.get("enabled") is False:
            continue
        svc_type = _normalize_auth_mode(svc.get("type") or svc.get("protocol"))
        if svc_type in {"TACACS", "RADIUS"} and _aaa_service_configured(svc):
            return True, svc_type

    return False, ""


def _analyse_edge_admin_aaa(ctx: dict, effective: dict, findings: list) -> None:
    configured, mechanism = _admin_aaa_configured(effective)
    if configured:
        return
    findings.append(_finding(
        ctx, "device_hardening", "low",
        "Centralised admin authentication not configured (no TACACS or RADIUS detected)",
        "deviceSettings.tacacs|radius|authServices",
        "absent",
        "Configure TACACS or RADIUS for edge administrative access.",
        methodology_ref="[System] Edge Admin AAA (TACACS/RADIUS)",
    ))


def _ha_control_fingerprint(label: str, effective: dict) -> str:
    if label == "ntp":
        ntp = effective.get("ntp") or {}
        servers = sorted(str(s) for s in (ntp.get("servers") or []))
        return json.dumps(
            {"enabled": ntp.get("enabled"), "servers": servers},
            sort_keys=True,
        )
    if label == "syslog":
        collectors = []
        for seg in effective.get("segments") or []:
            if not isinstance(seg, dict):
                continue
            syslog = seg.get("syslog") or {}
            if not syslog.get("enabled"):
                continue
            for collector in syslog.get("collectors") or []:
                if not isinstance(collector, dict):
                    continue
                collectors.append({
                    "host": collector.get("host") or collector.get("ip"),
                    "port": collector.get("port"),
                    "protocol": collector.get("protocol") or collector.get("proto"),
                })
        collectors.sort(key=lambda item: json.dumps(item, sort_keys=True))
        return json.dumps(collectors, sort_keys=True)
    if label == "snmp":
        snmp = effective.get("snmp") or {}
        v2c = snmp.get("snmpv2c") or {}
        v3 = snmp.get("snmpv3") or {}
        return json.dumps({
            "enabled": snmp.get("enabled"),
            "snmpv2cEnabled": v2c.get("enabled"),
            "snmpv3Enabled": v3.get("enabled") if isinstance(v3, dict) else None,
        }, sort_keys=True)
    if label == "admin_aaa":
        configured, mechanism = _admin_aaa_configured(effective)
        return json.dumps({"configured": configured, "mechanism": mechanism}, sort_keys=True)
    if label == "bfd":
        bfd = effective.get("bfd") or {}
        return json.dumps({"enabled": bfd.get("enabled")}, sort_keys=True)
    return ""


_HA_PARITY_CONTROLS = (
    ("NTP", "ntp"),
    ("syslog forwarding", "syslog"),
    ("SNMP", "snmp"),
    ("admin AAA", "admin_aaa"),
    ("BFD", "bfd"),
)


def _record_ctx(rec: dict) -> dict:
    return {k: rec[k] for k in
            ["edgeName", "edgeLogicalId", "profileName",
             "profileNumericId", "profileLogicalId",
             "edgeOverrides", "edgePortalRecord"]}


def build_ha_cpe_pairs(combined_records: list) -> list:
    """Match explicit CPE0/CPE1 hub pairs for parity comparison."""
    buckets = {}
    for rec in combined_records:
        name = rec.get("edgeName") or ""
        match = _HA_CPE_SUFFIX_RE.match(name)
        if not match:
            continue
        base = match.group("base").upper()
        role = "CPE0" if name.upper().endswith("-CPE0") else "CPE1"
        buckets.setdefault(base, {})[role] = rec
    pairs = []
    for members in buckets.values():
        if "CPE0" in members and "CPE1" in members:
            pairs.append((members["CPE0"], members["CPE1"]))
    return pairs


def analyse_ha_pair_parity(combined_records: list, findings: list) -> None:
    """[System] High Availability Configuration — compare HA peer effective configs."""
    for rec_a, rec_b in build_ha_cpe_pairs(combined_records):
        name_a = rec_a.get("edgeName", "?")
        name_b = rec_b.get("edgeName", "?")
        eff_a = _effective_config(rec_a.get("profileConfig") or {}, rec_a.get("edgeConfig") or {})
        eff_b = _effective_config(rec_b.get("profileConfig") or {}, rec_b.get("edgeConfig") or {})
        ctx = _record_ctx(rec_a)
        for label, key in _HA_PARITY_CONTROLS:
            fp_a = _ha_control_fingerprint(key, eff_a)
            fp_b = _ha_control_fingerprint(key, eff_b)
            if fp_a == fp_b:
                continue
            findings.append(_finding(
                ctx, "ha_config", "medium",
                f"HA pair config mismatch on {label}: {name_a} vs {name_b}",
                "ha.pairParity",
                json.dumps({
                    "peerA": name_a,
                    "peerB": name_b,
                    "control": label,
                    "peerAConfig": json.loads(fp_a) if fp_a else None,
                    "peerBConfig": json.loads(fp_b) if fp_b else None,
                }, ensure_ascii=False),
                "Align HA peer configuration for symmetric failover behaviour.",
                methodology_ref="[System] High Availability Configuration",
            ))


def analyse_edge_certificate_auth(
    ctx: dict,
    edge_certs: list,
    enterprise: dict,
    findings: list,
) -> None:
    """[System] Edge Certificate Authentication — portal auth mode + cert health."""
    portal = ctx.get("edgePortalRecord") or {}
    auth_mode = _edge_portal_auth_mode(portal)
    pki_mode = _normalize_auth_mode((enterprise or {}).get("endpointPkiMode"))
    serials = set()
    for key in ("serialNumber", "haSerialNumber"):
        val = portal.get(key)
        if val:
            serials.add(str(val))

    policy_evidence = {
        "authenticationMode": portal.get("authenticationMode"),
        "endpointPkiMode": (enterprise or {}).get("endpointPkiMode"),
    }

    if _is_psk_auth_mode(auth_mode):
        findings.append(_finding(
            ctx, "edge_certificate_auth", "medium",
            "Edge uses pre-shared key authentication (Certificate Deactivated)",
            "edgePortalRecord.authenticationMode",
            json.dumps(policy_evidence, ensure_ascii=False),
            "Migrate to Certificate Acquire or Certificate Required per hardening guidance.",
            methodology_ref="[System] Edge Certificate Authentication",
            automation="partial",
        ))
    elif not _is_cert_auth_mode(auth_mode):
        if pki_mode in {"", "CERTIFICATE_DISABLED", "CERTIFICATE_OPTIONAL"}:
            findings.append(_finding(
                ctx, "edge_certificate_auth", "low",
                "Edge certificate authentication mode not set to Certificate Acquire/Required",
                "edgePortalRecord.authenticationMode|enterprise.endpointPkiMode",
                json.dumps(policy_evidence, ensure_ascii=False),
                "Use Certificate Acquire or Certificate Required instead of pre-shared key only.",
                methodology_ref="[System] Edge Certificate Authentication",
                automation="partial",
            ))

    if not _is_cert_auth_mode(auth_mode) and not _is_psk_auth_mode(auth_mode):
        return

    certs_by_serial = _certs_by_serial(edge_certs or [])
    now = datetime.now(timezone.utc)
    targets = serials or {"unknown"}
    for serial in sorted(targets):
        latest = _latest_cert(certs_by_serial.get(serial, []))
        if not latest:
            sev = "high" if auth_mode.endswith("REQUIRED") or "REQUIRED" in auth_mode else "medium"
            findings.append(_finding(
                ctx, "edge_certificate_auth", sev,
                f"No edge certificate found for serial {serial}",
                "edge/getEdgeCertificates",
                json.dumps({"edgeSerialNumber": serial, **policy_evidence}, ensure_ascii=False),
                "Ensure the edge has a valid VCO-issued certificate for authentication.",
                methodology_ref="[System] Edge Certificate Authentication",
                automation="partial",
            ))
            continue
        valid_to = _parse_vco_timestamp(latest.get("validTo"))
        if not valid_to:
            continue
        days_left = (valid_to - now).days
        if days_left < 0:
            findings.append(_finding(
                ctx, "edge_certificate_auth", "high",
                f"Edge certificate expired for serial {serial}",
                "edge/getEdgeCertificates",
                json.dumps({
                    "edgeSerialNumber": serial,
                    "validTo": latest.get("validTo"),
                    "daysExpired": abs(days_left),
                }, ensure_ascii=False),
                "Renew edge certificates before authentication or overlay tunnels fail.",
                methodology_ref="[System] Edge Certificate Authentication",
                automation="partial",
            ))
        elif days_left <= _CERT_EXPIRY_WARN_DAYS:
            findings.append(_finding(
                ctx, "edge_certificate_auth", "medium",
                f"Edge certificate expiring within {_CERT_EXPIRY_WARN_DAYS} days for serial {serial}",
                "edge/getEdgeCertificates",
                json.dumps({
                    "edgeSerialNumber": serial,
                    "validTo": latest.get("validTo"),
                    "daysRemaining": days_left,
                }, ensure_ascii=False),
                "Confirm automatic certificate renewal is active.",
                methodology_ref="[System] Edge Certificate Authentication",
                automation="partial",
            ))


def analyse_device_hardening(
    ctx: dict,
    profile_cfg: dict,
    edge_cfg: dict,
    findings: list,
    firewall_cfg: dict = None,
) -> None:
    effective = _effective_config(profile_cfg, edge_cfg)

    ntp = effective.get("ntp") or {}
    if not ntp.get("enabled") or not (ntp.get("servers") or []):
        findings.append(_finding(
            ctx, "device_hardening", "medium",
            "NTP not configured or disabled",
            "ntp",
            f"enabled={ntp.get('enabled')}, servers={len(ntp.get('servers') or [])}",
            "Configure NTP servers on the profile or edge for accurate time synchronisation.",
            methodology_ref="[System] NTP Time Synchronisation",
        ))

    if not _segment_syslog_configured(effective):
        findings.append(_finding(
            ctx, "device_hardening", "medium",
            "Syslog forwarding not configured",
            "segments[].syslog",
            "No segment with syslog enabled and collectors defined",
            "Enable syslog collectors under Device > Telemetry > Syslog for central log collection.",
            methodology_ref="[Monitoring] Central Logging",
        ))

    snmp = effective.get("snmp") or {}
    v2c = snmp.get("snmpv2c") or {}
    v3 = snmp.get("snmpv3") or {}
    if v2c.get("enabled"):
        findings.append(_finding(
            ctx, "device_hardening", "medium",
            "SNMPv2c enabled",
            "snmp.snmpv2c",
            f"enabled=True, allowedIp={v2c.get('allowedIp', [])}",
            "Prefer SNMPv3 with authentication or disable SNMPv2c community access.",
            methodology_ref="[System] SNMP Hardening",
        ))
    elif snmp.get("enabled") and not (isinstance(v3, dict) and v3.get("enabled")):
        findings.append(_finding(
            ctx, "device_hardening", "low",
            "SNMP enabled without SNMPv3 authentication",
            "snmp",
            f"snmpv3.enabled={v3.get('enabled') if isinstance(v3, dict) else None}",
            "Use SNMPv3 with authentication or disable SNMP.",
            methodology_ref="[System] SNMP Hardening",
        ))

    if not _firewall_module_active(firewall_cfg):
        cc_fw = effective.get("ccFirewall") or {}
        if cc_fw.get("enabled") is False:
            findings.append(_finding(
                ctx, "device_hardening", "high",
                "Edge firewall control block disabled",
                "ccFirewall.enabled",
                "enabled=False",
                "Enable the firewall control block unless a documented exception exists.",
            ))

    _analyse_edge_admin_aaa(ctx, effective, findings)

    for iface in effective.get("routedInterfaces") or []:
        if not isinstance(iface, dict) or iface.get("disabled"):
            continue
        overlay = str(iface.get("wanOverlay", "")).upper()
        if overlay in {"USER_DISABLED", "DISABLED", "NONE", ""}:
            continue
        if iface.get("encryptOverlay") is False:
            name = iface.get("name", "?")
            findings.append(_finding(
                ctx, "device_hardening", "high",
                f"WAN overlay encryption disabled on interface {name}",
                f"routedInterfaces[{name}].encryptOverlay",
                f"wanOverlay={iface.get('wanOverlay')}, encryptOverlay=False",
                "Enable encryptOverlay on WAN interfaces carrying overlay traffic.",
                methodology_ref="[Net] Overlay Traffic Encryption",
            ))


def analyse_bfd(ctx: dict, profile_cfg: dict, edge_cfg: dict, findings: list) -> None:
    """[System] BFD Link Detection."""
    effective = _effective_config(profile_cfg, edge_cfg)
    bfd = effective.get("bfd") or {}
    if bfd.get("enabled"):
        return
    for iface in effective.get("routedInterfaces") or []:
        if not isinstance(iface, dict) or iface.get("disabled"):
            continue
        overlay = str(iface.get("wanOverlay", "")).upper()
        if overlay in {"USER_DISABLED", "DISABLED", "NONE", ""}:
            continue
        name = iface.get("name", "?")
        findings.append(_finding(
            ctx, "device_hardening", "low",
            f"BFD not enabled; active WAN interface {name} may lack fast failure detection",
            "bfd",
            f"bfd.enabled={bfd.get('enabled')}, interface={name}",
            "Enable BFD on WAN links where fast failure detection is required.",
            methodology_ref="[System] BFD Link Detection",
        ))
        return


def analyse_dns(ctx: dict, profile_cfg: dict, edge_cfg: dict, findings: list) -> None:
    """[Net] DNS Configuration."""
    effective = _effective_config(profile_cfg, edge_cfg)
    dns = effective.get("dns") or {}
    servers = dns.get("servers") or dns.get("dnsServers") or []
    if isinstance(servers, list) and servers:
        return
    for seg in effective.get("segments") or []:
        if isinstance(seg, dict):
            seg_dns = seg.get("dns") or {}
            seg_servers = seg_dns.get("servers") or seg_dns.get("dnsServers") or []
            if isinstance(seg_servers, list) and seg_servers:
                return
    findings.append(_finding(
        ctx, "device_hardening", "low",
        "DNS servers not configured on edge effective config",
        "dns",
        "No DNS servers in top-level or segment config",
        "Configure trusted DNS resolvers appropriate for the site.",
        methodology_ref="[Net] DNS Configuration",
    ))


def analyse_ha_config(ctx: dict, findings: list) -> None:
    """[System] High Availability Configuration."""
    portal = ctx.get("edgePortalRecord") or {}
    ha = portal.get("ha") or {}
    ha_type = str(ha.get("type", "")).upper()
    if ha_type in {"", "NONE"}:
        return
    ha_data = ha.get("data") or {}
    ha_state = str(ha_data.get("haState", "")).upper()
    if ha_type in {"ACTIVE_STANDBY", "ACTIVE_ACTIVE"} and not portal.get("haSerialNumber"):
        findings.append(_finding(
            ctx, "edge_inventory", "low",
            "Enhanced HA edge missing standby serial metadata",
            "edgePortalRecord.haSerialNumber",
            json.dumps({"haType": ha_type, "serialNumber": portal.get("serialNumber")}, ensure_ascii=False),
            "Verify HA standby unit is paired and reporting.",
            methodology_ref="[System] High Availability Configuration",
        ))
    if ha_state and ha_state not in {"READY", "ACTIVE", "STANDBY"}:
        findings.append(_finding(
            ctx, "edge_inventory", "medium",
            f"HA edge state is {ha_state}",
            "edgePortalRecord.ha.data.haState",
            json.dumps(ha_data, ensure_ascii=False)[:200],
            "Verify HA pair health and failover readiness.",
            methodology_ref="[System] High Availability Configuration",
        ))
    last_contact = ha_data.get("haLastContact")
    if last_contact:
        try:
            lc = datetime.fromisoformat(str(last_contact).replace("Z", "+00:00"))
            age_days = (datetime.now(timezone.utc) - lc).days
            if age_days >= STALE_EDGE_DAYS:
                findings.append(_finding(
                    ctx, "edge_inventory", "medium",
                    f"HA peer last contact {age_days} days ago",
                    "edgePortalRecord.ha.data.haLastContact",
                    f"haLastContact={last_contact}",
                    "Investigate HA peer connectivity.",
                    methodology_ref="[System] High Availability Configuration",
                ))
        except ValueError:
            pass


def analyse_security_features(ctx: dict, firewall_cfg: dict, findings: list) -> None:
    """[FW] Advanced Security Features — ATP/IDPS/URL filtering."""
    if not isinstance(firewall_cfg, dict):
        return
    features = firewall_cfg.get("securityFeatures") or {}
    if not isinstance(features, dict):
        return
    disabled = [
        k for k, v in features.items()
        if k.endswith("Enabled") and v is False
    ]
    if disabled:
        findings.append(_finding(
            ctx, "security_features", "low",
            "Advanced security features disabled",
            "firewall.securityFeatures",
            json.dumps(features, ensure_ascii=False),
            "Consider enabling IDPS, URL filtering, or malicious IP filtering if licensed.",
            methodology_ref="[FW] Advanced Security Features (ATP/IDPS/URL)",
            automation="partial",
        ))


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 4 — EDGE INVENTORY (portal metadata)
# ─────────────────────────────────────────────────────────────────────────────

_OFFLINE_STATES = {"OFFLINE", "DISCONNECTED", "NEVER_ACTIVATED"}


def analyse_edge_inventory(ctx: dict, findings: list) -> None:
    portal = ctx.get("edgePortalRecord") or {}
    edge_name = ctx.get("edgeName")
    state = str(portal.get("edgeState") or "").upper()
    activation = str(portal.get("activationState") or "").upper()

    if state in _OFFLINE_STATES:
        findings.append(_finding(
            ctx, "edge_inventory", "high",
            f"Edge is {state}",
            "edgePortalRecord.edgeState",
            f"edgeState={state}, lastContact={portal.get('lastContact')}",
            "Investigate offline edges; decommission or restore as appropriate.",
        ))

    last_contact = portal.get("lastContact")
    if last_contact and state not in _OFFLINE_STATES:
        try:
            lc = datetime.fromisoformat(str(last_contact).replace("Z", "+00:00"))
            age_days = (datetime.now(timezone.utc) - lc).days
            if age_days >= STALE_EDGE_DAYS:
                findings.append(_finding(
                    ctx, "edge_inventory", "medium",
                    f"Edge not contacted orchestrator in {age_days} days",
                    "edgePortalRecord.lastContact",
                    f"lastContact={last_contact}, age_days={age_days}",
                    f"Review edge health; stale contact exceeds {STALE_EDGE_DAYS} day threshold.",
                ))
        except ValueError:
            pass

    if activation and activation != "ACTIVATED":
        findings.append(_finding(
            ctx, "edge_inventory", "medium",
            f"Edge activation state is {activation}",
            "edgePortalRecord.activationState",
            f"activationState={activation}",
            "Confirm whether non-activated edges should remain in inventory.",
        ))

    sw_version = portal.get("softwareVersion") or ""
    if MIN_EDGE_VERSION and _version_below_minimum(sw_version, MIN_EDGE_VERSION):
        findings.append(_finding(
            ctx, "edge_inventory", "high",
            f"Edge software version below minimum ({MIN_EDGE_VERSION})",
            "edgePortalRecord.softwareVersion",
            f"softwareVersion={sw_version}, minimum={MIN_EDGE_VERSION}",
            "Upgrade edge to a supported software release.",
        ))
    elif sw_version and _parse_version_tuple(sw_version) and _parse_version_tuple(sw_version)[0] < 6:
        findings.append(_finding(
            ctx, "edge_inventory", "low",
            "Edge software version may be end-of-support",
            "edgePortalRecord.softwareVersion",
            f"softwareVersion={sw_version}",
            "Verify the edge release is still supported by the vendor.",
        ))


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 5 — ENTERPRISE EVENTS (optional --events)
# ─────────────────────────────────────────────────────────────────────────────

def _enterprise_ctx() -> dict:
    return {
        "edgeName":          "(enterprise)",
        "edgeLogicalId":     VCO_ENTERPRISE_LOGICAL_ID,
        "profileName":       None,
        "profileNumericId":  None,
        "profileLogicalId":  None,
    }


def _event_name(event: dict) -> str:
    return str(
        event.get("event")
        or event.get("eventName")
        or event.get("name")
        or ""
    ).strip()


def _is_aggregate_failure_event(name: str) -> bool:
    upper = name.upper()
    return (
        upper in _AGGREGATE_FAILURE_EVENTS
        or (upper.endswith("_FAILURE") and "LOGIN" in upper)
        or (upper.endswith("_FAILED") and "LOGIN" in upper)
    )


def _event_methodology_refs(name: str) -> list:
    """Map an event name to one or more monitoring methodology checks."""
    upper = name.upper()
    if upper in _IGNORED_EVENT_NAMES:
        return []
    refs = []
    if (
        _is_aggregate_failure_event(name)
        or upper in _ADMIN_ACTIVITY_EVENTS
        or ("LOGIN" in upper and "FAIL" in upper)
        or upper in {"EDGE_LOCALUI_LOGIN", "EDGE_SSH_LOGIN", "EDGE_MGMT_LOGIN"}
    ):
        refs.append("[Monitoring] Admin Activity Logs")
    if upper.startswith("MGD_CONF_") or upper in _CONFIG_CHANGE_EVENTS:
        refs.append("[Monitoring] Config Change Logs")
    if upper not in _IGNORED_EVENT_NAMES:
        refs.append("[Monitoring] Event Review")
    return list(dict.fromkeys(refs))


def _event_severity(name: str, aggregate: bool = False) -> str:
    upper = name.upper()
    if "FAILED" in upper or "ROLLBACK" in upper or "PENDING" in upper:
        return "high"
    if upper in _EVENT_REVIEW_ALERT_EVENTS or upper in _ADMIN_ACTIVITY_EVENTS:
        return "medium" if not aggregate else "info"
    if upper == "MGD_CONF_APPLIED":
        return "info"
    return "info" if aggregate else "medium"


def _append_event_finding(
    findings: list,
    ctx: dict,
    name: str,
    methodology_ref: str,
    severity: str,
    title: str,
    evidence: dict,
    note: str,
) -> None:
    findings.append(_finding(
        ctx, "enterprise_events", severity,
        title,
        "event/getEnterpriseEvents",
        json.dumps(evidence, ensure_ascii=False),
        note,
        methodology_ref=methodology_ref,
        automation="partial",
    ))


def analyse_enterprise_events(events: list, hours: int = EVENTS_HOURS_DEFAULT) -> list:
    """
    Build enterprise event findings for Monitoring checks.
    Emits Event Review summary, admin-activity aggregates, and config-change events.
    """
    ctx = _enterprise_ctx()
    findings = []
    all_counts = {}
    review_counts = {}
    aggregate_buckets = {}
    seen_individual = set()

    for ev in events:
        if not isinstance(ev, dict):
            continue
        name = _event_name(ev)
        if not name:
            continue
        upper = name.upper()
        all_counts[upper] = all_counts.get(upper, 0) + 1
        if upper in _IGNORED_EVENT_NAMES:
            continue
        review_counts[upper] = review_counts.get(upper, 0) + 1

        edge = ev.get("edgeName") or ev.get("edge") or "(enterprise)"
        event_time = ev.get("eventTime") or ev.get("time")
        refs = _event_methodology_refs(name)
        if not refs:
            continue

        if upper in _AGGREGATE_EVENT_NAMES or _is_aggregate_failure_event(name):
            for ref in refs:
                bucket_key = (upper, edge, ref)
                bucket = aggregate_buckets.setdefault(bucket_key, {
                    "count": 0,
                    "users": set(),
                    "first_time": event_time,
                    "last_time": event_time,
                    "sample_message": (ev.get("message") or "")[:200],
                })
                bucket["count"] += 1
                user = ev.get("user")
                if user:
                    bucket["users"].add(str(user))
                if event_time and (not bucket["first_time"] or event_time < bucket["first_time"]):
                    bucket["first_time"] = event_time
                if event_time and (not bucket["last_time"] or event_time > bucket["last_time"]):
                    bucket["last_time"] = event_time
            continue

        key = (upper, edge, event_time, tuple(refs))
        if key in seen_individual:
            continue
        seen_individual.add(key)

        sev = _event_severity(name)
        evidence = {
            "event": name,
            "edge": edge,
            "user": ev.get("user"),
            "severity": ev.get("severity"),
            "time": event_time,
            "message": (ev.get("message") or "")[:200],
        }
        for ref in refs:
            if ref == "[Monitoring] Event Review" and upper not in _EVENT_REVIEW_ALERT_EVENTS:
                continue
            note = "Review orchestrator event for security or configuration impact."
            if ref == "[Monitoring] Admin Activity Logs":
                note = "Review administrative activity for unauthorised or unexpected changes."
            elif ref == "[Monitoring] Config Change Logs":
                note = "Review configuration change event and confirm it was authorised."
            _append_event_finding(
                findings, ctx, name, ref, sev,
                f"Enterprise event: {name}",
                evidence, note,
            )

    for (name, edge, ref), bucket in sorted(aggregate_buckets.items()):
        sev = _event_severity(name, aggregate=True)
        if _is_aggregate_failure_event(name) or ("LOGIN" in name and "FAIL" in name.upper()):
            sev = "high"
        if ref == "[Monitoring] Admin Activity Logs" and name in {
            "MGD_CONF_APPLIED", "FIREWALL_ENABLE",
        }:
            continue
        if ref == "[Monitoring] Config Change Logs" and name in {
            "USER_LOGIN_FAILURE", "USER_LOGIN_FAILED", "LOGIN_FAILED",
            "CREATE_API_TOKEN", "DOWNLOAD_API_TOKEN", "REMOTE_ACTION", "ROLE_RESET",
        }:
            continue
        if ref == "[Monitoring] Event Review" and name not in _EVENT_REVIEW_ALERT_EVENTS and name != "MGD_CONF_APPLIED":
            continue

        title = f"Enterprise event: {name} ({bucket['count']} occurrence(s))"
        evidence = {
            "event": name,
            "edge": edge,
            "count": bucket["count"],
            "users": sorted(bucket["users"])[:10],
            "first_time": bucket["first_time"],
            "last_time": bucket["last_time"],
            "sample_message": bucket["sample_message"],
        }
        note = "Review aggregated orchestrator events in the lookback window."
        if ref == "[Monitoring] Admin Activity Logs":
            note = "Review repeated administrative or authentication events."
        elif ref == "[Monitoring] Config Change Logs":
            note = "Review aggregated configuration change activity."
        elif ref == "[Monitoring] Event Review":
            note = "Security-relevant events detected during automated event review."
        _append_event_finding(findings, ctx, name, ref, sev, title, evidence, note)

    top_types = sorted(review_counts.items(), key=lambda kv: (-kv[1], kv[0]))[:15]
    security_count = sum(
        count for event_name, count in review_counts.items()
        if event_name in _EVENT_REVIEW_ALERT_EVENTS
        or event_name.startswith("MGD_CONF_")
        or event_name in _ADMIN_ACTIVITY_EVENTS
    )
    _append_event_finding(
        findings, ctx, "EVENT_REVIEW_SUMMARY", "[Monitoring] Event Review", "info",
        f"Enterprise event review: {sum(review_counts.values())} reviewable event(s) in {hours}h window",
        {
            "hours": hours,
            "totalEvents": len(events),
            "reviewableEvents": sum(review_counts.values()),
            "securityRelevantEvents": security_count,
            "eventTypes": [{"event": k, "count": v} for k, v in top_types],
            "ignoredRoutineLogins": all_counts.get("BROWSER_ENTERPRISE_LOGIN", 0),
        },
        "Review event inventory; investigate security-relevant types and confirm alerting coverage.",
    )

    return findings


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 6 — HARDENING NOTES (informational)
#
# Emitted into hardening.json separately — not deduplicated into findings.csv
# as primary signal (device_hardening module covers key items as findings).
# ─────────────────────────────────────────────────────────────────────────────

HARDENING_CONTROLS = {
    "ntp":           "NTP / time synchronisation",
    "softwareUpdate":"Software update policy",
    "snmp":          "SNMP / monitoring telemetry",
    "tacacs":        "Centralised admin authentication (TACACS/RADIUS)",
    "ccFirewall":    "Firewall control block",
    "segments":      "Network segmentation",
    "bfd":           "BFD link-failure detection",
}

def evaluate_hardening(ctx: dict, profile_cfg: dict, edge_cfg: dict) -> list:
    """Informational hardening notes for hardening.json."""
    notes = []
    effective = _effective_config(profile_cfg, edge_cfg)

    for key, label in HARDENING_CONTROLS.items():
        val = effective.get(key)
        absent = val in (None, {}, [], "")
        disabled = False
        if isinstance(val, dict):
            disabled = val.get("enabled") is False
        elif isinstance(val, list) and len(val) == 0:
            absent = True

        if absent or disabled:
            notes.append({
                "edgeName":         ctx["edgeName"],
                "edgeLogicalId":    ctx["edgeLogicalId"],
                "profileName":      ctx["profileName"],
                "control":          key,
                "label":            label,
                "status":           "disabled" if disabled else "absent",
                "recommendation":   f"Review whether '{label}' should be configured for this edge. "
                                    "If genuinely not required, document the exception.",
            })
    return notes


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 7 — STRUCTURAL DIFF (from velo_profile_diff.py)
#
# Compares LAN networks, LAN interfaces, and routed interfaces between
# profile baseline and edge override — only when the override flag is set.
# ─────────────────────────────────────────────────────────────────────────────

_LAN_NET_FIELDS  = [("disabled",), ("advertise",), ("pingResponse",), ("dnsProxy",),
                    ("dhcp", "enabled"), ("ospf", "enabled")]
_LAN_IF_FIELDS   = [("disabled",), ("portMode",), ("untaggedVlan",), ("vlanIds",)]
_WAN_IF_FIELDS   = [("disabled",), ("wanOverlay",), ("encryptOverlay",),
                    ("natDirect",), ("pingResponse",), ("addressing", "type")]
_HA_DIFF_SUPPRESS_FIELDS = {
    "lan.networks": {
        ("pingResponse",), ("advertise",), ("disabled",),
        ("dhcp", "enabled"), ("dnsProxy",),
    },
    "lan.interfaces": {
        ("disabled",), ("vlanIds",), ("untaggedVlan",), ("portMode",),
    },
    "routedInterfaces": {
        ("pingResponse",), ("disabled",), ("addressing", "type"),
    },
}


def _downgrade_severity(severity: str) -> str:
    return {"high": "medium", "medium": "low", "low": "info"}.get(severity, severity)


def _is_ha_edge(ctx: dict) -> bool:
    """True for enhanced HA / CPE pair edges where config drift is often expected."""
    portal = ctx.get("edgePortalRecord") or {}
    ha = portal.get("ha") or {}
    ha_type = str(ha.get("type", "")).upper()
    if ha_type in {"ACTIVE_STANDBY", "ACTIVE_ACTIVE", "HOT_STANDBY"}:
        return True
    ha_data = ha.get("data") or {}
    if str(ha_data.get("haMode", "")).lower() == "enhanced":
        return True
    name = (ctx.get("edgeName") or "").upper()
    return any(marker in name for marker in ("CPE0", "CPE1", "-EHA", "_EHA"))


def _apply_ha_severity(ctx: dict, severity: str) -> str:
    return _downgrade_severity(severity) if _is_ha_edge(ctx) else severity


def _compare_keyed_list(ctx, p_list, e_list, key_field, object_label,
                        check_fields, category, findings, severity_map=None,
                        methodology_ref="[System] Config Consistency Real") -> None:
    p_map = index_by(p_list or [], key_field)
    e_map = index_by(e_list or [], key_field)
    all_keys = sorted(set(p_map) | set(e_map))
    for k in all_keys:
        p = p_map.get(k)
        e = e_map.get(k)
        loc = f"{object_label}[{key_field}={k}]"
        if p and not e:
            findings.append(_finding(ctx, "config_diff", _apply_ha_severity(ctx, "high"),
                f"{object_label} present in profile but missing on edge", loc,
                json.dumps(p, ensure_ascii=False)[:200],
                "Verify whether this was intentionally removed at the edge level.",
                methodology_ref=methodology_ref))
            continue
        if e and not p:
            if _is_ha_edge(ctx) and object_label == "routedInterfaces":
                continue
            findings.append(_finding(ctx, "config_diff", _apply_ha_severity(ctx, "low"),
                f"Additional {object_label} on edge not in profile baseline", loc,
                json.dumps(e, ensure_ascii=False)[:200],
                "Confirm this edge-specific addition is intended and has been reviewed.",
                methodology_ref=methodology_ref))
            continue
        for path in check_fields:
            pv, ev = p, e
            valid = True
            for part in path:
                pv = pv.get(part) if isinstance(pv, dict) else None
                ev = ev.get(part) if isinstance(ev, dict) else None
                if pv is None:
                    valid = False
                    break
            if not valid:
                continue
            if pv != ev:
                if _is_ha_edge(ctx):
                    suppress = _HA_DIFF_SUPPRESS_FIELDS.get(object_label, set())
                    if tuple(path) in suppress:
                        continue
                sev = "medium"
                if severity_map and path in severity_map:
                    sev = severity_map[path]
                findings.append(_finding(ctx, "config_diff", _apply_ha_severity(ctx, sev),
                    f"{object_label} field differs from profile baseline",
                    f"{loc}.{'.'.join(path)}",
                    f"profile={json.dumps(pv)}, edge={json.dumps(ev)}",
                    "Review whether this edge-level override is intentional and approved.",
                    methodology_ref=methodology_ref))


def _infer_profile_routed_interfaces(profile_cfg: dict, edge_portal: dict) -> list:
    models = profile_cfg.get("models") or {}
    family = edge_portal.get("deviceFamily", "")
    model  = (edge_portal.get("modelNumber") or "").lower()
    if "6x0" in family.lower() or "610" in model:
        candidates = ["edge6X0", "edge610lte"]
    elif "3x10" in model or "3810" in model:
        candidates = ["edge3X10", "edge3X00"]
    elif "3x00" in family.lower() or "3800" in model:
        candidates = ["edge3X00", "edge3X10"]
    else:
        candidates = [k for k, v in models.items() if isinstance(v, dict) and "routedInterfaces" in v]
    for c in candidates:
        ri = get_nested(models, c, "routedInterfaces")
        if isinstance(ri, list):
            return ri
    return []


def analyse_structural_diff(ctx: dict, profile_cfg: dict, edge_cfg: dict, findings: list) -> None:
    """Only run if the edge has deviceSettings override enabled."""
    overrides = ctx.get("edgeOverrides") or {}
    if not overrides.get("deviceSettings", False):
        return

    if not isinstance(profile_cfg, dict) or not isinstance(edge_cfg, dict):
        return

    _compare_keyed_list(
        ctx,
        get_nested(profile_cfg, "lan", "networks"),
        get_nested(edge_cfg, "lan", "networks"),
        "vlanId", "lan.networks", _LAN_NET_FIELDS, "config_diff", findings,
    )
    _compare_keyed_list(
        ctx,
        get_nested(profile_cfg, "lan", "interfaces"),
        get_nested(edge_cfg, "lan", "interfaces"),
        "name", "lan.interfaces", _LAN_IF_FIELDS, "config_diff", findings,
    )
    p_ri = _infer_profile_routed_interfaces(profile_cfg, ctx.get("edgePortalRecord") or {})
    e_ri = edge_cfg.get("routedInterfaces") or []
    _compare_keyed_list(
        ctx, p_ri, e_ri, "name", "routedInterfaces", _WAN_IF_FIELDS, "config_diff", findings,
        severity_map={("encryptOverlay",): "high", ("natDirect",): "high"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 8 — BUSINESS POLICY / WAN
# ─────────────────────────────────────────────────────────────────────────────

def analyse_business_policy(
    ctx: dict,
    wan_cfg: dict,
    profile_cfg: dict,
    edge_cfg: dict,
    findings: list,
) -> None:
    """[Net] Default Segment Behaviour / Business Policy Override."""
    effective = _effective_config(profile_cfg, edge_cfg)
    for seg in effective.get("segments") or []:
        if not isinstance(seg, dict):
            continue
        seg_label = _segment_name(seg)
        rules = seg.get("rules") or seg.get("businessPolicyRules") or []
        for i, rule in enumerate(rules if isinstance(rules, list) else []):
            if not isinstance(rule, dict):
                continue
            flat = _flatten_firewall_rule(rule, default_action="allow")
            nr = _normalise_rule(flat)
            if _is_allow(nr["action"]) and _is_any(nr["src"]) and _is_any(nr["dst"]):
                findings.append(_finding(
                    ctx, "business_policy", "high",
                    f"Business policy allow-all rule on segment {seg_label}",
                    f"segments[{seg_label}].rules[{i}]",
                    f"name={nr['name']}, src={nr['src']}, dst={nr['dst']}",
                    "Restrict business policy to required flows only.",
                    methodology_ref="[Net] Default Segment Behaviour",
                ))

    overrides = ctx.get("edgeOverrides") or {}
    if overrides.get("qos"):
        findings.append(_finding(
            ctx, "business_policy", "medium",
            "QoS/business policy override active on edge",
            "edgeOverrides.qos",
            json.dumps(overrides, ensure_ascii=False),
            "Verify business policy overrides do not bypass segmentation intent.",
            methodology_ref="[Net] Business Policy Override",
        ))

    if isinstance(wan_cfg, dict):
        for link in wan_cfg.get("links") or []:
            if not isinstance(link, dict):
                continue
            if link.get("encryptOverlay") is False:
                findings.append(_finding(
                    ctx, "business_policy", "high",
                    f"WAN link {link.get('name', '?')} has overlay encryption disabled",
                    f"WAN.links[{link.get('name')}].encryptOverlay",
                    json.dumps({"interface": link.get("interfaces"), "mode": link.get("mode")}, ensure_ascii=False),
                    "Enable overlay encryption on WAN links.",
                    methodology_ref="[Net] Overlay Traffic Encryption",
                ))


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 9 — PHASE 4 DEEP COLLECTORS
# ─────────────────────────────────────────────────────────────────────────────

def get_enterprise_route_table() -> dict:
    result = portal_rpc(
        "enterprise/getEnterpriseRouteTable",
        {"enterpriseId": int(VCO_ENTERPRISE_NUMERIC_ID)},
        request_id=70,
    )
    return result if isinstance(result, dict) else {}


def get_gateway_edge_assignments(gateway_id: int) -> list:
    result = portal_rpc(
        "gateway/getGatewayEdgeAssignments",
        {"gatewayId": int(gateway_id)},
        request_id=71,
    )
    if isinstance(result, list):
        return result
    if isinstance(result, dict):
        return result.get("data", [])
    return []


def _extract_gateways_from_services(services_payload) -> list:
    gateways = []

    def walk(obj):
        if isinstance(obj, dict):
            if (
                isinstance(obj.get("id"), int)
                and isinstance(obj.get("name"), str)
                and "gatewayState" in obj
            ):
                gateways.append(obj)
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for v in obj:
                walk(v)

    walk(services_payload)
    uniq = {g["id"]: g for g in gateways}
    return list(uniq.values())


def _route_entry_cidr(entry: dict) -> str:
    """Extract CIDR string from enterprise route table subnet entry."""
    if not isinstance(entry, dict):
        return ""
    learned = entry.get("learnedRoute") or {}
    if isinstance(learned, dict):
        cidr_ip = learned.get("cidrIp")
        cidr_prefix = learned.get("cidrPrefix")
        if cidr_ip is not None and cidr_prefix is not None:
            return f"{cidr_ip}/{cidr_prefix}"
    raw = entry.get("subnet") or entry.get("prefix") or ""
    return str(raw).split(":")[0].strip()


def _route_prefix_length(cidr: str) -> int:
    try:
        return ipaddress.ip_network(cidr, strict=False).prefixlen
    except ValueError:
        return -1


def _advertised_route_concern(cidr: str) -> tuple:
    """
    Return (is_concern, reason) for advertised routes.
    Flags default route (/0) and prefixes shorter than /16.
    """
    prefix_len = _route_prefix_length(cidr)
    if prefix_len < 0:
        return False, ""
    if prefix_len == 0:
        return True, "advertised default route (0.0.0.0/0)"
    if prefix_len < BROAD_CIDR_THRESHOLD:
        return True, f"advertised prefix /{prefix_len} is broader than /{BROAD_CIDR_THRESHOLD}"
    return False, ""


def _advertised_exits(entry: dict) -> list:
    """Return preferredExits with advertise=True."""
    exits = []
    for exit_info in entry.get("preferredExits") or []:
        if isinstance(exit_info, dict) and exit_info.get("advertise"):
            exits.append(exit_info)
    return exits


def analyse_route_table(route_data: dict, findings: list) -> None:
    """
    [Net] Route Table Review — partial automation.
    Only flags advertised /0 and advertised prefixes shorter than /16.
    Peers, BGP policy, and gateway design remain manual review.
    """
    ctx = _enterprise_ctx()
    if not isinstance(route_data, dict):
        return

    seen = set()
    subnets = route_data.get("subnets") or []
    if not isinstance(subnets, list):
        subnets = []

    for i, entry in enumerate(subnets):
        if not isinstance(entry, dict):
            continue
        advertised_exits = _advertised_exits(entry)
        if not advertised_exits:
            continue

        cidr = _route_entry_cidr(entry)
        if not cidr:
            continue

        concern, reason = _advertised_route_concern(cidr)
        if not concern:
            continue

        key = (cidr, reason)
        if key in seen:
            continue
        seen.add(key)

        sev = "high" if _route_prefix_length(cidr) == 0 else "medium"
        findings.append(_finding(
            ctx, "route_table", sev,
            f"Enterprise route table: {reason}",
            f"enterprise.subnets[{i}].preferredExits",
            json.dumps({
                "cidr": cidr,
                "reason": reason,
                "advertisedExits": len(advertised_exits),
                "exits": [
                    {
                        "edgeId": x.get("edgeId"),
                        "profileId": x.get("profileId"),
                        "segmentId": x.get("segmentId"),
                        "protocol": x.get("protocol"),
                        "entity": x.get("entity"),
                    }
                    for x in advertised_exits[:5]
                ],
            }, ensure_ascii=False),
            "Review whether this advertised reachability is intentional. "
            "Peers, BGP, and gateway pools require manual review.",
            methodology_ref="[Net] Route Table Review",
            automation="partial",
        ))

    # Legacy API shape: only flag when route dict indicates advertisement
    routes = route_data.get("routes") or route_data.get("data") or []
    if isinstance(routes, dict):
        routes = routes.get("routes", [])
    if isinstance(routes, list):
        for i, route in enumerate(routes):
            if not isinstance(route, dict):
                continue
            if not route.get("advertise") and not route.get("advertised"):
                continue
            prefix = str(route.get("prefix") or route.get("network") or route.get("cidr") or "")
            concern, reason = _advertised_route_concern(prefix)
            if concern and (prefix, reason) not in seen:
                seen.add((prefix, reason))
                sev = "high" if _route_prefix_length(prefix) == 0 else "medium"
                findings.append(_finding(
                    ctx, "route_table", sev,
                    f"Enterprise route table: {reason}",
                    f"enterprise.routes[{i}]",
                    json.dumps(route, ensure_ascii=False)[:200],
                    "Review whether this advertised reachability is intentional.",
                    methodology_ref="[Net] Route Table Review",
                    automation="partial",
                ))


def analyse_gateway_assignments(
    gateways: list,
    assignments_by_gateway: dict,
    scope_edge_names: set,
    findings: list,
) -> None:
    """[Net] Gateway Assignment Review — partial automation."""
    ctx = _enterprise_ctx()
    for gw in gateways:
        if not isinstance(gw, dict):
            continue
        gw_id = gw.get("id")
        gw_name = gw.get("name", str(gw_id))
        assigned = assignments_by_gateway.get(gw_id, [])
        assigned_names = {
            a.get("edgeName") or a.get("name")
            for a in assigned if isinstance(a, dict)
        }
        assigned_names.discard(None)
        in_scope = assigned_names & scope_edge_names
        if not assigned_names:
            findings.append(_finding(
                ctx, "gateway_review", "low",
                f"Gateway {gw_name} has no edge assignments",
                f"gateway[{gw_id}]",
                f"gatewayState={gw.get('gatewayState')}",
                "Review whether dormant gateways should be decommissioned.",
                methodology_ref="[Net] Gateway Assignment Review",
                automation="partial",
            ))
        elif in_scope:
            findings.append(_finding(
                ctx, "gateway_review", "info",
                f"Edges in scope assigned to gateway {gw_name}",
                f"gateway[{gw_id}].assignments",
                f"edges={sorted(in_scope)}",
                "Confirm gateway pool assignment matches topology design.",
                methodology_ref="[Net] Gateway Assignment Review",
                automation="partial",
            ))


_CERT_EXPIRY_WARN_DAYS = 30
_CERT_MAX_KEY_AGE_DAYS = 120
_CERT_ROTATION_GAP_DAYS = 120
_CERT_HA_REBUILD_TOLERANCE_DAYS = 7
_WEAK_VPN_CRYPTO_VALUES = {
    "des", "3des", "md5", "sha1", "rc4", "null", "none", "disabled", "psk",
}
_ACCEPTABLE_VPN_ENCRYPTION_PROTOCOLS = {
    "GROUP_IPSEC", "IPSEC", "AES256", "AES128", "AES-256", "AES-128",
}


def _parse_vco_timestamp(value) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None


def _sanitize_cert_records(certs: list) -> list:
    sanitized = []
    for cert in certs:
        if not isinstance(cert, dict):
            continue
        sanitized.append({k: v for k, v in cert.items() if k != "certificate"})
    return sanitized


def _walk_weak_vpn_crypto(obj, path: str = "") -> list:
    hits = []
    if isinstance(obj, dict):
        for key, val in obj.items():
            child = f"{path}.{key}" if path else key
            if isinstance(val, str) and val.lower() in _WEAK_VPN_CRYPTO_VALUES:
                hits.append((child, val))
            else:
                hits.extend(_walk_weak_vpn_crypto(val, child))
    elif isinstance(obj, list):
        for idx, val in enumerate(obj):
            hits.extend(_walk_weak_vpn_crypto(val, f"{path}[{idx}]"))
    return hits


def _certs_by_serial(certs: list) -> dict:
    grouped = {}
    for cert in certs:
        if not isinstance(cert, dict):
            continue
        serial = cert.get("edgeSerialNumber") or "unknown"
        grouped.setdefault(serial, []).append(cert)
    return grouped


def _latest_cert(certs: list) -> dict | None:
    if not certs:
        return None
    return max(
        certs,
        key=lambda c: _parse_vco_timestamp(c.get("validTo"))
        or datetime.min.replace(tzinfo=timezone.utc),
    )


def analyse_vpn_encryption_strength(
    ctx: dict,
    control_plane_cfg: dict,
    device_settings_cfg: dict,
    findings: list,
) -> None:
    """[VPN] Encryption Strength — controlPlane VPN protocol + weak crypto scan."""
    if isinstance(control_plane_cfg, dict):
        for field_path, val in _walk_weak_vpn_crypto(
            control_plane_cfg.get("vpn") or {}, "controlPlane.vpn",
        ):
            findings.append(_finding(
                ctx, "vpn_crypto", "high",
                f"Weak VPN crypto setting: {field_path}={val}",
                field_path,
                f"{field_path}={val}",
                "Upgrade to modern cipher suites and hashing algorithms.",
                methodology_ref="[VPN] Encryption Strength",
            ))

        for idx, seg in enumerate(control_plane_cfg.get("segments") or []):
            if not isinstance(seg, dict):
                continue
            vpn = seg.get("vpn") or {}
            if not vpn.get("enabled") or not vpn.get("edgeToEdge"):
                continue
            seg_name = seg.get("name") or seg.get("segmentLogicalId") or f"segment[{idx}]"
            proto = (vpn.get("edgeToEdgeDetail") or {}).get("encryptionProtocol")
            if not proto:
                findings.append(_finding(
                    ctx, "vpn_crypto", "low",
                    f"Branch-to-branch VPN enabled without declared encryption protocol ({seg_name})",
                    f"controlPlane.segments[{seg_name}].vpn.edgeToEdgeDetail.encryptionProtocol",
                    json.dumps({"segment": seg_name, "edgeToEdge": True}, ensure_ascii=False),
                    "Confirm overlay VPN uses strong encryption (e.g. GROUP_IPSEC).",
                    methodology_ref="[VPN] Encryption Strength",
                ))
            elif str(proto).upper() not in _ACCEPTABLE_VPN_ENCRYPTION_PROTOCOLS:
                sev = "high" if str(proto).lower() in _WEAK_VPN_CRYPTO_VALUES else "medium"
                findings.append(_finding(
                    ctx, "vpn_crypto", sev,
                    f"Non-standard VPN encryption protocol on segment {seg_name}: {proto}",
                    f"controlPlane.segments[{seg_name}].vpn.edgeToEdgeDetail.encryptionProtocol",
                    f"encryptionProtocol={proto}",
                    "Use VeloCloud default strong overlay encryption (GROUP_IPSEC).",
                    methodology_ref="[VPN] Encryption Strength",
                ))

    if isinstance(device_settings_cfg, dict):
        for idx, iface in enumerate(device_settings_cfg.get("routedInterfaces") or []):
            if not isinstance(iface, dict):
                continue
            if iface.get("edgeToEdgeEncryption") is False:
                findings.append(_finding(
                    ctx, "vpn_crypto", "high",
                    f"Edge-to-edge encryption disabled on interface {iface.get('name', idx)}",
                    f"deviceSettings.routedInterfaces[{idx}].edgeToEdgeEncryption",
                    json.dumps(
                        {"interface": iface.get("name"), "edgeToEdgeEncryption": False},
                        ensure_ascii=False,
                    ),
                    "Enable edge-to-edge encryption on routed interfaces.",
                    methodology_ref="[VPN] Encryption Strength",
                ))


def analyse_vpn_certificate_validation(ctx: dict, certs: list, findings: list) -> None:
    """[VPN] Certificate Validation — edge/getEdgeCertificates presence and expiry."""
    sanitized = [c for c in (certs or []) if isinstance(c, dict)]
    if not sanitized:
        findings.append(_finding(
            ctx, "vpn_crypto", "high",
            "No edge certificates found (certificate-based VPN auth not evidenced)",
            "edge/getEdgeCertificates",
            "empty certificate list",
            "Ensure edges use VCO-issued certificates for overlay authentication.",
            methodology_ref="[VPN] Certificate Validation",
        ))
        return

    now = datetime.now(timezone.utc)
    for serial, serial_certs in _certs_by_serial(sanitized).items():
        latest = _latest_cert(serial_certs)
        if not latest:
            continue
        valid_to = _parse_vco_timestamp(latest.get("validTo"))
        if not valid_to:
            findings.append(_finding(
                ctx, "vpn_crypto", "medium",
                f"Edge certificate for serial {serial} missing validTo metadata",
                "edge/getEdgeCertificates",
                json.dumps(_sanitize_cert_records([latest])[0], ensure_ascii=False),
                "Verify certificate validity in Orchestrator certificate management.",
                methodology_ref="[VPN] Certificate Validation",
            ))
            continue

        days_left = (valid_to - now).days
        if days_left < 0:
            findings.append(_finding(
                ctx, "vpn_crypto", "high",
                f"Edge certificate expired for serial {serial}",
                "edge/getEdgeCertificates",
                json.dumps({
                    "edgeSerialNumber": serial,
                    "validFrom": latest.get("validFrom"),
                    "validTo": latest.get("validTo"),
                    "daysExpired": abs(days_left),
                }, ensure_ascii=False),
                "Renew or redeploy edge certificates before overlay authentication fails.",
                methodology_ref="[VPN] Certificate Validation",
            ))
        elif days_left <= _CERT_EXPIRY_WARN_DAYS:
            findings.append(_finding(
                ctx, "vpn_crypto", "medium",
                f"Edge certificate expiring within {_CERT_EXPIRY_WARN_DAYS} days for serial {serial}",
                "edge/getEdgeCertificates",
                json.dumps({
                    "edgeSerialNumber": serial,
                    "validTo": latest.get("validTo"),
                    "daysRemaining": days_left,
                }, ensure_ascii=False),
                "Confirm automatic certificate renewal is scheduled.",
                methodology_ref="[VPN] Certificate Validation",
            ))
        elif not latest.get("authorityKeyId"):
            findings.append(_finding(
                ctx, "vpn_crypto", "low",
                f"Edge certificate for serial {serial} lacks authorityKeyId metadata",
                "edge/getEdgeCertificates",
                json.dumps(_sanitize_cert_records([latest])[0], ensure_ascii=False),
                "Confirm certificate is issued by the enterprise CA.",
                methodology_ref="[VPN] Certificate Validation",
            ))


def _ha_cert_context(ctx: dict) -> tuple:
    """Return (is_enhanced_ha, set of physical serial numbers for the HA pair)."""
    portal = ctx.get("edgePortalRecord") or {}
    ha = portal.get("ha") or {}
    ha_type = str(ha.get("type", "")).upper()
    serials = set()
    for key in ("serialNumber", "haSerialNumber"):
        val = portal.get(key)
        if val:
            serials.add(str(val))
    is_enhanced = ha_type in {"ACTIVE_STANDBY", "ACTIVE_ACTIVE"} and len(serials) >= 2
    return is_enhanced, serials


def _ordered_certs(certs: list) -> list:
    return sorted(
        certs,
        key=lambda c: _parse_vco_timestamp(c.get("validFrom"))
        or datetime.min.replace(tzinfo=timezone.utc),
    )


def _current_rotation_cluster(ordered: list) -> list:
    """
    Certificates in the current renewal cluster — everything after the last
    major gap (> _CERT_ROTATION_GAP_DAYS) in the serial's history.
    """
    if not ordered:
        return []
    cluster_start = 0
    for idx, (prev, curr) in enumerate(zip(ordered, ordered[1:])):
        prev_to = _parse_vco_timestamp(prev.get("validTo"))
        curr_from = _parse_vco_timestamp(curr.get("validFrom"))
        if prev_to and curr_from and (curr_from - prev_to).days > _CERT_ROTATION_GAP_DAYS:
            cluster_start = idx + 1
    return ordered[cluster_start:]


def _cluster_start_time(ordered: list) -> datetime | None:
    cluster = _current_rotation_cluster(ordered)
    if not cluster:
        return None
    return _parse_vco_timestamp(cluster[0].get("validFrom"))


def _ha_coordinated_rebuild(certs_by_serial: dict, ha_serials: set) -> bool:
    """
    True when all enhanced-HA member serials began their current renewal cluster
    within a short window (typical pair re-enrollment / rebuild event).
    """
    starts = []
    for serial in ha_serials:
        serial_certs = certs_by_serial.get(serial)
        if not serial_certs:
            return False
        start = _cluster_start_time(_ordered_certs(serial_certs))
        if not start:
            return False
        starts.append(start)
    if len(starts) < 2:
        return False
    return (max(starts) - min(starts)).days <= _CERT_HA_REBUILD_TOLERANCE_DAYS


def _cert_rotation_gaps(ordered: list) -> list:
    gaps = []
    for prev, curr in zip(ordered, ordered[1:]):
        prev_to = _parse_vco_timestamp(prev.get("validTo"))
        curr_from = _parse_vco_timestamp(curr.get("validFrom"))
        if prev_to and curr_from:
            gaps.append({
                "gapDays": (curr_from - prev_to).days,
                "previousValidTo": prev.get("validTo"),
                "nextValidFrom": curr.get("validFrom"),
            })
    return gaps


def _latest_cert_health(latest: dict, now: datetime) -> dict:
    valid_to = _parse_vco_timestamp(latest.get("validTo"))
    valid_from = _parse_vco_timestamp(latest.get("validFrom"))
    expired = bool(valid_to and valid_to < now)
    key_age_days = (now - valid_from).days if valid_from else None
    stale = bool(key_age_days is not None and key_age_days > _CERT_MAX_KEY_AGE_DAYS)
    return {
        "expired": expired,
        "stale": stale,
        "unhealthy": expired or stale,
        "keyAgeDays": key_age_days,
    }


def analyse_vpn_key_rotation(ctx: dict, certs: list, findings: list) -> None:
    """
    [VPN] Key Rotation — certificate renewal history from edge/getEdgeCertificates.

    Flags current rotation problems only:
      • latest cert stale (>120 days) or expired
      • gaps >120 days within the current renewal cluster (not historical)
    Enhanced HA: suppresses historical pair-rebuild gaps when member serials
    re-enrolled together; still flags unhealthy latest material.
    """
    sanitized = [c for c in (certs or []) if isinstance(c, dict)]
    if not sanitized:
        return

    now = datetime.now(timezone.utc)
    certs_by_serial = _certs_by_serial(sanitized)
    ha_enhanced, ha_serials = _ha_cert_context(ctx)
    ha_rebuild = (
        _ha_coordinated_rebuild(certs_by_serial, ha_serials)
        if ha_enhanced else False
    )

    for serial, serial_certs in certs_by_serial.items():
        ordered = _ordered_certs(serial_certs)
        latest = _latest_cert(ordered)
        if not latest:
            continue

        health = _latest_cert_health(latest, now)
        is_ha_member = ha_enhanced and serial in ha_serials

        if len(ordered) < 2:
            if health["unhealthy"]:
                findings.append(_finding(
                    ctx, "vpn_crypto", "medium",
                    f"No certificate rotation history and latest cert unhealthy for serial {serial}",
                    "edge/getEdgeCertificates",
                    json.dumps({
                        "edgeSerialNumber": serial,
                        "certificateCount": len(ordered),
                        "validTo": latest.get("validTo"),
                        "keyAgeDays": health.get("keyAgeDays"),
                    }, ensure_ascii=False),
                    "Verify VeloCloud automatic certificate renewal is functioning.",
                    methodology_ref="[VPN] Key Rotation",
                ))
            continue

        if not health["unhealthy"]:
            continue

        if health["stale"]:
            findings.append(_finding(
                ctx, "vpn_crypto", "high",
                f"Edge certificate/key material older than {_CERT_MAX_KEY_AGE_DAYS} days "
                f"for serial {serial}",
                "edge/getEdgeCertificates",
                json.dumps({
                    "edgeSerialNumber": serial,
                    "validFrom": latest.get("validFrom"),
                    "keyAgeDays": health["keyAgeDays"],
                }, ensure_ascii=False),
                "Trigger certificate renewal or investigate failed auto-rotation.",
                methodology_ref="[VPN] Key Rotation",
            ))

        cluster = _current_rotation_cluster(ordered)
        had_historical_rebuild = len(cluster) < len(ordered)

        if is_ha_member and ha_rebuild and had_historical_rebuild:
            continue

        gap_found = False
        for gap in _cert_rotation_gaps(cluster):
            if gap["gapDays"] > _CERT_ROTATION_GAP_DAYS:
                findings.append(_finding(
                    ctx, "vpn_crypto", "medium",
                    f"Rotation gap ({gap['gapDays']} days) in current renewal period for serial {serial}",
                    "edge/getEdgeCertificates",
                    json.dumps({
                        "edgeSerialNumber": serial,
                        "previousValidTo": gap["previousValidTo"],
                        "nextValidFrom": gap["nextValidFrom"],
                        "gapDays": gap["gapDays"],
                        "clusterSize": len(cluster),
                    }, ensure_ascii=False),
                    "Investigate delayed certificate renewal in the active rotation window.",
                    methodology_ref="[VPN] Key Rotation",
                ))
                gap_found = True
                break
        if not gap_found and len(cluster) < 2 and health["unhealthy"]:
            findings.append(_finding(
                ctx, "vpn_crypto", "medium",
                f"No certificate renewals in current rotation period for serial {serial}",
                "edge/getEdgeCertificates",
                json.dumps({
                    "edgeSerialNumber": serial,
                    "clusterStart": cluster[0].get("validFrom") if cluster else None,
                    "clusterSize": len(cluster),
                    "validTo": latest.get("validTo"),
                }, ensure_ascii=False),
                "Confirm automatic certificate renewal is active for this edge.",
                methodology_ref="[VPN] Key Rotation",
            ))


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 10 — ENTERPRISE MANAGEMENT (users, API tokens, auth mode)
# ─────────────────────────────────────────────────────────────────────────────

def _sanitize_enterprise_user(user: dict) -> dict:
    if not isinstance(user, dict):
        return {}
    return {k: v for k, v in user.items() if k not in {"password", "salt"}}


def _sanitize_api_token_record(token: dict) -> dict:
    if not isinstance(token, dict):
        return {}
    return {k: v for k, v in token.items() if k.lower() != "token"}


def _api_token_active(token: dict) -> bool:
    if not isinstance(token, dict):
        return False
    if token.get("isRevoked"):
        return False
    return str(token.get("state", "")).upper() not in {"REVOKED", "DISABLED"}


def analyse_token_enumeration(ctx: dict, tokens: list, findings: list) -> None:
    """
    [Mgmt] Token Enumeration — inventory via enterprise/getApiTokens.
    Emits assisted inventory plus flags for expired or non-expiring active tokens.
    """
    token_list = [t for t in (tokens or []) if isinstance(t, dict)]
    active = [t for t in token_list if _api_token_active(t)]
    now = datetime.now(timezone.utc)

    inventory = [
        {
            "name": t.get("name"),
            "state": t.get("state"),
            "expiration": t.get("expiration"),
            "created": t.get("created"),
            "downloaded": t.get("downloaded"),
            "createdForEnterpriseUserId": t.get("createdForEnterpriseUserId"),
        }
        for t in token_list
    ]
    findings.append(_finding(
        ctx, "mgmt_tokens", "info",
        f"Enterprise API token inventory: {len(active)} active, {len(token_list)} total",
        "enterprise/getApiTokens",
        json.dumps(inventory, ensure_ascii=False),
        "Review each API token for least privilege, owner, and expiry.",
        methodology_ref="[Mgmt] Token Enumeration",
        automation="partial",
    ))

    for token in token_list:
        if not _api_token_active(token):
            continue
        name = token.get("name") or token.get("tokenUuid") or str(token.get("id"))
        expiration = _parse_vco_timestamp(token.get("expiration"))
        if not expiration:
            findings.append(_finding(
                ctx, "mgmt_tokens", "medium",
                f"Active API token '{name}' has no expiration date",
                f"enterprise/getApiTokens[{token.get('id')}].expiration",
                json.dumps(_sanitize_api_token_record(token), ensure_ascii=False),
                "Set an expiry on API tokens and rotate regularly.",
                methodology_ref="[Mgmt] Token Enumeration",
                automation="partial",
            ))
        elif expiration < now:
            findings.append(_finding(
                ctx, "mgmt_tokens", "high",
                f"Active API token '{name}' is past expiration",
                f"enterprise/getApiTokens[{token.get('id')}].expiration",
                json.dumps({
                    "name": name,
                    "expiration": token.get("expiration"),
                    "state": token.get("state"),
                }, ensure_ascii=False),
                "Revoke expired API tokens.",
                methodology_ref="[Mgmt] Token Enumeration",
                automation="partial",
            ))


def analyse_enterprise_auth_mode(
    ctx: dict, enterprise: dict, users: list, findings: list,
) -> None:
    """
    [Mgmt] Enterprise Authentication Mode — infer native vs federated from
    enterprise domain and user isNative flags (SAML/LDAP detail not in API).
    """
    user_list = [u for u in (users or []) if isinstance(u, dict)]
    if not user_list:
        findings.append(_finding(
            ctx, "mgmt_auth", "medium",
            "No enterprise users returned from Orchestrator API",
            "enterprise/getEnterpriseUsers",
            "empty user list",
            "Verify API token can read enterprise user configuration.",
            methodology_ref="[Mgmt] Enterprise Authentication Mode",
            automation="partial",
        ))
        return

    domain = (enterprise or {}).get("domain")
    pki_mode = (enterprise or {}).get("endpointPkiMode")
    native_count = sum(1 for u in user_list if u.get("isNative"))
    federated_count = len(user_list) - native_count

    if domain or federated_count > 0:
        if native_count > 0:
            mode = "MIXED"
        else:
            mode = "FEDERATED"
    else:
        mode = "NATIVE"

    evidence = {
        "detectedMode": mode,
        "domain": domain,
        "endpointPkiMode": pki_mode,
        "nativeUsers": native_count,
        "federatedUsers": federated_count,
        "totalUsers": len(user_list),
    }

    if mode == "NATIVE":
        findings.append(_finding(
            ctx, "mgmt_auth", "low",
            "Enterprise authentication mode is native VeloCloud (no IdP domain configured)",
            "enterprise/getEnterprise.domain",
            json.dumps(evidence, ensure_ascii=False),
            "Confirm native authentication is acceptable or migrate to SAML/LDAP.",
            methodology_ref="[Mgmt] Enterprise Authentication Mode",
            automation="partial",
        ))
    elif mode == "MIXED":
        findings.append(_finding(
            ctx, "mgmt_auth", "low",
            "Enterprise has mixed native and federated user accounts",
            "enterprise/getEnterpriseUsers.isNative",
            json.dumps(evidence, ensure_ascii=False),
            "Review whether native accounts are still required.",
            methodology_ref="[Mgmt] Enterprise Authentication Mode",
            automation="partial",
        ))


def analyse_dormant_users(ctx: dict, users: list, findings: list) -> None:
    """[Mgmt] Dormant User Account Review — lastLogin age and inactive accounts."""
    now = datetime.now(timezone.utc)
    for user in users or []:
        if not isinstance(user, dict):
            continue
        username = user.get("username") or user.get("email") or str(user.get("id"))

        if not user.get("isActive"):
            findings.append(_finding(
                ctx, "mgmt_users", "low",
                f"Inactive enterprise user account still present: {username}",
                f"enterprise/getEnterpriseUsers[{user.get('id')}].isActive",
                json.dumps(_sanitize_enterprise_user(user), ensure_ascii=False)[:300],
                "Remove or formally decommission dormant accounts.",
                methodology_ref="[Mgmt] Dormant User Account Review",
            ))
            continue

        if user.get("isServiceAccount"):
            continue

        last_login = _parse_vco_timestamp(user.get("lastLogin"))
        if not last_login:
            findings.append(_finding(
                ctx, "mgmt_users", "medium",
                f"Active enterprise user has never logged in: {username}",
                f"enterprise/getEnterpriseUsers[{user.get('id')}].lastLogin",
                json.dumps({
                    "username": username,
                    "created": user.get("created"),
                    "roleName": user.get("roleName"),
                    "accessLevel": user.get("accessLevel"),
                }, ensure_ascii=False),
                "Confirm account is required or disable until onboarded.",
                methodology_ref="[Mgmt] Dormant User Account Review",
            ))
            continue

        idle_days = (now - last_login).days
        if idle_days >= DORMANT_USER_DAYS:
            severity = "high" if idle_days >= 180 else "medium"
            findings.append(_finding(
                ctx, "mgmt_users", severity,
                f"Enterprise user dormant {idle_days} days: {username}",
                f"enterprise/getEnterpriseUsers[{user.get('id')}].lastLogin",
                json.dumps({
                    "username": username,
                    "lastLogin": user.get("lastLogin"),
                    "idleDays": idle_days,
                    "roleName": user.get("roleName"),
                    "accessLevel": user.get("accessLevel"),
                }, ensure_ascii=False),
                f"Review account necessity (threshold {DORMANT_USER_DAYS} days).",
                methodology_ref="[Mgmt] Dormant User Account Review",
            ))


# ─────────────────────────────────────────────────────────────────────────────
# MODULE 11 — ISOLATION (shared resources, wireless security)
# ─────────────────────────────────────────────────────────────────────────────

_WIFI_MODULE_HINTS = ("wifi", "wlan", "wireless", "radio")
_WEAK_WIFI_MODES = {
    "open", "wpa", "wpa-psk", "wpa2", "wpa2-psk", "wpa3-psk", "psk", "personal", "none",
}
_STRONG_WIFI_HINTS = ("enterprise", "802.1x", "wpa2-enterprise", "wpa-enterprise", "wpa3-enterprise")
_NO_WIFI_MODEL_SUFFIXES = ("-n", "-nw")
_WIFI_CAPABLE_MODEL_MARKERS = ("710", "7105g", "710-w", "710w", "-lte", "lte", "wifi", "wlan")


def _edge_has_wifi_hardware(edge_portal: dict) -> bool:
    """True only for edge models that can actually serve integrated Wi-Fi."""
    if not isinstance(edge_portal, dict):
        return False
    model = (edge_portal.get("modelNumber") or "").lower().strip()
    if not model:
        return False
    if any(model.endswith(suffix) for suffix in _NO_WIFI_MODEL_SUFFIXES):
        return False
    if model in {"edge3800", "edge3810", "edge3400", "edge2000"}:
        return False
    if any(marker in model for marker in _WIFI_CAPABLE_MODEL_MARKERS):
        return True
    # Non-N 5x0/6x0 branch models may ship with integrated Wi-Fi (610/620/640/680 without -N).
    if any(token in model for token in ("edge510", "edge520", "edge540", "edge610", "edge620", "edge640", "edge680")):
        return True
    return False


def _is_wifi_enabled(cfg: dict) -> bool:
    if not isinstance(cfg, dict):
        return False
    disabled = cfg.get("disabled")
    if disabled is True or str(disabled).lower() in {"1", "true", "yes"}:
        return False
    enabled = cfg.get("enabled", cfg.get("isEnabled"))
    if enabled is False or str(enabled).lower() in {"0", "false", "disabled"}:
        return False
    if enabled in (True, 1) or str(enabled).lower() in {"1", "true", "enabled"}:
        return True
    if disabled is False or str(disabled).lower() in {"0", "false", "no"}:
        return True
    return False


def _wifi_security_mode(cfg: dict) -> str:
    if not isinstance(cfg, dict):
        return ""
    for key in ("securityMode", "security", "authentication", "authMode", "encryption", "wpaMode"):
        val = cfg.get(key)
        if val is not None and str(val).strip():
            return str(val).strip().lower().replace("_", "-")
    return ""


def _wifi_uses_radius(cfg: dict) -> bool:
    if not isinstance(cfg, dict):
        return False
    for key in ("radiusAuthentication", "useRadius", "radiusEnabled"):
        val = cfg.get(key)
        if val in (True, 1) or str(val).lower() in {"1", "true", "enabled"}:
            return True
    if cfg.get("radius") or cfg.get("radiusServers"):
        return True
    return False


def _collect_wifi_configs(obj, path: str = "", results: list = None) -> list:
    """Return (path, config_dict) tuples for Wi-Fi SSID/radio settings."""
    results = results if results is not None else []
    if isinstance(obj, dict):
        if _is_wireless_interface(obj):
            results.append((path, obj))
            return results
        for key, val in obj.items():
            if key.lower() in _WIFI_MODULE_HINTS or key.lower() in {"ssids", "radios"}:
                _collect_wifi_configs(val, f"{path}.{key}" if path else key, results)
    elif isinstance(obj, list):
        for idx, val in enumerate(obj):
            if isinstance(val, dict) and _is_wireless_interface(val):
                results.append((f"{path}[{idx}]", val))
            else:
                _collect_wifi_configs(val, f"{path}[{idx}]", results)
    return results


def _profile_lan_interfaces_for_edge(profile_cfg: dict, edge_portal: dict) -> list:
    """Return profile LAN interfaces for the edge hardware model only."""
    if not _edge_has_wifi_hardware(edge_portal):
        return []
    models = profile_cfg.get("models") or {}
    family = edge_portal.get("deviceFamily", "")
    model = (edge_portal.get("modelNumber") or "").lower()
    if "-lte" in model or model.endswith("lte"):
        candidates = ["edge610lte", "edge6X0"]
    elif "6x0" in family.lower() or "610" in model:
        candidates = ["edge6X0", "edge610lte"]
    elif "3x10" in model or "3810" in model:
        candidates = ["edge3X10", "edge3X00"]
    elif "3x00" in family.lower() or "3800" in model:
        candidates = ["edge3X00", "edge3X10"]
    elif "500" in model or "510" in model:
        candidates = ["edge500", "edge5X0"]
    elif any(marker in model for marker in ("710", "7105g")):
        candidates = ["edge710", "edge7105g", "edge7X0", "edge6X0"]
    else:
        candidates = [k for k, v in models.items() if isinstance(v, dict) and "lan" in v]
    for candidate in candidates:
        interfaces = get_nested(models, candidate, "lan", "interfaces")
        if isinstance(interfaces, list):
            return interfaces
    return []


def _is_wireless_interface(cfg: dict) -> bool:
    if not isinstance(cfg, dict):
        return False
    if str(cfg.get("type", "")).lower() in {"wifi", "wlan", "wireless"}:
        return True
    if cfg.get("ssid") and (
        cfg.get("securityMode") or cfg.get("security") or cfg.get("authentication")
    ):
        return True
    return False


def _walk_config_roots(
    profile_cfg: dict,
    edge_cfg: dict,
    edge_modules: dict,
    profile_modules: dict,
    edge_portal: dict,
) -> list:
    roots = []
    if isinstance(edge_cfg, dict):
        roots.append(("edgeConfig", edge_cfg))
    for ifc in get_nested(edge_cfg or {}, "lan", "interfaces") or []:
        if _is_wireless_interface(ifc):
            roots.append((f"edgeConfig.lan.interfaces[{ifc.get('name', '?')}]", ifc))
    for ifc in _profile_lan_interfaces_for_edge(profile_cfg or {}, edge_portal or {}):
        if _is_wireless_interface(ifc):
            roots.append((f"profileConfig.lan.interfaces[{ifc.get('name', '?')}]", ifc))
    for container, label in ((edge_modules, "edgeModules"), (profile_modules, "profileModules")):
        if not isinstance(container, dict):
            continue
        for mod_name in container:
            if any(hint in mod_name.lower() for hint in _WIFI_MODULE_HINTS):
                data = _module_data(container, mod_name)
                if data:
                    roots.append((f"{label}.{mod_name}", data))
    return roots


def analyse_wireless_security(
    ctx: dict,
    profile_cfg: dict,
    edge_cfg: dict,
    edge_modules: dict,
    profile_modules: dict,
    findings: list,
) -> None:
    """
    [Isolation] Wireless Security — only evaluates edges with Wi-Fi hardware and
    actively enabled WLAN interfaces. Flags weak modes (PSK/open) and missing RADIUS.
    """
    edge_portal = ctx.get("edgePortalRecord") or {}
    if not _edge_has_wifi_hardware(edge_portal):
        return

    wifi_configs = []
    for root_path, root_obj in _walk_config_roots(
        profile_cfg, edge_cfg, edge_modules or {}, profile_modules or {}, edge_portal,
    ):
        for path, cfg in _collect_wifi_configs(root_obj, root_path):
            if _is_wifi_enabled(cfg):
                wifi_configs.append((path, cfg))

    seen_ssids = set()
    for path, cfg in wifi_configs:
        ssid = cfg.get("ssid") or cfg.get("name") or path
        if ssid in seen_ssids:
            continue
        seen_ssids.add(ssid)

        mode = _wifi_security_mode(cfg)
        has_radius = _wifi_uses_radius(cfg)
        if not mode:
            findings.append(_finding(
                ctx, "wireless_security", "medium",
                f"Wireless enabled but security mode not declared for SSID '{ssid}'",
                path,
                json.dumps({"ssid": ssid, "path": path}, ensure_ascii=False),
                "Confirm WPA-Enterprise with RADIUS AAA is configured for wireless.",
                methodology_ref="[Isolation] Wireless Security",
                automation="partial",
            ))
            continue

        if mode in _WEAK_WIFI_MODES or "psk" in mode or "personal" in mode or mode == "open":
            findings.append(_finding(
                ctx, "wireless_security", "high",
                f"Wireless SSID '{ssid}' uses weak security mode: {mode}",
                path,
                json.dumps({"ssid": ssid, "securityMode": mode}, ensure_ascii=False),
                "Use WPA-Enterprise with RADIUS AAA per wireless hardening guidance.",
                methodology_ref="[Isolation] Wireless Security",
                automation="partial",
            ))
            continue

        if not any(hint in mode for hint in _STRONG_WIFI_HINTS):
            findings.append(_finding(
                ctx, "wireless_security", "medium",
                f"Wireless SSID '{ssid}' does not use WPA-Enterprise: {mode}",
                path,
                json.dumps({"ssid": ssid, "securityMode": mode}, ensure_ascii=False),
                "Configure WPA-Enterprise with RADIUS for wireless authentication.",
                methodology_ref="[Isolation] Wireless Security",
                automation="partial",
            ))
            continue

        if not has_radius:
            findings.append(_finding(
                ctx, "wireless_security", "medium",
                f"Wireless SSID '{ssid}' lacks RADIUS AAA configuration",
                path,
                json.dumps({"ssid": ssid, "securityMode": mode}, ensure_ascii=False),
                "Bind wireless authentication to enterprise RADIUS/AAA servers.",
                methodology_ref="[Isolation] Wireless Security",
                automation="partial",
            ))


def analyse_shared_resources(
    ctx: dict,
    enterprise_record: dict,
    gateways: list,
    assignments_by_gateway: dict,
    scope_edge_names: set,
    findings: list,
) -> None:
    """
    [Isolation] Shared Resources — enterprise gateway pool and gateway assignment inventory.
    Complements [Net] Gateway Assignment Review with enterprise-level shared-resource context.
    """
    enterprise = enterprise_record if isinstance(enterprise_record, dict) else {}
    gateway_list = [g for g in (gateways or []) if isinstance(g, dict)]
    pool_id = enterprise.get("gatewayPoolId")
    network_id = enterprise.get("networkId")

    gateway_summary = [
        {
            "id": g.get("id"),
            "name": g.get("name"),
            "gatewayState": g.get("gatewayState"),
            "assignedEdgeCount": len(assignments_by_gateway.get(g.get("id"), []) or []),
        }
        for g in gateway_list
    ]
    findings.append(_finding(
        ctx, "shared_resources", "info",
        f"Enterprise shared resources: gatewayPoolId={pool_id}, gateways={len(gateway_list)}",
        "enterprise/getEnterprise",
        json.dumps({
            "gatewayPoolId": pool_id,
            "networkId": network_id,
            "gateways": gateway_summary,
        }, ensure_ascii=False),
        "Review gateway pool and shared service assignments for cross-topology exposure.",
        methodology_ref="[Isolation] Shared Resources",
        automation="partial",
    ))

    if not gateway_list:
        return

    if len(gateway_list) > 1:
        findings.append(_finding(
            ctx, "shared_resources", "low",
            f"Enterprise has {len(gateway_list)} gateways in shared pool {pool_id}",
            "enterprise/getEnterpriseServices",
            json.dumps(gateway_summary, ensure_ascii=False),
            "Confirm multiple gateways in the pool are intentional and segmented appropriately.",
            methodology_ref="[Isolation] Shared Resources",
            automation="partial",
        ))

    for gw in gateway_list:
        gw_id = gw.get("id")
        gw_name = gw.get("name", str(gw_id))
        assigned = assignments_by_gateway.get(gw_id, []) or []
        assigned_names = {
            a.get("edgeName") or a.get("name")
            for a in assigned if isinstance(a, dict)
        }
        assigned_names.discard(None)
        out_of_scope = sorted(assigned_names - scope_edge_names)
        if out_of_scope:
            findings.append(_finding(
                ctx, "shared_resources", "medium",
                f"Gateway {gw_name} has assignments outside review scope",
                f"gateway[{gw_id}].assignments",
                json.dumps({"gateway": gw_name, "outOfScopeEdges": out_of_scope}, ensure_ascii=False),
                "Verify shared gateway assignments are limited to intended edges/tenants.",
                methodology_ref="[Isolation] Shared Resources",
                automation="partial",
            ))


# ─────────────────────────────────────────────────────────────────────────────
# METHODOLOGY — XLSX loader, coverage, manual checklist, Dradis export
# ─────────────────────────────────────────────────────────────────────────────

_AUTOMATION_TIER_BY_TITLE = {
    "[Mgmt] MFA Enforcement": "manual",
    "[Mgmt] RBAC Least Privilege": "manual",
    "[Mgmt] API Token Expiry": "manual",
    "[Mgmt] API Token Privileges": "manual",
    "[Mgmt] Orchestrator Exposure": "manual",
    "[Mgmt] Token Enumeration": "partial",
    "[Mgmt] Enterprise Authentication Mode": "partial",
    "[Mgmt] Dormant User Account Review": "automated",
    "[Net] Overlay to Management Access": "partial",
    "[Net] Gateway Assignment Review": "partial",
    "[Net] Route Table Review": "partial",
    "[Net] Segment Isolation": "automated",
    "[Net] Default Segment Behaviour": "automated",
    "[Net] Business Policy Override": "automated",
    "[Net] Edge-to-Edge Communication": "automated",
    "[Net] Overlay Traffic Encryption": "automated",
    "[Net] DNS Configuration": "automated",
    "[FW] Default Deny": "automated",
    "[FW] Rule Scope": "automated",
    "[FW] NAT Exposure": "automated",
    "[FW] Segment Enforcement": "partial",
    "[FW] Edge Local Access Restrictions": "automated",
    "[FW] Firewall Event Logging": "automated",
    "[VPN] Encryption Strength": "automated",
    "[VPN] Certificate Validation": "automated",
    "[VPN] Key Rotation": "automated",
    "[Monitoring] Central Logging": "automated",
    "[Monitoring] Admin Activity Logs": "partial",
    "[Monitoring] Event Review": "partial",
    "[Monitoring] Config Change Logs": "partial",
    "[Monitoring] Critical Event Alerts": "manual",
    "[Monitoring] Firewall Log Collection": "automated",
    "[System] Edge Versions": "automated",
    "[System] Patch Levels": "automated",
    "[System] Config Consistency Real": "partial",
    "[System] Edge Config Stack Review": "partial",
    "[System] Edge Certificate Authentication": "partial",
    "[System] NTP Time Synchronisation": "automated",
    "[System] Edge Override Governance": "partial",
    "[System] Edge Admin AAA (TACACS/RADIUS)": "automated",
    "[System] SNMP Hardening": "automated",
    "[System] BFD Link Detection": "automated",
    "[System] High Availability Configuration": "automated",
    "[System] Inactive Edge Review": "automated",
    "[Isolation] Tenant Separation": "manual",
    "[Isolation] Shared Resources": "partial",
    "[Isolation] Profile Inheritance": "partial",
    "[Isolation] Wireless Security": "partial",
    "[FW] Advanced Security Features (ATP/IDPS/URL)": "partial",
}

_SCRIPT_MODULE_BY_TITLE = {
    "[Net] Gateway Assignment Review": "analyse_gateway_assignments",
    "[Net] Route Table Review": "analyse_route_table",
    "[Net] Segment Isolation": "analyse_segment_isolation",
    "[Net] Default Segment Behaviour": "analyse_business_policy",
    "[Net] Business Policy Override": "analyse_business_policy",
    "[Net] Edge-to-Edge Communication": "analyse_edge_to_edge_communication",
    "[Net] Overlay to Management Access": "analyse_overlay_management_access",
    "[Net] Overlay Traffic Encryption": "analyse_device_hardening, analyse_business_policy",
    "[Net] DNS Configuration": "analyse_dns",
    "[FW] Default Deny": "analyse_firewall_rules",
    "[FW] Rule Scope": "analyse_firewall_rules",
    "[FW] NAT Exposure": "analyse_nat_exposure",
    "[FW] Segment Enforcement": "analyse_segment_enforcement",
    "[FW] Edge Local Access Restrictions": "analyse_edge_access",
    "[FW] Firewall Event Logging": "analyse_firewall_logging",
    "[VPN] Encryption Strength": "analyse_vpn_encryption_strength",
    "[VPN] Certificate Validation": "analyse_vpn_certificate_validation",
    "[VPN] Key Rotation": "analyse_vpn_key_rotation",
    "[Mgmt] Token Enumeration": "analyse_token_enumeration",
    "[Mgmt] Enterprise Authentication Mode": "analyse_enterprise_auth_mode",
    "[Mgmt] Dormant User Account Review": "analyse_dormant_users",
    "[Monitoring] Central Logging": "analyse_device_hardening",
    "[Monitoring] Admin Activity Logs": "analyse_enterprise_events",
    "[Monitoring] Event Review": "analyse_enterprise_events",
    "[Monitoring] Config Change Logs": "analyse_enterprise_events",
    "[Monitoring] Firewall Log Collection": "analyse_firewall_logging",
    "[System] Edge Versions": "analyse_edge_inventory",
    "[System] Patch Levels": "analyse_software_update",
    "[System] Config Consistency Real": "analyse_structural_diff",
    "[System] Edge Config Stack Review": "get_edge_configuration_modules",
    "[System] Edge Certificate Authentication": "analyse_edge_certificate_auth",
    "[System] NTP Time Synchronisation": "analyse_device_hardening",
    "[System] Edge Override Governance": "analyse_override_governance",
    "[System] Edge Admin AAA (TACACS/RADIUS)": "analyse_device_hardening",
    "[System] SNMP Hardening": "analyse_device_hardening",
    "[System] BFD Link Detection": "analyse_bfd",
    "[System] High Availability Configuration": "analyse_ha_config, analyse_ha_pair_parity",
    "[System] Inactive Edge Review": "analyse_edge_inventory",
    "[Isolation] Profile Inheritance": "analyse_segmentation",
    "[Isolation] Shared Resources": "analyse_shared_resources",
    "[Isolation] Wireless Security": "analyse_wireless_security",
    "[FW] Advanced Security Features (ATP/IDPS/URL)": "analyse_security_features",
}

_TITLE_INFERENCE = [
    ("any/any", "[FW] Rule Scope"),
    ("overly broad", "[FW] Rule Scope"),
    ("missing protocol", "[FW] Rule Scope"),
    ("default action", "[FW] Default Deny"),
    ("Stateful firewall", "[FW] Default Deny"),
    ("Firewall disabled", "[FW] Default Deny"),
    ("Segment firewall disabled on edge", "[FW] Segment Enforcement"),
    ("Sensitive VRF segment", "[FW] Segment Enforcement"),
    ("Edge segment firewall diverges", "[FW] Segment Enforcement"),
    ("permits cross-segment flow", "[FW] Segment Enforcement"),
    ("Inter-segment firewall allow", "[FW] Segment Enforcement"),
    ("NTP not", "[System] NTP Time Synchronisation"),
    ("Syslog forwarding", "[Monitoring] Central Logging"),
    ("SNMPv2c", "[System] SNMP Hardening"),
    ("SNMP enabled", "[System] SNMP Hardening"),
    ("TACACS", "[System] Edge Admin AAA (TACACS/RADIUS)"),
    ("admin authentication not configured", "[System] Edge Admin AAA (TACACS/RADIUS)"),
    ("pre-shared key authentication", "[System] Edge Certificate Authentication"),
    ("certificate authentication mode", "[System] Edge Certificate Authentication"),
    ("Edge certificate expiring", "[System] Edge Certificate Authentication"),
    ("Edge certificate expired", "[System] Edge Certificate Authentication"),
    ("No edge certificate found", "[System] Edge Certificate Authentication"),
    ("overlay encryption disabled", "[Net] Overlay Traffic Encryption"),
    ("software version below", "[System] Edge Versions"),
    ("end-of-support", "[System] Edge Versions"),
    ("Edge is OFFLINE", "[System] Inactive Edge Review"),
    ("not contacted", "[System] Inactive Edge Review"),
    ("activation state", "[System] Inactive Edge Review"),
    ("VPN edge-to-edge", "[Net] Edge-to-Edge Communication"),
    ("branch-to-branch VPN wider than profile", "[Net] Edge-to-Edge Communication"),
    ("Branch-to-branch VPN enabled without segment isolation", "[Net] Segment Isolation"),
    ("Edge override enables branch-to-branch VPN without isolation", "[Net] Segment Isolation"),
    ("allow list overlaps overlay-routable peer prefix", "[Net] Overlay to Management Access"),
    ("unrestricted with branch-to-branch VPN enabled", "[Net] Overlay to Management Access"),
    ("management service", "[FW] Edge Local Access Restrictions"),
    ("segment missing", "[Isolation] Profile Inheritance"),
    ("segment not defined", "[Isolation] Profile Inheritance"),
    ("differs from profile", "[System] Config Consistency Real"),
    ("present in profile but missing", "[System] Config Consistency Real"),
    ("Additional ", "[System] Config Consistency Real"),
    ("Enterprise event review:", "[Monitoring] Event Review"),
    ("Enterprise event:", "[Monitoring] Event Review"),
    ("USER_LOGIN_FAILURE", "[Monitoring] Admin Activity Logs"),
    ("CREATE_API_TOKEN", "[Monitoring] Admin Activity Logs"),
    ("REMOTE_ACTION", "[Monitoring] Admin Activity Logs"),
    ("MGD_CONF_APPLIED", "[Monitoring] Config Change Logs"),
    ("BFD not enabled", "[System] BFD Link Detection"),
    ("DNS servers not", "[Net] DNS Configuration"),
    ("Weak VPN crypto", "[VPN] Encryption Strength"),
    ("Edge-to-edge encryption disabled", "[VPN] Encryption Strength"),
    ("encryption protocol", "[VPN] Encryption Strength"),
    ("No edge certificates", "[VPN] Certificate Validation"),
    ("certificate expired", "[VPN] Certificate Validation"),
    ("certificate expiring", "[VPN] Certificate Validation"),
    ("certificate rotation", "[VPN] Key Rotation"),
    ("certificate/key material older", "[VPN] Key Rotation"),
    ("Rotation gap", "[VPN] Key Rotation"),
    ("current rotation period", "[VPN] Key Rotation"),
    ("API token inventory", "[Mgmt] Token Enumeration"),
    ("API token", "[Mgmt] Token Enumeration"),
    ("authentication mode", "[Mgmt] Enterprise Authentication Mode"),
    ("native VeloCloud", "[Mgmt] Enterprise Authentication Mode"),
    ("Enterprise user dormant", "[Mgmt] Dormant User Account Review"),
    ("never logged in", "[Mgmt] Dormant User Account Review"),
    ("Inactive enterprise user", "[Mgmt] Dormant User Account Review"),
    ("Wireless SSID", "[Isolation] Wireless Security"),
    ("Wireless enabled", "[Isolation] Wireless Security"),
    ("shared resources", "[Isolation] Shared Resources"),
    ("gateway pool", "[Isolation] Shared Resources"),
    ("Profile-defined segment missing", "[Isolation] Profile Inheritance"),
    ("Edge contains segment not defined", "[Isolation] Profile Inheritance"),
    ("VPN edge-to-datacentre setting differs", "[Isolation] Profile Inheritance"),
    ("more NAT rules than profile", "[Isolation] Profile Inheritance"),
    ("HA edge state", "[System] High Availability Configuration"),
    ("HA peer last contact", "[System] High Availability Configuration"),
    ("HA pair config mismatch", "[System] High Availability Configuration"),
    ("standby serial metadata", "[System] High Availability Configuration"),
    ("override active", "[System] Edge Override Governance"),
    ("Software update", "[System] Patch Levels"),
    ("NAT rule", "[FW] NAT Exposure"),
    ("Firewall active but logging", "[FW] Firewall Event Logging"),
    ("Firewall logging enabled but syslog", "[Monitoring] Firewall Log Collection"),
    ("Business policy allow-all", "[Net] Default Segment Behaviour"),
    ("QoS/business policy override", "[Net] Business Policy Override"),
]


def _normalize_automation_tier(value, title: str = "") -> str:
    """Normalise XLSX Automation cell to automated|partial|manual."""
    if value is not None and str(value).strip():
        v = str(value).strip().lower()
        if v in {"automated", "partial", "manual"}:
            return v
    if title:
        return _AUTOMATION_TIER_BY_TITLE.get(title, "manual")
    return "manual"


def _automation_tier_for_check(row: tuple, idx: dict, title: str) -> str:
    if "Automation" in idx and idx["Automation"] < len(row):
        cell = row[idx["Automation"]]
        if cell is not None and str(cell).strip():
            return _normalize_automation_tier(cell, title)
    return _AUTOMATION_TIER_BY_TITLE.get(title, "manual")


def _script_module_for_check(row: tuple, idx: dict, title: str) -> str:
    if "ScriptModule" in idx and idx["ScriptModule"] < len(row):
        cell = row[idx["ScriptModule"]]
        if cell is not None and str(cell).strip():
            return str(cell).strip()
    return _SCRIPT_MODULE_BY_TITLE.get(title, "")


def _methodology_builtin_drift(checks: list) -> list:
    """Return human-readable drift messages when XLSX rows differ from built-in maps."""
    drift = []
    by_title = {c["title"]: c for c in checks}
    for title, tier in _AUTOMATION_TIER_BY_TITLE.items():
        row = by_title.get(title)
        if not row:
            drift.append(f"built-in only (missing from XLSX): {title}")
            continue
        xlsx_tier = row.get("automationTier", "")
        if xlsx_tier and xlsx_tier != tier:
            drift.append(f"tier mismatch for {title}: XLSX={xlsx_tier} builtin={tier}")
        xlsx_mod = (row.get("scriptModule") or "").strip()
        builtin_mod = _SCRIPT_MODULE_BY_TITLE.get(title, "")
        if xlsx_mod and xlsx_mod != builtin_mod:
            drift.append(f"module mismatch for {title}: XLSX={xlsx_mod} builtin={builtin_mod}")
    for title in sorted(set(by_title) - set(_AUTOMATION_TIER_BY_TITLE)):
        drift.append(f"XLSX only (missing from built-in): {title}")
    return drift


def load_methodology_checks(xlsx_path: str = None) -> list:
    """Load methodology rows from XLSX (Automation + ScriptModule); fall back to built-in map."""
    path = xlsx_path or METHODOLOGY_XLSX
    checks = []
    if os.path.isfile(path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {h: i for i, h in enumerate(headers) if h}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[idx.get("Title", 4)]:
                    continue
                title = str(row[idx["Title"]])
                checks.append({
                    "scope":          row[idx.get("Scope", 0)],
                    "title":          title,
                    "description":    row[idx.get("Description", 5)] or "",
                    "tooling":        row[idx.get("Tooling", 6)] or "",
                    "priority":       row[idx.get("Priority", 8)] or "",
                    "automationTier": _automation_tier_for_check(row, idx, title),
                    "scriptModule":   _script_module_for_check(row, idx, title),
                })
            wb.close()
            if checks:
                drift = _methodology_builtin_drift(checks)
                if drift:
                    print(
                        "[WARN] Methodology XLSX differs from built-in maps in velo_final.py "
                        f"({len(drift)} item(s)). XLSX is used for coverage; update built-in fallback "
                        "if you rely on openpyxl-less runs:"
                    )
                    for msg in drift[:8]:
                        print(f"       - {msg}")
                    if len(drift) > 8:
                        print(f"       - ... +{len(drift) - 8} more")
                return checks
        except ImportError:
            print("[WARN] openpyxl not installed — using built-in methodology titles")
        except Exception as exc:
            print(f"[WARN] Could not load methodology XLSX: {exc}")
    for title, tier in _AUTOMATION_TIER_BY_TITLE.items():
        checks.append({
            "scope":          "Unknown",
            "title":          title,
            "description":    "",
            "tooling":        "",
            "priority":       "Mandatory",
            "automationTier": tier,
            "scriptModule":   _SCRIPT_MODULE_BY_TITLE.get(title, ""),
        })
    return checks


def methodology_tier_lookup(checks: list) -> dict:
    return {c["title"]: c.get("automationTier", "manual") for c in checks}


def enrich_findings_methodology(findings: list, tier_lookup: dict = None) -> None:
    """Backfill methodologyRef and automation on findings that lack them."""
    lookup = tier_lookup or _AUTOMATION_TIER_BY_TITLE
    for f in findings:
        if not f.get("methodologyRef"):
            title = f.get("title", "")
            for needle, ref in _TITLE_INFERENCE:
                if needle.lower() in title.lower():
                    f["methodologyRef"] = ref
                    break
        ref = f.get("methodologyRef", "")
        if ref in lookup:
            f["automation"] = lookup[ref]
        elif ref and not f.get("automation"):
            f["automation"] = _AUTOMATION_TIER_BY_TITLE.get(ref, "automated")


def build_methodology_coverage(
    methodology_checks: list,
    edge_findings: list,
    event_findings: list,
    events_enabled: bool,
    deep_ran: bool,
) -> list:
    """Build 48-row methodology coverage matrix."""
    all_findings = edge_findings + event_findings
    by_ref = {}
    for f in all_findings:
        ref = f.get("methodologyRef") or ""
        if ref:
            by_ref.setdefault(ref, []).append(f)

    coverage = []
    for check in methodology_checks:
        title = check["title"]
        tier = check.get("automationTier", "manual")
        matched = by_ref.get(title, [])
        fail_findings = [
            f for f in matched
            if f.get("severity") in {"high", "medium"}
        ]

        if tier == "manual":
            status = "manual"
        elif tier == "partial":
            if not matched:
                if title in {"[Monitoring] Event Review", "[Monitoring] Config Change Logs",
                             "[Monitoring] Admin Activity Logs"} and not events_enabled:
                    status = "not_run"
                elif title in {"[Net] Route Table Review", "[Net] Gateway Assignment Review"} and not deep_ran:
                    status = "not_run"
                else:
                    status = "partial"
            else:
                # Partial-tier checks are script-assisted; findings need human confirmation
                status = "assisted"
        elif fail_findings:
            status = "fail"
        elif matched:
            sevs = {f.get("severity") for f in matched}
            status = "partial" if sevs <= {"low", "info"} else "pass"
        else:
            status = "pass"

        edges_affected = sorted({
            f.get("edgeName") for f in matched
            if f.get("edgeName") and f.get("edgeName") != "(enterprise)"
        })

        coverage.append({
            "scope":          check.get("scope"),
            "title":          title,
            "priority":       check.get("priority"),
            "automationTier": tier,
            "scriptModule":   check.get("scriptModule", ""),
            "status":         status,
            "findingCount":   len(matched),
            "edgesAffected":  edges_affected,
        })
    return coverage


def build_manual_review_checklist(
    methodology_checks: list,
    edge_findings: list,
    event_findings: list,
) -> list:
    all_findings = edge_findings + event_findings
    by_ref = {}
    for f in all_findings:
        ref = f.get("methodologyRef") or ""
        if ref:
            by_ref.setdefault(ref, []).append(f)

    checklist = []
    for check in methodology_checks:
        tier = check.get("automationTier", "manual")
        if tier not in {"manual", "partial"}:
            continue
        related = by_ref.get(check["title"], [])
        status = "pending"
        if related:
            status = "assisted"
        checklist.append({
            "scope":            check.get("scope"),
            "title":              check["title"],
            "priority":           check.get("priority"),
            "tooling":            check.get("tooling"),
            "description":        check.get("description"),
            "automationTier":     tier,
            "scriptModule":       check.get("scriptModule", ""),
            "status":             status,
            "evidenceTemplate":   f"Record evidence for: {check['title']}",
            "relatedFindingCount": len(related),
        })
    return checklist


def build_review_summary_by_scope(coverage: list) -> dict:
    summary = {}
    for row in coverage:
        scope = row.get("scope") or "Unknown"
        bucket = summary.setdefault(scope, {
            "total": 0, "mandatory": 0, "automated_pass": 0,
            "automated_fail": 0, "assisted": 0, "manual": 0, "partial": 0, "not_run": 0,
        })
        bucket["total"] += 1
        if row.get("priority") == "Mandatory":
            bucket["mandatory"] += 1
        status = row.get("status", "")
        tier = row.get("automationTier", "")
        if status == "manual":
            bucket["manual"] += 1
        elif status == "not_run":
            bucket["not_run"] += 1
        elif status == "assisted":
            bucket["assisted"] += 1
        elif status == "partial":
            bucket["partial"] += 1
        elif status == "fail":
            bucket["automated_fail"] += 1
        elif status == "pass":
            bucket["automated_pass"] += 1
    return summary


_COVERAGE_KEYS = [
    "scope", "title", "priority", "automationTier", "scriptModule",
    "status", "findingCount", "edgesAffected",
]

_DRADIS_KEYS = ["Scope", "Title", "Priority", "Status", "Executor", "Comments", "Evidence"]


def build_dradis_import(coverage: list, methodology_checks: list) -> list:
    desc_by_title = {c["title"]: c.get("tooling", "") for c in methodology_checks}
    rows = []
    for row in coverage:
        title = row["title"]
        status = row["status"]
        if status == "manual":
            comments = "Manual review required — see tooling column in methodology"
            evidence = desc_by_title.get(title, "")
        elif status == "pass":
            comments = "No findings — check passed (automated)"
            evidence = ""
        elif status == "not_run":
            comments = "Automated check not run (enable --events or --deep)"
            evidence = ""
        elif status == "assisted":
            comments = (
                f"Script-assisted — {row['findingCount']} finding(s); "
                "confirm whether drift/overrides are intentional"
            )
            evidence = ", ".join(row.get("edgesAffected") or [])[:500]
        elif status == "fail":
            comments = f"{row['findingCount']} finding(s) on {len(row.get('edgesAffected') or [])} edge(s)"
            evidence = ", ".join(row.get("edgesAffected") or [])[:500]
        else:
            comments = f"Partial coverage — {row['findingCount']} informational/low finding(s)"
            evidence = ", ".join(row.get("edgesAffected") or [])[:500]
        rows.append({
            "Scope":    row.get("scope"),
            "Title":    title,
            "Priority": row.get("priority"),
            "Status":   status,
            "Executor": "velo_final.py",
            "Comments": comments,
            "Evidence": evidence,
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="VeloCloud profile-aware config collector + security review",
    )
    parser.add_argument(
        "--events",
        dest="events",
        action="store_true",
        default=EVENTS_DEFAULT,
        help="Fetch enterprise events (admin/config monitoring) via Portal API (default: on)",
    )
    parser.add_argument(
        "--no-events",
        dest="events",
        action="store_false",
        help="Skip enterprise event fetch and event-based monitoring checks",
    )
    parser.add_argument(
        "--events-hours",
        type=int,
        default=EVENTS_HOURS_DEFAULT,
        metavar="N",
        help=f"Event lookback window in hours (default: {EVENTS_HOURS_DEFAULT})",
    )
    parser.add_argument(
        "--deep",
        dest="deep",
        action="store_true",
        default=PHASE4_DEFAULT,
        help="Route table + gateway assignment collectors (default: on)",
    )
    parser.add_argument(
        "--no-deep",
        dest="deep",
        action="store_false",
        help="Skip route table and gateway assignment collectors",
    )
    parser.add_argument(
        "--no-coverage-report",
        action="store_true",
        help="Skip methodology coverage report generation",
    )
    parser.add_argument(
        "--scope-edges",
        default=None,
        metavar="NAMES",
        help="Comma-separated edge names to limit scope (overrides VCO_SCOPE_EDGES)",
    )
    parser.add_argument(
        "--methodology-xlsx",
        "--xlsx",
        dest="methodology_xlsx",
        default=None,
        metavar="PATH",
        help="Methodology workbook path (overrides VCO_METHODOLOGY_XLSX)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        metavar="PATH",
        help="Write all output to this directory (default: vco_output/<UTC timestamp>)",
    )
    parser.add_argument(
        "--min-edge-version",
        default=None,
        metavar="VER",
        help="Minimum acceptable edge software version, e.g. 6.0.0 (overrides VCO_MIN_EDGE_VERSION)",
    )
    parser.add_argument(
        "--verify-tls",
        action="store_true",
        default=None,
        help="Verify TLS certificates (overrides VCO_VERIFY_TLS)",
    )
    parser.add_argument(
        "--no-verify-tls",
        dest="verify_tls",
        action="store_false",
        help="Disable TLS certificate verification",
    )
    parser.add_argument(
        "--dormant-user-days",
        type=int,
        default=None,
        metavar="N",
        help=f"Dormant user threshold in days (default: {DORMANT_USER_DAYS})",
    )
    parser.add_argument(
        "--stale-edge-days",
        type=int,
        default=None,
        metavar="N",
        help=f"Stale edge threshold in days (default: {STALE_EDGE_DAYS})",
    )
    return parser.parse_args()


def apply_cli_overrides(args) -> None:
    """Apply CLI flags over environment defaults for this run."""
    global OUTPUT_DIR, VCO_SCOPE_EDGES, METHODOLOGY_XLSX
    global MIN_EDGE_VERSION, VERIFY_TLS, DORMANT_USER_DAYS, STALE_EDGE_DAYS

    OUTPUT_DIR = _resolve_output_dir(args.output_dir)
    if args.scope_edges is not None:
        VCO_SCOPE_EDGES = [x.strip() for x in args.scope_edges.split(",") if x.strip()]
    if args.methodology_xlsx:
        METHODOLOGY_XLSX = args.methodology_xlsx
    if args.min_edge_version is not None:
        MIN_EDGE_VERSION = args.min_edge_version.strip()
    if args.verify_tls is not None:
        VERIFY_TLS = args.verify_tls
        api2_client.verify = VERIFY_TLS
        portal_client.verify = VERIFY_TLS
        if not VERIFY_TLS:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    if args.dormant_user_days is not None:
        DORMANT_USER_DAYS = args.dormant_user_days
    if args.stale_edge_days is not None:
        STALE_EDGE_DAYS = args.stale_edge_days


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    if not any(a in sys.argv for a in ("-h", "--help")):
        _require_env()
    args = parse_args()
    apply_cli_overrides(args)
    print("=" * 72)
    print("  velo_final.py — VeloCloud profile-aware collector + security review")
    print("=" * 72)
    print(f"  Host:       {VCO_HOSTNAME}")
    print(f"  Enterprise: {VCO_ENTERPRISE_LOGICAL_ID} (numeric: {VCO_ENTERPRISE_NUMERIC_ID})")
    if VCO_SCOPE_EDGES:
        print(f"  Scope:      {len(VCO_SCOPE_EDGES)} edge(s)")
    if args.events:
        print(f"  Events:     last {args.events_hours}h")
    else:
        print("  Events:     disabled (--no-events)")
    if args.deep:
        print("  Deep:       route table + gateway assignments")
    else:
        print("  Deep:       disabled (--no-deep)")
    if MIN_EDGE_VERSION:
        print(f"  Min version: {MIN_EDGE_VERSION}")
    print(f"  Output:     {OUTPUT_DIR}")

    # ── 1. Portal edge list ──────────────────────────────────────────────────
    print("\n[1/6] Fetching Portal edge list...")
    edges = get_portal_edge_list()
    if not edges:
        print("[ERROR] No edges returned from Portal API. Check VCO_TOKEN and numeric enterprise ID.")
        sys.exit(1)
    save_json(os.path.join(OUTPUT_DIR, "portal_edge_list.json"), edges)
    print(f"       {len(edges)} edge record(s) retrieved")

    # ── 2. Profile catalog ───────────────────────────────────────────────────
    print("\n[2/6] Fetching Portal enterprise services / profile catalog...")
    catalog = get_profile_catalog()
    if not catalog:
        print("[ERROR] No profile records extracted. Check VCO_TOKEN and enterprise ID.")
        sys.exit(1)
    save_json(os.path.join(OUTPUT_DIR, "portal_profile_catalog.json"), catalog)
    print(f"       {len(catalog)} unique profile record(s) extracted")

    # ── 3. Build edge/profile join ───────────────────────────────────────────
    print("\n[3/6] Building edge/profile join...")
    joined = build_edge_profile_join(edges, catalog)
    save_json(os.path.join(OUTPUT_DIR, "edge_profile_join.json"), joined)
    print(f"       {len(joined)} edge(s) in scope")

    # ── 4. Fetch APIv2 configs + Portal configuration modules ──────────────
    print("\n[4/6] Fetching profile + edge configs (APIv2 + Portal modules)...")
    profile_cache         = {}
    profile_modules_cache = {}
    profiles_dir          = os.path.join(OUTPUT_DIR, "profiles")
    profile_firewall_dir  = os.path.join(OUTPUT_DIR, "profile_firewall")
    edges_dir             = os.path.join(OUTPUT_DIR, "edges")
    firewall_dir          = os.path.join(OUTPUT_DIR, "firewall")
    wan_dir               = os.path.join(OUTPUT_DIR, "wan")
    certs_dir             = os.path.join(OUTPUT_DIR, "certs")
    combined_dir          = os.path.join(OUTPUT_DIR, "combined")
    combined_records      = []
    enterprise_services = get_enterprise_services_payload()
    if enterprise_services is not None and enterprise_services != {}:
        save_json(os.path.join(OUTPUT_DIR, "enterprise_services.json"), enterprise_services)

    for item in joined:
        en   = item["edgeName"]
        elid = item["edgeLogicalId"]
        enid = item.get("edgeNumericId")
        plid = item["profileLogicalId"]
        pnid = item.get("profileNumericId")

        print(f"  → {en}")

        if plid and plid not in profile_cache:
            pcfg = get_profile_device_settings(plid)
            profile_cache[plid] = pcfg
            if pcfg:
                save_json(os.path.join(profiles_dir, f"{plid}.json"), pcfg)

        if pnid and pnid not in profile_modules_cache:
            raw_modules = get_profile_configuration_modules(pnid)
            mod_dict = _portal_modules_dict(raw_modules)
            profile_modules_cache[pnid] = mod_dict
            pfw = _module_data(mod_dict, "firewall")
            if pfw:
                save_json(
                    os.path.join(profile_firewall_dir, f"profile_{pnid}_firewall.json"),
                    pfw,
                )

        pcfg = profile_cache.get(plid) or {}
        profile_mods = profile_modules_cache.get(pnid) or {}
        profile_fw = _module_data(profile_mods, "firewall")
        ecfg = get_edge_device_settings(elid) or {}
        if ecfg:
            save_json(os.path.join(edges_dir, f"{safe_name(en)}.json"), ecfg)

        fw_cfg = {}
        wan_cfg = {}
        cp_cfg = {}
        ds_cfg = {}
        qos_cfg = {}
        edge_certs = []
        modules = {}
        if enid:
            modules = get_edge_configuration_modules(enid) or {}
            fw_cfg = _module_data(modules, "firewall")
            wan_cfg = _module_data(modules, "WAN")
            cp_cfg = _module_data(modules, "controlPlane")
            ds_cfg = _module_data(modules, "deviceSettings")
            qos_cfg = _module_data(modules, "QOS")
            edge_certs = get_edge_certificates(enid) or []
            if fw_cfg:
                save_json(os.path.join(firewall_dir, f"{safe_name(en)}.json"), fw_cfg)
            if wan_cfg:
                save_json(os.path.join(wan_dir, f"{safe_name(en)}.json"), wan_cfg)
            if edge_certs:
                save_json(
                    os.path.join(certs_dir, f"{safe_name(en)}.json"),
                    _sanitize_cert_records(edge_certs),
                )

        rec = deepcopy(item)
        rec["profileConfig"]       = pcfg
        rec["profileFirewall"]     = profile_fw
        rec["edgeConfig"]          = ecfg
        rec["firewallConfig"]      = fw_cfg
        rec["wanConfig"]           = wan_cfg
        rec["controlPlaneConfig"]  = cp_cfg
        rec["deviceSettingsModule"] = ds_cfg
        rec["qosConfig"]           = qos_cfg
        rec["edgeCertificates"]    = _sanitize_cert_records(edge_certs)
        rec["configurationModules"] = modules if enid else {}
        save_json(os.path.join(combined_dir, f"{safe_name(en)}_combined.json"), rec)
        combined_records.append(rec)

    # ── 5. Detection ─────────────────────────────────────────────────────────
    print("\n[5/6] Running detection modules...")
    all_findings  = []
    all_hardening = []
    enterprise_events = []
    enterprise_event_findings = []
    scope_edge_names = {item["edgeName"] for item in joined}
    overlay_peer_prefixes = build_overlay_peer_prefixes(combined_records)
    enterprise_record = get_enterprise_record()
    if enterprise_record:
        save_json(os.path.join(OUTPUT_DIR, "mgmt", "enterprise.json"), enterprise_record)

    for rec in combined_records:
        ctx  = {k: rec[k] for k in
                ["edgeName", "edgeLogicalId", "profileName",
                 "profileNumericId", "profileLogicalId",
                 "edgeOverrides", "edgePortalRecord"]}
        pcfg = rec.get("profileConfig") or {}
        ecfg = rec.get("edgeConfig") or {}
        fw_cfg = rec.get("firewallConfig") or {}
        wan_cfg = rec.get("wanConfig") or {}
        cp_cfg = rec.get("controlPlaneConfig") or {}
        ds_mod = rec.get("deviceSettingsModule") or {}
        edge_certs = rec.get("edgeCertificates") or []
        effective = _effective_config(pcfg, ecfg)

        analyse_firewall_rules(
            ctx, pcfg, ecfg, all_findings,
            firewall_cfg=fw_cfg,
            profile_fw=rec.get("profileFirewall") or {},
            edge_ds=rec.get("deviceSettingsModule") or {},
            profile_modules=profile_modules_cache.get(ctx.get("profileNumericId")) or {},
        )
        analyse_segment_enforcement(
            ctx, pcfg, ecfg, all_findings,
            firewall_cfg=fw_cfg,
            profile_fw=rec.get("profileFirewall") or {},
            edge_ds=rec.get("deviceSettingsModule") or {},
            profile_ds=_module_data(
                profile_modules_cache.get(ctx.get("profileNumericId")) or {},
                "deviceSettings",
            ),
            edge_qos=rec.get("qosConfig") or {},
            profile_modules=profile_modules_cache.get(ctx.get("profileNumericId")) or {},
            profile_fw_module=rec.get("profileFirewall") or {},
        )
        analyse_edge_access(ctx, fw_cfg, all_findings)
        analyse_overlay_management_access(
            ctx, fw_cfg, pcfg, ecfg,
            overlay_peer_prefixes.get(ctx["edgeName"], []),
            all_findings,
        )
        analyse_firewall_logging(ctx, fw_cfg, effective, all_findings)
        analyse_override_governance(ctx, all_findings)
        analyse_software_update(ctx, pcfg, ecfg, all_findings)
        analyse_nat_exposure(ctx, pcfg, ecfg, all_findings)
        analyse_business_policy(ctx, wan_cfg, pcfg, ecfg, all_findings)
        analyse_segmentation(ctx, pcfg, ecfg, all_findings)
        analyse_edge_to_edge_communication(
            ctx, pcfg, ecfg, all_findings,
            profile_modules=profile_modules_cache.get(ctx.get("profileNumericId")) or {},
            profile_fw=rec.get("profileFirewall") or {},
            edge_fw=fw_cfg,
        )
        analyse_segment_isolation(
            ctx, pcfg, ecfg, all_findings,
            profile_modules=profile_modules_cache.get(ctx.get("profileNumericId")) or {},
            profile_fw=rec.get("profileFirewall") or {},
            edge_fw=fw_cfg,
        )
        analyse_structural_diff(ctx, pcfg, ecfg, all_findings)
        analyse_device_hardening(ctx, pcfg, ecfg, all_findings, firewall_cfg=fw_cfg)
        analyse_bfd(ctx, pcfg, ecfg, all_findings)
        analyse_dns(ctx, pcfg, ecfg, all_findings)
        analyse_ha_config(ctx, all_findings)
        analyse_security_features(ctx, fw_cfg, all_findings)
        analyse_edge_inventory(ctx, all_findings)
        analyse_edge_certificate_auth(ctx, edge_certs, enterprise_record, all_findings)
        analyse_vpn_encryption_strength(ctx, cp_cfg, ds_mod, all_findings)
        analyse_vpn_certificate_validation(ctx, edge_certs, all_findings)
        analyse_vpn_key_rotation(ctx, edge_certs, all_findings)
        analyse_wireless_security(
            ctx, pcfg, ecfg,
            rec.get("configurationModules") or {},
            profile_modules_cache.get(ctx.get("profileNumericId")) or {},
            all_findings,
        )
        all_hardening.extend(evaluate_hardening(ctx, pcfg, ecfg))

    analyse_ha_pair_parity(combined_records, all_findings)

    print("       Running enterprise management collectors (users, tokens, auth)...")
    mgmt_dir = os.path.join(OUTPUT_DIR, "mgmt")
    if not enterprise_record:
        enterprise_record = get_enterprise_record()
    enterprise_users = get_enterprise_users()
    api_tokens = get_enterprise_api_tokens()
    if enterprise_record:
        if not os.path.exists(os.path.join(mgmt_dir, "enterprise.json")):
            save_json(os.path.join(mgmt_dir, "enterprise.json"), enterprise_record)
    if enterprise_users:
        save_json(
            os.path.join(mgmt_dir, "enterprise_users.json"),
            [_sanitize_enterprise_user(u) for u in enterprise_users],
        )
    if api_tokens:
        save_json(
            os.path.join(mgmt_dir, "api_tokens.json"),
            [_sanitize_api_token_record(t) for t in api_tokens],
        )
    mgmt_ctx = _enterprise_ctx()
    analyse_token_enumeration(mgmt_ctx, api_tokens, all_findings)
    analyse_enterprise_auth_mode(mgmt_ctx, enterprise_record, enterprise_users, all_findings)
    analyse_dormant_users(mgmt_ctx, enterprise_users, all_findings)

    print("       Running isolation collectors (shared resources)...")
    gateways = _extract_gateways_from_services(enterprise_services) if enterprise_services else []
    assignments_by_gateway = {}
    for gw in gateways:
        gw_id = gw.get("id")
        if gw_id is None:
            continue
        assignments_by_gateway[gw_id] = get_gateway_edge_assignments(gw_id) or []
    analyse_shared_resources(
        mgmt_ctx, enterprise_record, gateways, assignments_by_gateway, scope_edge_names, all_findings,
    )

    if args.deep:
        print("       Running Phase 4 deep collectors (route table, gateways)...")
        routes_dir = os.path.join(OUTPUT_DIR, "routes")
        gateways_dir = os.path.join(OUTPUT_DIR, "gateways")
        route_data = get_enterprise_route_table()
        if route_data:
            save_json(os.path.join(routes_dir, "enterprise_route_table.json"), route_data)
            analyse_route_table(route_data, all_findings)
        gateways = _extract_gateways_from_services(enterprise_services)
        if not assignments_by_gateway:
            for gw in gateways:
                gw_id = gw.get("id")
                if gw_id is None:
                    continue
                assignments_by_gateway[gw_id] = get_gateway_edge_assignments(gw_id) or []
        for gw in gateways:
            gw_id = gw.get("id")
            if gw_id is None:
                continue
            assigned = assignments_by_gateway.get(gw_id)
            if assigned is None:
                assigned = get_gateway_edge_assignments(gw_id) or []
                assignments_by_gateway[gw_id] = assigned
            save_json(
                os.path.join(gateways_dir, f"gateway_{gw_id}_assignments.json"),
                assigned,
            )
        analyse_gateway_assignments(
            gateways, assignments_by_gateway, scope_edge_names, all_findings,
        )

    if args.events:
        print(f"\n[6/6] Fetching enterprise events ({args.events_hours}h)...")
        enterprise_events = get_enterprise_events(args.events_hours)
        events_path = os.path.join(OUTPUT_DIR, "enterprise_events.json")
        save_json(events_path, enterprise_events)
        print(f"       {len(enterprise_events)} event(s) retrieved")
        enterprise_event_findings = analyse_enterprise_events(
            enterprise_events, args.events_hours,
        )
        event_findings_path = os.path.join(OUTPUT_DIR, "enterprise_event_findings.json")
        save_json(event_findings_path, enterprise_event_findings)
        print(f"       {len(enterprise_event_findings)} event finding(s) after filtering")

    methodology_checks = []
    tier_lookup = _AUTOMATION_TIER_BY_TITLE
    if not args.no_coverage_report:
        methodology_checks = load_methodology_checks(METHODOLOGY_XLSX)
        tier_lookup = methodology_tier_lookup(methodology_checks)

    enrich_findings_methodology(all_findings, tier_lookup)
    enrich_findings_methodology(enterprise_event_findings, tier_lookup)

    # Deduplicate findings
    seen = set()
    deduped = []
    for f in all_findings:
        key = (
            f["category"], f["edgeLogicalId"], f.get("methodologyRef", ""),
            f["title"], f["fieldPath"], f["evidence"],
        )
        if key not in seen:
            seen.add(key)
            deduped.append(f)

    findings_dir = os.path.join(OUTPUT_DIR, "findings")
    save_json(os.path.join(findings_dir, "findings.json"), deduped)
    save_csv(os.path.join(findings_dir, "findings.csv"), deduped, _FINDING_KEYS)
    save_json(os.path.join(findings_dir, "hardening.json"), all_hardening)

    coverage = []
    manual_checklist = []
    scope_summary = {}
    if not args.no_coverage_report and methodology_checks:
        coverage = build_methodology_coverage(
            methodology_checks,
            deduped,
            enterprise_event_findings,
            args.events,
            args.deep,
        )
        manual_checklist = build_manual_review_checklist(
            methodology_checks, deduped, enterprise_event_findings,
        )
        scope_summary = build_review_summary_by_scope(coverage)
        save_json(os.path.join(findings_dir, "methodology_coverage.json"), coverage)
        coverage_csv = [
            {**row, "edgesAffected": ",".join(row.get("edgesAffected") or [])}
            for row in coverage
        ]
        save_csv(os.path.join(findings_dir, "methodology_coverage.csv"), coverage_csv, _COVERAGE_KEYS)
        save_json(os.path.join(findings_dir, "manual_review_checklist.json"), manual_checklist)
        save_json(os.path.join(findings_dir, "review_summary_by_scope.json"), scope_summary)
        dradis_rows = build_dradis_import(coverage, methodology_checks)
        save_csv(os.path.join(findings_dir, "dradis_import.csv"), dradis_rows, _DRADIS_KEYS)

    sev_counts = {}
    for f in deduped:
        sev_counts[f["severity"]] = sev_counts.get(f["severity"], 0) + 1

    cat_counts = {}
    for f in deduped:
        cat_counts[f["category"]] = cat_counts.get(f["category"], 0) + 1

    auto_pass = sum(1 for r in coverage if r.get("status") == "pass")
    auto_fail = sum(1 for r in coverage if r.get("status") == "fail")
    automated_total = sum(
        1 for r in coverage if r.get("automationTier") == "automated"
    )
    methodology_pct = round(100 * auto_pass / max(automated_total, 1), 1)

    summary = {
        "edges_in_scope":              len(joined),
        "profiles_fetched":            len(profile_cache),
        "findings_total":              len(deduped),
        "severity_counts":             sev_counts,
        "category_counts":             cat_counts,
        "hardening_notes":             len(all_hardening),
        "events_enabled":              args.events,
        "events_fetched":              len(enterprise_events),
        "enterprise_event_findings":   len(enterprise_event_findings),
        "deep_enabled":                args.deep,
        "methodology_checks":          len(methodology_checks),
        "methodology_automated_pass":  auto_pass,
        "methodology_automated_fail":  auto_fail,
        "methodology_automated_pct":   methodology_pct,
        "manual_review_pending":       sum(1 for c in manual_checklist if c.get("status") == "pending"),
        "output_dir":                  OUTPUT_DIR,
    }
    save_json(os.path.join(findings_dir, "summary.json"), summary)

    print("\n✅  DONE")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
